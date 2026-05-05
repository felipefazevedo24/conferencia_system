from __future__ import annotations

import io
import json
import math
import os
import re
import time
import unicodedata
from datetime import datetime
from pathlib import Path
from urllib.parse import quote

from flask import current_app
import requests
from sqlalchemy import or_

from ..extensions import db
from ..models import (
    AgendamentoCliente,
    AgendamentoFornecedor,
    AgendamentoMotorista,
    AgendamentoVeiculo,
    ItemNota,
    Usuario,
)
from .consyste_service import listar_nfes_consyste_por_caixa
from .pedidos_service import (
    PEDIDOS_FONTE_GOOGLE_SHEETS,
    buscar_linhas_pedido,
    label_fonte_pedidos,
    obter_fonte_pedidos_google_sheets,
)

try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover - dependencia opcional em alguns ambientes
    load_workbook = None


TIPOS_SOLICITACAO = ("COLETA", "ENTREGA", "AVULSA")
STATUS_SOLICITACAO = ("Pendente", "EmAnalise", "Alocada", "EmRota", "Concluida", "Cancelada")
PRIORIDADES_SOLICITACAO = ("Baixa", "Media", "Alta", "Critica")
STATUS_ATIVOS = {"Pendente", "EmAnalise", "Alocada", "EmRota"}
VEICULOS_KANBAN = ("IVECO", "SAVEIRO")
_GEOCODE_LAST_REQUEST_TS = 0.0

_CADASTRO_MODELS = {
    "fornecedor": AgendamentoFornecedor,
    "cliente": AgendamentoCliente,
}

_HEADER_ALIASES = {
    "fornecedor": {
        "codigo": {"codigo"},
        "nome": {"nome"},
        "razao_social": {"razaosocial", "rsocial"},
        "cnpj_cpf": {"cnpjcpf"},
        "tipo_pessoa": {"tipodepessoa", "pessoa"},
        "contato": {"contato"},
        "telefone": {"telefone1"},
        "telefone_secundario": {"telefone2"},
        "email": {"email"},
        "logradouro": {"endereco"},
        "bairro": {"bairro"},
        "cidade": {"cidade"},
        "uf": {"estado", "uf"},
        "cep": {"cep"},
        "observacoes": {"geral", "observacoes"},
        "janela_atendimento": {"janelaatendimento"},
        "transportadora": {"transportadora"},
        "codigo_integracao": {"codintegracao"},
        "inativo": {"inativo"},
    },
    "cliente": {
        "codigo": {"codigo"},
        "nome": {"nome"},
        "razao_social": {"rsocial", "razaosocial"},
        "cnpj_cpf": {"cnpjcpf"},
        "tipo_pessoa": {"pessoa"},
        "contato": {"contatoprincipal", "contato"},
        "telefone": {"telefone1"},
        "telefone_secundario": {"telefone2"},
        "email": {"email", "emaildeentrega"},
        "logradouro": {"endereco"},
        "bairro": {"bairro"},
        "cidade": {"cidade"},
        "uf": {"uf", "estado"},
        "cep": {"cep"},
        "observacoes": {"anotacoesdefaturamentodocliente", "observacoes"},
        "municipio_entrega": {"municipiodeentrega"},
        "codigo_integracao": {"codigodeintegracao", "codintegracao"},
        "inativo": {"inativo"},
    },
}

_CONSYSTE_FIELDS = (
    "id,chave,numero,dest_nome,dest_cnpj,dest_endereco,dest_numero,dest_complemento,"
    "dest_bairro,dest_cidade,dest_uf,dest_cep,dest_telefone,dest_email,contato,"
    "cliente_nome,cliente_cnpj,endereco,numero_endereco,bairro,cidade,uf,cep,"
    "observacao,informacoes_complementares,itens"
)


def status_label_agendamento(status: str | None) -> str:
    mapping = {
        "Pendente": "Pendente",
        "EmAnalise": "Em análise",
        "Alocada": "Alocada",
        "EmRota": "Em rota",
        "Concluida": "Concluída",
        "Cancelada": "Cancelada",
    }
    return mapping.get(str(status or "").strip(), str(status or "").strip() or "---")


def prioridade_label_agendamento(prioridade: str | None) -> str:
    mapping = {
        "Baixa": "Baixa",
        "Media": "Média",
        "Alta": "Alta",
        "Critica": "Crítica",
    }
    return mapping.get(str(prioridade or "").strip(), str(prioridade or "").strip() or "---")


def normalizar_texto(valor) -> str:
    texto = str(valor or "").strip().lower()
    texto = unicodedata.normalize("NFKD", texto)
    texto = "".join(ch for ch in texto if not unicodedata.combining(ch))
    return re.sub(r"[^a-z0-9]+", "", texto)


def limpar_documento(valor) -> str:
    return re.sub(r"\D", "", str(valor or ""))


def valor_texto(valor) -> str:
    if valor is None:
        return ""
    if isinstance(valor, float):
        if valor.is_integer():
            return str(int(valor))
        return str(valor).strip()
    return str(valor).strip()


def bool_from_excel(valor) -> bool:
    texto = normalizar_texto(valor)
    return texto in {"sim", "s", "true", "1", "ativo", "nao0"}


def formatar_endereco_logistico(dados: dict | None) -> str:
    dados = dados or {}
    primeira_linha = ", ".join(
        parte
        for parte in [
            str(dados.get("logradouro") or "").strip(),
            str(dados.get("numero") or "").strip(),
        ]
        if parte
    )
    segunda_linha = " - ".join(
        parte
        for parte in [
            str(dados.get("bairro") or "").strip(),
            " / ".join(
                parte
                for parte in [
                    str(dados.get("cidade") or "").strip(),
                    str(dados.get("uf") or "").strip(),
                ]
                if parte
            ).strip(" /"),
        ]
        if parte
    )
    cep = str(dados.get("cep") or "").strip()
    partes = [parte for parte in [primeira_linha, segunda_linha, f"CEP {cep}" if cep else ""] if parte]
    return " | ".join(partes)


def serializar_motorista(row: AgendamentoMotorista | None) -> dict | None:
    if not row:
        return None
    return {
        "id": row.id,
        "nome": str(row.nome or "").strip(),
        "telefone": str(row.telefone or "").strip(),
        "cnh": str(row.cnh or "").strip(),
        "observacoes": str(row.observacoes or "").strip(),
        "ativo": bool(row.ativo),
    }


def listar_motoristas_agendamento(q: str = "", incluir_inativos: bool = False) -> list[dict]:
    termo = str(q or "").strip()
    query = AgendamentoMotorista.query
    if not incluir_inativos:
        query = query.filter_by(ativo=True)
    if termo:
        like = f"%{termo}%"
        query = query.filter(or_(AgendamentoMotorista.nome.ilike(like), AgendamentoMotorista.telefone.ilike(like), AgendamentoMotorista.cnh.ilike(like)))
    rows = query.order_by(AgendamentoMotorista.ativo.desc(), AgendamentoMotorista.nome.asc()).all()
    return [serializar_motorista(row) for row in rows]


def salvar_motorista_agendamento(payload: dict) -> AgendamentoMotorista:
    motorista_id = payload.get("id")
    row = None
    if motorista_id not in (None, ""):
        try:
            row = AgendamentoMotorista.query.get(int(motorista_id))
        except (TypeError, ValueError):
            row = None
    if not row:
        row = AgendamentoMotorista(created_at=datetime.now())
        db.session.add(row)

    nome = str(payload.get("nome") or "").strip()
    if not nome:
        raise ValueError("Informe o nome do motorista.")

    row.nome = nome
    row.telefone = str(payload.get("telefone") or "").strip() or None
    row.cnh = str(payload.get("cnh") or "").strip() or None
    row.observacoes = str(payload.get("observacoes") or "").strip() or None
    row.ativo = bool(payload.get("ativo", True))
    row.updated_at = datetime.now()
    db.session.flush()
    return row


def sincronizar_motoristas_usuarios(*, commit: bool = False) -> None:
    """Sincroniza motoristas com usuarios de role Motorista.

    Para cada usuario com role 'Motorista', verifica se existe um
    AgendamentoMotorista vinculado via usuario_username. Se nao existir,
    tenta encontrar um motorista pelo nome igual ao username ou cria um novo.
    """
    usuarios_motoristas = Usuario.query.filter_by(role="Motorista").all()
    for usuario in usuarios_motoristas:
        username = str(usuario.username or "").strip()
        if not username:
            continue
        # Verifica se ja existe motorista vinculado a este usuario
        motorista = AgendamentoMotorista.query.filter_by(usuario_username=username).first()
        if motorista:
            continue
        # Tenta encontrar um motorista pelo nome igual ao username
        motorista = AgendamentoMotorista.query.filter(
            AgendamentoMotorista.nome.ilike(username),
            AgendamentoMotorista.usuario_username.is_(None),
        ).first()
        if motorista:
            motorista.usuario_username = username
            motorista.updated_at = datetime.now()
        else:
            # Cria novo motorista vinculado ao usuario
            motorista = AgendamentoMotorista(
                nome=username,
                usuario_username=username,
                ativo=True,
                created_at=datetime.now(),
                updated_at=datetime.now(),
            )
            db.session.add(motorista)
    if commit:
        db.session.commit()


def _normalizar_float(valor) -> float | None:
    if valor in (None, ""):
        return None
    try:
        return float(str(valor).replace(",", "."))
    except (TypeError, ValueError):
        return None


def _cache_geocode_path() -> Path:
    return Path(current_app.instance_path) / "agendamento" / "geocode_cache.json"


def _load_geocode_cache() -> dict:
    caminho = _cache_geocode_path()
    if not caminho.exists():
        return {}
    try:
        return json.loads(caminho.read_text(encoding="utf-8")) or {}
    except Exception:
        return {}


def _save_geocode_cache(cache: dict) -> None:
    try:
        caminho = _cache_geocode_path()
        caminho.parent.mkdir(parents=True, exist_ok=True)
        caminho.write_text(json.dumps(cache, ensure_ascii=True, indent=2), encoding="utf-8")
    except Exception:
        pass


def montar_endereco_rota(dados: dict | None) -> str:
    dados = dados or {}
    partes = [
        str(dados.get("logradouro") or "").strip(),
        str(dados.get("numero") or "").strip(),
        str(dados.get("bairro") or "").strip(),
        str(dados.get("cidade") or "").strip(),
        str(dados.get("uf") or "").strip(),
        str(dados.get("cep") or "").strip(),
    ]
    texto = ", ".join(parte for parte in partes if parte)
    return re.sub(r"\s+", " ", texto).strip(" ,")


def _geocode_endereco(endereco: str) -> dict | None:
    global _GEOCODE_LAST_REQUEST_TS

    consulta = str(endereco or "").strip()
    if not consulta:
        return None

    chave = normalizar_texto(consulta)
    cache = _load_geocode_cache()
    cached = cache.get(chave)
    if isinstance(cached, dict) and cached.get("lat") not in (None, "") and cached.get("lng") not in (None, ""):
        return cached

    espera = 1.05 - (time.time() - _GEOCODE_LAST_REQUEST_TS)
    if espera > 0:
        time.sleep(espera)

    try:
        response = requests.get(
            current_app.config.get("AGENDAMENTO_GEOCODE_URL"),
            params={"q": consulta, "format": "jsonv2", "limit": 1},
            headers={"User-Agent": "ColumbiaSync/1.0 (agendamento-logistica)"},
            timeout=int(current_app.config.get("AGENDAMENTO_GEOCODE_TIMEOUT_SECONDS", 8)),
        )
        _GEOCODE_LAST_REQUEST_TS = time.time()
        response.raise_for_status()
        payload = response.json()
    except Exception:
        return None

    if not isinstance(payload, list) or not payload:
        return None

    row = payload[0]
    lat = _normalizar_float(row.get("lat"))
    lng = _normalizar_float(row.get("lon"))
    if lat is None or lng is None:
        return None

    resultado = {
        "lat": lat,
        "lng": lng,
        "display_name": str(row.get("display_name") or "").strip(),
    }
    cache[chave] = resultado
    _save_geocode_cache(cache)
    return resultado


def _haversine_km(lat1: float, lng1: float, lat2: float, lng2: float) -> float:
    raio = 6371.0
    dlat = math.radians(lat2 - lat1)
    dlng = math.radians(lng2 - lng1)
    a = (
        math.sin(dlat / 2) ** 2
        + math.cos(math.radians(lat1)) * math.cos(math.radians(lat2)) * math.sin(dlng / 2) ** 2
    )
    return 2 * raio * math.asin(math.sqrt(a))


def montar_waze_url_agendamento(destino: dict | None) -> str:
    destino = destino or {}
    lat = _normalizar_float(destino.get("latitude"))
    lng = _normalizar_float(destino.get("longitude"))
    if lat is not None and lng is not None:
        return f"https://www.google.com/maps/dir/?api=1&destination={lat:.6f},{lng:.6f}"
    busca = quote(montar_endereco_rota(destino))
    if not busca:
        return ""
    return f"https://www.google.com/maps/dir/?api=1&destination={busca}"


def estimar_rota_agendamento(destino: dict | None, *, origem_latitude=None, origem_longitude=None) -> dict:
    destino = dict(destino or {})
    resultado = {
        "disponivel": False,
        "km_estimado": None,
        "km_estimado_retorno": None,
        "origem_fonte": "",
        "waze_url": montar_waze_url_agendamento(destino),
        "destino_latitude": None,
        "destino_longitude": None,
        "origem_latitude": None,
        "origem_longitude": None,
    }

    origem_lat = _normalizar_float(origem_latitude)
    origem_lng = _normalizar_float(origem_longitude)
    if origem_lat is not None and origem_lng is not None:
        resultado["origem_fonte"] = "gps_atual"
    else:
        cfg_lat = _normalizar_float(current_app.config.get("AGENDAMENTO_BASE_LATITUDE"))
        cfg_lng = _normalizar_float(current_app.config.get("AGENDAMENTO_BASE_LONGITUDE"))
        if cfg_lat is not None and cfg_lng is not None and cfg_lat != 0 and cfg_lng != 0:
            origem_lat = cfg_lat
            origem_lng = cfg_lng
            resultado["origem_fonte"] = "base_configurada"
        else:
            base_origem = str(current_app.config.get("AGENDAMENTO_BASE_ORIGEM") or "").strip()
            geocode_origem = _geocode_endereco(base_origem) if base_origem else None
            if geocode_origem:
                origem_lat = geocode_origem["lat"]
                origem_lng = geocode_origem["lng"]
                resultado["origem_fonte"] = "base_geocodificada"

    destino_lat = _normalizar_float(destino.get("latitude"))
    destino_lng = _normalizar_float(destino.get("longitude"))
    if destino_lat is None or destino_lng is None:
        geocode_destino = _geocode_endereco(montar_endereco_rota(destino))
        if geocode_destino:
            destino_lat = geocode_destino["lat"]
            destino_lng = geocode_destino["lng"]

    resultado["origem_latitude"] = origem_lat
    resultado["origem_longitude"] = origem_lng
    resultado["destino_latitude"] = destino_lat
    resultado["destino_longitude"] = destino_lng
    resultado["waze_url"] = montar_waze_url_agendamento(
        {
            **destino,
            "latitude": destino_lat,
            "longitude": destino_lng,
        }
    )

    if origem_lat is None or origem_lng is None or destino_lat is None or destino_lng is None:
        return resultado

    fator = max(1.0, float(current_app.config.get("AGENDAMENTO_ESTIMATIVA_KM_FATOR", 1.28)))
    km_base = _haversine_km(origem_lat, origem_lng, destino_lat, destino_lng)
    km_estimado = round(km_base * fator, 1)
    resultado["disponivel"] = True
    resultado["km_estimado"] = km_estimado
    resultado["km_estimado_retorno"] = round(km_estimado * 2, 1)
    return resultado


def _path_padrao_cadastro(tipo: str) -> Path:
    if tipo == "fornecedor":
        return Path(str(current_app.config.get("AGENDAMENTO_FORNECEDORES_XLSX") or "")).expanduser()
    return Path(str(current_app.config.get("AGENDAMENTO_CLIENTES_XLSX") or "")).expanduser()


def _model_por_tipo(tipo: str):
    tipo_limpo = str(tipo or "").strip().lower()
    if tipo_limpo not in _CADASTRO_MODELS:
        raise ValueError("Tipo de cadastro inválido.")
    return tipo_limpo, _CADASTRO_MODELS[tipo_limpo]


def _resolve_header_map(tipo: str, headers: list) -> dict[str, int]:
    aliases = _HEADER_ALIASES[tipo]
    normalized = {normalizar_texto(header): idx for idx, header in enumerate(headers)}
    mapping = {}
    for campo, allowed in aliases.items():
        for alias in allowed:
            if alias in normalized:
                mapping[campo] = normalized[alias]
                break
    if "nome" not in mapping:
        raise RuntimeError(f"Não foi possível identificar a coluna principal de nome em {tipo}.")
    return mapping


def _abrir_workbook(arquivo=None, caminho: Path | None = None):
    if load_workbook is None:
        raise RuntimeError("Dependência openpyxl não disponível neste ambiente.")
    if arquivo is not None:
        data = arquivo.read()
        try:
            arquivo.seek(0)
        except Exception:
            pass
        return load_workbook(filename=io.BytesIO(data), read_only=True, data_only=True)
    if not caminho or not caminho.exists():
        raise RuntimeError(f"Arquivo não encontrado: {caminho}")
    return load_workbook(filename=str(caminho), read_only=True, data_only=True)


def _row_to_payload(tipo: str, row: list, header_map: dict[str, int]) -> dict:
    def get(campo: str) -> str:
        idx = header_map.get(campo)
        if idx is None or idx >= len(row):
            return ""
        return valor_texto(row[idx])

    nome = get("nome")
    razao_social = get("razao_social") or nome
    payload = {
        "codigo": get("codigo"),
        "nome": nome or razao_social,
        "razao_social": razao_social or nome,
        "cnpj_cpf": limpar_documento(get("cnpj_cpf"))[:20],
        "tipo_pessoa": get("tipo_pessoa"),
        "contato": get("contato"),
        "telefone": get("telefone"),
        "telefone_secundario": get("telefone_secundario"),
        "email": get("email"),
        "logradouro": get("logradouro"),
        "numero": "",
        "complemento": "",
        "bairro": get("bairro"),
        "cidade": get("cidade"),
        "uf": get("uf")[:2].upper(),
        "cep": limpar_documento(get("cep"))[:10],
        "observacoes": get("observacoes"),
        "codigo_integracao": get("codigo_integracao"),
        "ativo": not bool_from_excel(get("inativo")),
    }
    if tipo == "fornecedor":
        payload["janela_atendimento"] = get("janela_atendimento")
        payload["transportadora"] = get("transportadora")
    else:
        payload["municipio_entrega"] = get("municipio_entrega")
    return payload


def importar_cadastros_excel(tipo: str, arquivo=None, nome_arquivo: str | None = None) -> dict:
    tipo_limpo, model = _model_por_tipo(tipo)
    caminho_padrao = _path_padrao_cadastro(tipo_limpo)
    workbook = _abrir_workbook(arquivo=arquivo, caminho=None if arquivo is not None else caminho_padrao)

    try:
        ws = workbook.active
        rows = ws.iter_rows(values_only=True)
        header_row = next(rows, None)
        if not header_row:
            raise RuntimeError("O arquivo está vazio.")
        header_map = _resolve_header_map(tipo_limpo, list(header_row))

        registros = []
        inconsistencias = []
        importado_em = datetime.now()
        arquivo_origem = nome_arquivo or (arquivo.filename if getattr(arquivo, "filename", None) else caminho_padrao.name)

        for idx, row in enumerate(rows, start=2):
            payload = _row_to_payload(tipo_limpo, list(row or []), header_map)
            if not any(str(valor or "").strip() for valor in payload.values()):
                continue
            if not payload["nome"]:
                inconsistencias.append(f"Linha {idx}: nome não informado.")
                continue
            payload["fonte_arquivo"] = arquivo_origem
            payload["importado_em"] = importado_em
            registros.append(model(**payload))

        db.session.query(model).delete()
        if registros:
            db.session.bulk_save_objects(registros)
        db.session.commit()

        return {
            "tipo": tipo_limpo,
            "arquivo": arquivo_origem,
            "importados": len(registros),
            "ativos": sum(1 for registro in registros if bool(registro.ativo)),
            "inconsistencias": inconsistencias[:25],
            "total_inconsistencias": len(inconsistencias),
            "colunas_identificadas": sorted(header_map.keys()),
        }
    finally:
        workbook.close()


def ensure_cadastros_base_carregados() -> None:
    for tipo, model in _CADASTRO_MODELS.items():
        try:
            if model.query.count() > 0:
                continue
            caminho = _path_padrao_cadastro(tipo)
            if caminho.exists():
                importar_cadastros_excel(tipo)
        except Exception:
            db.session.rollback()


def listar_cadastros(tipo: str, q: str = "", limit: int = 30) -> list[dict]:
    ensure_cadastros_base_carregados()
    tipo_limpo, model = _model_por_tipo(tipo)
    limit = max(1, min(int(limit or 30), 100))
    termo = str(q or "").strip()
    query = model.query
    if termo:
        like = f"%{termo}%"
        termo_doc = limpar_documento(termo)
        conditions = [
            model.nome.ilike(like),
            model.razao_social.ilike(like),
            model.codigo.ilike(like),
        ]
        if termo_doc:
            conditions.append(model.cnpj_cpf.ilike(f"%{termo_doc}%"))
        query = query.filter(or_(*conditions))
    rows = query.order_by(model.ativo.desc(), model.nome.asc()).limit(limit).all()
    return [serializar_cadastro(row, tipo_limpo) for row in rows]


def resumo_cadastros() -> dict:
    ensure_cadastros_base_carregados()
    ultimo_fornecedor = (
        AgendamentoFornecedor.query.order_by(AgendamentoFornecedor.importado_em.desc()).first()
    )
    ultimo_cliente = AgendamentoCliente.query.order_by(AgendamentoCliente.importado_em.desc()).first()
    return {
        "fornecedores": AgendamentoFornecedor.query.count(),
        "clientes": AgendamentoCliente.query.count(),
        "fornecedores_ativos": AgendamentoFornecedor.query.filter_by(ativo=True).count(),
        "clientes_ativos": AgendamentoCliente.query.filter_by(ativo=True).count(),
        "ultima_importacao_fornecedores": ultimo_fornecedor.importado_em if ultimo_fornecedor else None,
        "ultima_importacao_clientes": ultimo_cliente.importado_em if ultimo_cliente else None,
        "arquivo_fornecedores": ultimo_fornecedor.fonte_arquivo if ultimo_fornecedor else os.path.basename(_path_padrao_cadastro("fornecedor")),
        "arquivo_clientes": ultimo_cliente.fonte_arquivo if ultimo_cliente else os.path.basename(_path_padrao_cadastro("cliente")),
    }


def serializar_cadastro(row, tipo: str | None = None) -> dict:
    tipo_valor = tipo or ("fornecedor" if isinstance(row, AgendamentoFornecedor) else "cliente")
    payload = {
        "id": row.id,
        "tipo": tipo_valor,
        "codigo": str(row.codigo or "").strip(),
        "nome": str(row.nome or "").strip(),
        "razao_social": str(getattr(row, "razao_social", "") or "").strip(),
        "cnpj_cpf": str(getattr(row, "cnpj_cpf", "") or "").strip(),
        "contato": str(getattr(row, "contato", "") or "").strip(),
        "telefone": str(getattr(row, "telefone", "") or "").strip(),
        "telefone_secundario": str(getattr(row, "telefone_secundario", "") or "").strip(),
        "email": str(getattr(row, "email", "") or "").strip(),
        "logradouro": str(getattr(row, "logradouro", "") or "").strip(),
        "numero": str(getattr(row, "numero", "") or "").strip(),
        "complemento": str(getattr(row, "complemento", "") or "").strip(),
        "bairro": str(getattr(row, "bairro", "") or "").strip(),
        "cidade": str(getattr(row, "cidade", "") or "").strip(),
        "uf": str(getattr(row, "uf", "") or "").strip(),
        "cep": str(getattr(row, "cep", "") or "").strip(),
        "observacoes": str(getattr(row, "observacoes", "") or "").strip(),
        "ativo": bool(getattr(row, "ativo", True)),
    }
    if hasattr(row, "janela_atendimento"):
        payload["janela_atendimento"] = str(getattr(row, "janela_atendimento", "") or "").strip()
    if hasattr(row, "transportadora"):
        payload["transportadora"] = str(getattr(row, "transportadora", "") or "").strip()
    if hasattr(row, "municipio_entrega"):
        payload["municipio_entrega"] = str(getattr(row, "municipio_entrega", "") or "").strip()
    payload["endereco_formatado"] = formatar_endereco_logistico(payload)
    return payload


def localizar_cadastro(tipo: str, *, codigo: str = "", documento: str = "", nome: str = ""):
    ensure_cadastros_base_carregados()
    tipo_limpo, model = _model_por_tipo(tipo)
    codigo_limpo = str(codigo or "").strip()
    documento_limpo = limpar_documento(documento)
    nome_limpo = str(nome or "").strip()

    if codigo_limpo:
        found = model.query.filter_by(codigo=codigo_limpo).order_by(model.ativo.desc(), model.id.asc()).first()
        if found:
            return found
    if documento_limpo:
        found = (
            model.query
            .filter(model.cnpj_cpf == documento_limpo)
            .order_by(model.ativo.desc(), model.id.asc())
            .first()
        )
        if found:
            return found
    if nome_limpo:
        like = f"%{nome_limpo}%"
        found = (
            model.query
            .filter(or_(model.nome.ilike(like), model.razao_social.ilike(like)))
            .order_by(model.ativo.desc(), model.nome.asc())
            .first()
        )
        if found:
            return found
    return None


def listar_veiculos_agendamento() -> list[AgendamentoVeiculo]:
    return (
        AgendamentoVeiculo.query
        .filter_by(ativo=True)
        .order_by(AgendamentoVeiculo.ordem_exibicao.asc(), AgendamentoVeiculo.nome_exibicao.asc())
        .all()
    )


def _payload_fornecedor_vazio() -> dict:
    return {
        "id": None,
        "tipo": "fornecedor",
        "codigo": "",
        "nome": "",
        "razao_social": "",
        "cnpj_cpf": "",
        "contato": "",
        "telefone": "",
        "telefone_secundario": "",
        "email": "",
        "logradouro": "",
        "numero": "",
        "complemento": "",
        "bairro": "",
        "cidade": "",
        "uf": "",
        "cep": "",
        "observacoes": "",
        "ativo": True,
        "endereco_formatado": "",
    }


def _inferir_fonte_oc(linhas: list[dict]) -> dict:
    if not linhas:
        return obter_fonte_pedidos_google_sheets()
    fonte_tipo = str(next((linha.get("fonte_dados") for linha in linhas if linha.get("fonte_dados")), "") or PEDIDOS_FONTE_GOOGLE_SHEETS)
    fonte = obter_fonte_pedidos_google_sheets()
    fonte["tipo"] = fonte_tipo
    fonte["label"] = label_fonte_pedidos(fonte_tipo)
    if fonte_tipo != PEDIDOS_FONTE_GOOGLE_SHEETS:
        fonte["url"] = ""
        fonte["csv_url"] = ""
    return fonte


def _enriquecer_parceiro_com_linha_oc(parceiro: dict, linha_referencia: dict | None) -> dict:
    parceiro = dict(parceiro or {})
    linha_referencia = linha_referencia or {}
    campos = {
        "codigo": "fornecedor_codigo",
        "nome": "fornecedor_nome",
        "razao_social": "fornecedor_nome",
        "cnpj_cpf": "fornecedor_cnpj",
        "contato": "contato",
        "telefone": "telefone",
        "email": "email",
        "logradouro": "logradouro",
        "numero": "numero",
        "complemento": "complemento",
        "bairro": "bairro",
        "cidade": "cidade",
        "uf": "uf",
        "cep": "cep",
        "observacoes": "observacoes",
    }
    for campo_destino, campo_origem in campos.items():
        valor_atual = str(parceiro.get(campo_destino) or "").strip()
        valor_novo = str(linha_referencia.get(campo_origem) or "").strip()
        if not valor_atual and valor_novo:
            parceiro[campo_destino] = valor_novo
    parceiro["endereco_formatado"] = formatar_endereco_logistico(parceiro)
    return parceiro


def _consultar_oc_agendamento_legacy(numero_oc: str) -> dict:
    numero_oc_limpo = str(numero_oc or "").strip()
    fonte = obter_fonte_pedidos_google_sheets()
    if not numero_oc_limpo:
        return {"encontrada": False, "error": "Informe o número da OC."}

    linhas = buscar_linhas_pedido(numero_oc_limpo)
    if not linhas:
        return {
            "encontrada": False,
            "error": "OC não encontrada na integração atual.",
            "numero_oc": numero_oc_limpo,
            "itens": [],
        }

    referencia = (
        ItemNota.query
        .filter(ItemNota.pedido_compra == numero_oc_limpo)
        .order_by(ItemNota.data_importacao.desc(), ItemNota.id.desc())
        .first()
    )
    fornecedor_nome = str(getattr(referencia, "fornecedor", "") or "").strip()
    fornecedor_doc = str(getattr(referencia, "cnpj_emitente", "") or "").strip()
    cadastro = localizar_cadastro("fornecedor", documento=fornecedor_doc, nome=fornecedor_nome)

    parceiro = serializar_cadastro(cadastro, "fornecedor") if cadastro else {
        "id": None,
        "tipo": "fornecedor",
        "codigo": "",
        "nome": fornecedor_nome,
        "razao_social": fornecedor_nome,
        "cnpj_cpf": fornecedor_doc,
        "contato": "",
        "telefone": "",
        "telefone_secundario": "",
        "email": "",
        "logradouro": "",
        "numero": "",
        "complemento": "",
        "bairro": "",
        "cidade": "",
        "uf": "",
        "cep": "",
        "observacoes": "",
        "ativo": True,
        "endereco_formatado": "",
    }

    itens = [
        {
            "codigo_item": str(linha.get("codigo_material") or "").strip(),
            "descricao": str(linha.get("descricao_material") or "").strip() or f"Item da OC {numero_oc_limpo}",
            "quantidade": float(linha.get("qtd") or 0.0),
            "unidade": "",
            "volumes": 0.0,
            "observacoes": "",
        }
        for linha in linhas
    ]

    warning = ""
    if not parceiro.get("nome"):
        warning = "Fornecedor não identificado automaticamente. Selecione ou digite o cadastro manualmente."
    elif not parceiro.get("logradouro"):
        warning = "Fornecedor localizado, mas com endereço incompleto no cadastro."

    return {
        "encontrada": True,
        "numero_oc": numero_oc_limpo,
        "fornecedor": parceiro,
        "itens": itens,
        "warning": warning,
    }


def consultar_oc_agendamento(numero_oc: str) -> dict:
    numero_oc_limpo = str(numero_oc or "").strip()
    fonte = obter_fonte_pedidos_google_sheets()
    if not numero_oc_limpo:
        return {"encontrada": False, "error": "Informe o numero da OC.", "fonte": fonte}

    linhas = buscar_linhas_pedido(numero_oc_limpo)
    fonte = _inferir_fonte_oc(linhas)
    if not linhas:
        return {
            "encontrada": False,
            "error": f"OC nao encontrada na fonte {fonte['label']}.",
            "numero_oc": numero_oc_limpo,
            "itens": [],
            "fonte": fonte,
        }

    linha_referencia = next(
        (
            linha
            for linha in linhas
            if any(
                str(linha.get(chave) or "").strip()
                for chave in ("fornecedor_codigo", "fornecedor_nome", "fornecedor_cnpj", "logradouro", "cidade", "uf")
            )
        ),
        {},
    )
    referencia = (
        ItemNota.query
        .filter(ItemNota.pedido_compra == numero_oc_limpo)
        .order_by(ItemNota.data_importacao.desc(), ItemNota.id.desc())
        .first()
    )
    fornecedor_nome = str(linha_referencia.get("fornecedor_nome") or getattr(referencia, "fornecedor", "") or "").strip()
    fornecedor_doc = str(linha_referencia.get("fornecedor_cnpj") or getattr(referencia, "cnpj_emitente", "") or "").strip()
    fornecedor_codigo = str(linha_referencia.get("fornecedor_codigo") or "").strip()
    cadastro = localizar_cadastro(
        "fornecedor",
        codigo=fornecedor_codigo,
        documento=fornecedor_doc,
        nome=fornecedor_nome,
    )

    parceiro = serializar_cadastro(cadastro, "fornecedor") if cadastro else _payload_fornecedor_vazio()
    if fornecedor_nome and not parceiro.get("nome"):
        parceiro["nome"] = fornecedor_nome
    if fornecedor_nome and not parceiro.get("razao_social"):
        parceiro["razao_social"] = fornecedor_nome
    if fornecedor_doc and not parceiro.get("cnpj_cpf"):
        parceiro["cnpj_cpf"] = fornecedor_doc
    parceiro = _enriquecer_parceiro_com_linha_oc(parceiro, linha_referencia)

    itens = [
        {
            "codigo_item": str(linha.get("codigo_material") or "").strip(),
            "descricao": str(linha.get("descricao_material") or "").strip() or f"Item da OC {numero_oc_limpo}",
            "quantidade": float(linha.get("qtd") or 0.0),
            "unidade": "",
            "volumes": 0.0,
            "observacoes": "",
        }
        for linha in linhas
    ]

    warning = ""
    if not parceiro.get("nome"):
        warning = f"Fornecedor nao identificado automaticamente na fonte {fonte['label']}. Selecione ou digite o cadastro manualmente."
    elif not parceiro.get("logradouro"):
        warning = f"Fornecedor localizado na fonte {fonte['label']}, mas com endereco incompleto."

    return {
        "encontrada": True,
        "numero_oc": numero_oc_limpo,
        "fornecedor": parceiro,
        "itens": itens,
        "warning": warning,
        "fonte": fonte,
    }


def _pick(documento: dict, *chaves: str) -> str:
    for chave in chaves:
        valor = documento.get(chave)
        if valor not in (None, ""):
            return valor_texto(valor)
    return ""


def _normalizar_itens_consyste(documento: dict) -> list[dict]:
    itens = documento.get("itens")
    if not isinstance(itens, list):
        return []
    normalizados = []
    for item in itens:
        if not isinstance(item, dict):
            continue
        normalizados.append(
            {
                "codigo_item": _pick(item, "codigo", "cProd", "codigo_item"),
                "descricao": _pick(item, "descricao", "xProd", "produto", "nome"),
                "quantidade": float(item.get("qtd") or item.get("quantidade") or item.get("qtd_comercial") or 0.0),
                "unidade": _pick(item, "unidade", "uCom", "sigla_unidade"),
                "volumes": float(item.get("volumes") or 0.0),
                "observacoes": _pick(item, "observacao", "obs"),
            }
        )
    return normalizados


def consultar_nf_agendamento(numero_nf: str) -> dict:
    numero_limpo = limpar_documento(numero_nf)
    if not numero_limpo:
        return {"encontrada": False, "error": "Informe o número da NF."}

    alguma_caixa_respondeu = False
    autenticacao_falhou = False
    documento_encontrado = None
    # "todos" e "recebidos" retornam HTTP 500 no ambiente atual — usamos apenas as caixas suportadas
    for caixa in ("emitidos", "recebidos"):
        try:
            ok, status_code, payload = listar_nfes_consyste_por_caixa(
                caixa=caixa,
                q=f"numero:{numero_limpo}",
                campos=_CONSYSTE_FIELDS,
                timeout=20,
            )
        except Exception:
            # Falha de rede nesta caixa: não marca como indisponível, tenta a próxima
            continue

        if status_code in {401, 403}:
            autenticacao_falhou = True
            continue
        if not ok:
            # 500 em uma caixa específica não significa que o serviço está fora —
            # basta que pelo menos uma caixa responda com 200
            continue

        alguma_caixa_respondeu = True
        documentos = payload.get("documentos") if isinstance(payload, dict) else payload
        for documento in documentos or []:
            if limpar_documento(documento.get("numero")) == numero_limpo:
                documento_encontrado = documento
                break
        if documento_encontrado:
            break

    if not documento_encontrado:
        if autenticacao_falhou:
            return {
                "encontrada": False,
                "numero_nf": numero_limpo,
                "error": "A integração Consyste recusou a consulta. Confira o token configurado.",
            }
        if not alguma_caixa_respondeu:
            return {
                "encontrada": False,
                "numero_nf": numero_limpo,
                "error": "Consyste indisponível no momento. Tente novamente ou preencha os dados manualmente.",
            }
        return {
            "encontrada": False,
            "numero_nf": numero_limpo,
            "error": f"NF {numero_limpo} não encontrada na Consyste. Verifique o número ou preencha os dados manualmente.",
        }

    nome_cliente = _pick(documento_encontrado, "dest_nome", "cliente_nome", "destinatario_nome", "razao_social")
    cnpj_cliente = limpar_documento(_pick(documento_encontrado, "dest_cnpj", "cliente_cnpj", "cnpj"))
    cadastro = localizar_cadastro("cliente", documento=cnpj_cliente, nome=nome_cliente)
    parceiro_base = serializar_cadastro(cadastro, "cliente") if cadastro else {
        "id": None,
        "tipo": "cliente",
        "codigo": "",
        "nome": "",
        "razao_social": "",
        "cnpj_cpf": "",
        "contato": "",
        "telefone": "",
        "telefone_secundario": "",
        "email": "",
        "logradouro": "",
        "numero": "",
        "complemento": "",
        "bairro": "",
        "cidade": "",
        "uf": "",
        "cep": "",
        "observacoes": "",
        "ativo": True,
        "endereco_formatado": "",
    }

    parceiro = dict(parceiro_base)
    parceiro["nome"] = nome_cliente or parceiro.get("nome") or parceiro.get("razao_social") or ""
    parceiro["razao_social"] = parceiro["razao_social"] or parceiro["nome"]
    parceiro["cnpj_cpf"] = cnpj_cliente or parceiro.get("cnpj_cpf") or ""
    parceiro["contato"] = parceiro.get("contato") or _pick(documento_encontrado, "contato", "dest_contato")
    parceiro["telefone"] = parceiro.get("telefone") or _pick(documento_encontrado, "dest_telefone", "telefone")
    parceiro["email"] = parceiro.get("email") or _pick(documento_encontrado, "dest_email", "email")
    parceiro["logradouro"] = parceiro.get("logradouro") or _pick(documento_encontrado, "dest_endereco", "endereco", "logradouro")
    parceiro["numero"] = parceiro.get("numero") or _pick(documento_encontrado, "dest_numero", "numero_endereco", "numero")
    parceiro["complemento"] = parceiro.get("complemento") or _pick(documento_encontrado, "dest_complemento", "complemento")
    parceiro["bairro"] = parceiro.get("bairro") or _pick(documento_encontrado, "dest_bairro", "bairro")
    parceiro["cidade"] = parceiro.get("cidade") or _pick(documento_encontrado, "dest_cidade", "cidade")
    parceiro["uf"] = (parceiro.get("uf") or _pick(documento_encontrado, "dest_uf", "uf"))[:2].upper()
    parceiro["cep"] = parceiro.get("cep") or limpar_documento(_pick(documento_encontrado, "dest_cep", "cep"))
    parceiro["observacoes"] = parceiro.get("observacoes") or _pick(
        documento_encontrado, "observacao", "informacoes_complementares"
    )
    parceiro["endereco_formatado"] = formatar_endereco_logistico(parceiro)

    warning = ""
    if not parceiro["logradouro"] or not parceiro["cidade"] or not parceiro["uf"]:
        warning = "A NF foi localizada, mas o endereço precisa ser complementado antes de salvar."

    return {
        "encontrada": True,
        "numero_nf": numero_limpo,
        "documento_id": str(documento_encontrado.get("id") or "").strip(),
        "chave": str(documento_encontrado.get("chave") or "").strip(),
        "cliente": parceiro,
        "itens": _normalizar_itens_consyste(documento_encontrado),
        "warning": warning,
    }
