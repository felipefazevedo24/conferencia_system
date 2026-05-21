from __future__ import annotations

import re
import unicodedata
import json
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl
from sqlalchemy import func

from ..extensions import db
from ..models import ClassificacaoContabilItem, ClassificacaoContabilPadrao, ItemNota, PlanoContaDominio


ARQUIVOS_PADRAO_2026 = [
    Path(r"z:\CUSTOS\ESTOQUE\2026\04-Abril 2026\Classificação entradas\Entradas 01 a 29 - FINAL PARA CONTABIL.xlsx"),
    Path(r"z:\CUSTOS\ESTOQUE\2026\03-Março 2026\Classificações entradas\Entradas 30 _ CHECk _ FINAL.xlsx"),
    Path(r"z:\CUSTOS\ESTOQUE\2026\02-Fevereiro 2026\Classificações entradas\Entradas Fevereiro  01 A 26.xlsx"),
]
BUNDLED_PADROES_PATH = Path(__file__).resolve().parents[1] / "data" / "classificacao_contabil_padroes_2026.json"
PLANO_CONTAS_DOMINIO_PATH = Path(__file__).resolve().parents[2] / "Plano contas dominio.xlsx"

CFOP_ENTRADA_DESCRICOES = {
    "1101": "Compra para industrializacao",
    "1102": "Compra para comercializacao",
    "1124": "Industrializacao efetuada por outra empresa",
    "1151": "Transferencia para industrializacao",
    "1152": "Transferencia para comercializacao",
    "1201": "Devolucao de venda de producao do estabelecimento",
    "1202": "Devolucao de venda de mercadoria adquirida ou recebida de terceiros",
    "1252": "Compra de energia eletrica",
    "1352": "Aquisicao de servico de transporte",
    "1403": "Compra para comercializacao em operacao com substituicao tributaria",
    "1551": "Compra de bem para o ativo imobilizado",
    "1556": "Compra de material para uso ou consumo",
    "1901": "Entrada para industrializacao por encomenda",
    "1902": "Retorno de mercadoria remetida para industrializacao por encomenda",
    "1915": "Entrada de mercadoria recebida para conserto ou reparo",
    "1916": "Retorno de mercadoria remetida para conserto ou reparo",
    "1924": "Entrada para industrializacao por conta e ordem do adquirente",
    "1933": "Aquisicao de servico tributado pelo ISSQN",
    "2101": "Compra para industrializacao de outro estado",
    "2102": "Compra para comercializacao de outro estado",
    "2124": "Industrializacao efetuada por outra empresa de outro estado",
    "2403": "Compra para comercializacao de outro estado com substituicao tributaria",
    "2551": "Compra de bem para o ativo imobilizado de outro estado",
    "2556": "Compra de material para uso ou consumo de outro estado",
    "2915": "Entrada de mercadoria recebida para conserto ou reparo de outro estado",
    "2916": "Retorno de mercadoria remetida para conserto ou reparo de outro estado",
    "2933": "Aquisicao de servico tributado pelo ISSQN de outro estado",
}


HEADER_ALIASES = {
    "fornecedor": ["Entrada:Fornecedor", "fornecedor", "FORNECEDOR"],
    "cfop": ["Itens da Entrada de NF:CFOP", "cfop", "CFOP"],
    "codigo": ["Itens da Entrada de NF:Cód. interno /Cód. fabricante", "cod_interno", "Material"],
    "descricao": ["Itens da Entrada de NF:Descrição", "produto"],
    "conta": ["Conta"],
    "nome_conta": ["Nome conta"],
    "comentario": ["comentario", "comentário"],
}


def normalizar_texto(value: Any) -> str:
    text = str(value or "").strip().upper()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = re.sub(r"[^A-Z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def _normalizar_codigo(value: Any) -> str:
    return normalizar_texto(value).replace(" ", "")


def _somente_digitos(value: Any) -> str:
    return re.sub(r"\D", "", str(value or ""))


def normalizar_conta(value: Any) -> str:
    texto = str(value or "").strip()
    if not texto:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return _somente_digitos(texto) or texto


def cfop_entrada(value: Any) -> str:
    variantes = _cfop_variantes(value)
    for variante in variantes:
        if variante.startswith(("1", "2", "3")):
            return variante
    return variantes[0] if variantes else ""


def descricao_cfop_entrada(value: Any) -> str:
    entrada = cfop_entrada(value)
    return CFOP_ENTRADA_DESCRICOES.get(entrada, "")


def buscar_nome_conta(conta: Any) -> str:
    codigo = normalizar_conta(conta)
    if not codigo:
        return ""
    plano = PlanoContaDominio.query.filter_by(codigo_conta=codigo).first()
    return plano.nome_conta if plano else ""


def importar_plano_contas_dominio(path: Path | None = None, forcar: bool = False) -> dict:
    path = path or PLANO_CONTAS_DOMINIO_PATH
    if not path.exists():
        return {"origem": "arquivo_indisponivel", "contas_criadas": 0, "contas_atualizadas": 0, "contas_total": PlanoContaDominio.query.count()}
    if not forcar and PlanoContaDominio.query.count() > 0:
        return {"origem": "banco", "contas_criadas": 0, "contas_atualizadas": 0, "contas_total": PlanoContaDominio.query.count()}

    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    criadas = 0
    atualizadas = 0
    agora = datetime.now()
    for sheet in workbook.worksheets:
        headers = []
        header_row = None
        for row_number, row in enumerate(sheet.iter_rows(min_row=1, max_row=10, values_only=True), start=1):
            headers = [normalizar_texto(value).replace(" ", "_").lower() for value in row]
            if "codigo_conta" in headers and "nome_conta" in headers:
                header_row = row_number
                break
        if not header_row:
            continue
        index = {header: pos for pos, header in enumerate(headers) if header}
        for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
            def cell(campo):
                pos = index.get(campo)
                return row[pos] if pos is not None and pos < len(row) else ""

            codigo = normalizar_conta(cell("codigo_conta"))
            nome = str(cell("nome_conta") or "").strip()
            if not codigo or not nome:
                continue
            plano = PlanoContaDominio.query.filter_by(codigo_conta=codigo).first()
            if plano is None:
                plano = PlanoContaDominio(codigo_conta=codigo)
                db.session.add(plano)
                criadas += 1
            else:
                atualizadas += 1
            plano.nome_conta = nome[:180]
            plano.classificacao_conta = str(cell("classificacao_conta") or "").strip()[:60]
            plano.tipo_conta = str(cell("tipo_conta") or "").strip()[:20]
            plano.origem = path.name[:160]
            plano.atualizado_em = agora
    db.session.commit()
    return {"origem": str(path), "contas_criadas": criadas, "contas_atualizadas": atualizadas, "contas_total": PlanoContaDominio.query.count()}


def garantir_plano_contas_dominio() -> None:
    if PlanoContaDominio.query.count() == 0:
        importar_plano_contas_dominio()


def _cfop_variantes(value: Any) -> list[str]:
    cfop = re.sub(r"\D", "", str(value or "").strip())
    if not cfop:
        return [""]
    variantes = {cfop}
    if len(cfop) == 4:
        if cfop.startswith("5"):
            variantes.add("1" + cfop[1:])
        elif cfop.startswith("6"):
            variantes.add("2" + cfop[1:])
        elif cfop.startswith("1"):
            variantes.add("5" + cfop[1:])
        elif cfop.startswith("2"):
            variantes.add("6" + cfop[1:])
    return list(variantes)


def _filtrar_cfop(query, variantes: list[str]):
    return query.filter(ClassificacaoContabilPadrao.cfop.in_(variantes))


def _valor(row: tuple, index: dict[str, int], campo: str) -> Any:
    for header in HEADER_ALIASES[campo]:
        pos = index.get(header)
        if pos is not None and pos < len(row):
            return row[pos]
    return None


def _linha_cabecalho(sheet) -> tuple[int, list[str]] | None:
    for row_number, row in enumerate(sheet.iter_rows(min_row=1, max_row=6, values_only=True), start=1):
        headers = [str(value).strip() if value is not None else "" for value in row]
        if "Conta" in headers and ("Nome conta" in headers) and any(h in headers for h in HEADER_ALIASES["fornecedor"]):
            return row_number, headers
    return None


def _iter_linhas_classificadas(workbook_source, origem_nome: str):
    workbook = openpyxl.load_workbook(workbook_source, read_only=True, data_only=True)
    for sheet in workbook.worksheets:
        header_info = _linha_cabecalho(sheet)
        if not header_info:
            continue
        header_row, headers = header_info
        index = {header: pos for pos, header in enumerate(headers) if header}

        for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
            conta = str(_valor(row, index, "conta") or "").strip()
            nome_conta = str(_valor(row, index, "nome_conta") or "").strip()
            if not conta or not nome_conta:
                continue
            fornecedor = _valor(row, index, "fornecedor")
            cfop = _valor(row, index, "cfop")
            codigo = _valor(row, index, "codigo")
            descricao = _valor(row, index, "descricao")
            if not any([fornecedor, cfop, codigo, descricao]):
                continue
            yield {
                "fornecedor_norm": normalizar_texto(fornecedor),
                "cfop": str(cfop or "").strip(),
                "codigo_norm": _normalizar_codigo(codigo),
                "descricao_norm": normalizar_texto(descricao),
                "conta": conta,
                "nome_conta": nome_conta,
                "comentario": str(_valor(row, index, "comentario") or "").strip(),
                "origem": f"{origem_nome} / {sheet.title}",
            }


def importar_padroes_excel(paths: list[Path] | None = None) -> dict:
    paths = paths or ARQUIVOS_PADRAO_2026
    agregadas: Counter[tuple] = Counter()
    exemplos: dict[tuple, dict] = {}
    arquivos_lidos = []

    for path in paths:
        if not path.exists():
            continue
        arquivos_lidos.append(str(path))
        for row in _iter_linhas_classificadas(path, path.name):
            key = (
                row["fornecedor_norm"],
                row["cfop"],
                row["codigo_norm"],
                row["descricao_norm"],
                row["conta"],
            )
            agregadas[key] += 1
            exemplos.setdefault(key, row)

    agora = datetime.now()
    criados = 0
    atualizados = 0
    for key, ocorrencias in agregadas.items():
        row = exemplos[key]
        padrao = ClassificacaoContabilPadrao.query.filter_by(
            fornecedor_norm=row["fornecedor_norm"],
            cfop=row["cfop"],
            codigo_norm=row["codigo_norm"],
            descricao_norm=row["descricao_norm"],
            conta=row["conta"],
        ).first()
        if padrao is None:
            padrao = ClassificacaoContabilPadrao(
                fornecedor_norm=row["fornecedor_norm"],
                cfop=row["cfop"],
                codigo_norm=row["codigo_norm"],
                descricao_norm=row["descricao_norm"],
                conta=row["conta"],
            )
            db.session.add(padrao)
            criados += 1
        else:
            atualizados += 1
        padrao.nome_conta = row["nome_conta"]
        padrao.comentario = row["comentario"][:500]
        padrao.ocorrencias = int(ocorrencias)
        padrao.origem = row["origem"][:120]
        padrao.atualizado_em = agora

    db.session.commit()
    return {
        "arquivos_lidos": len(arquivos_lidos),
        "linhas_agregadas": sum(agregadas.values()),
        "padroes_criados": criados,
        "padroes_atualizados": atualizados,
        "padroes_total": ClassificacaoContabilPadrao.query.count(),
    }


def importar_padroes_uploads(arquivos) -> dict:
    agregadas: Counter[tuple] = Counter()
    exemplos: dict[tuple, dict] = {}
    arquivos_lidos = []

    for arquivo in arquivos:
        filename = str(getattr(arquivo, "filename", "") or "upload.xlsx").strip()
        if not filename.lower().endswith((".xlsx", ".xlsm")):
            continue
        try:
            arquivo.stream.seek(0)
        except Exception:
            pass
        arquivos_lidos.append(filename)
        for row in _iter_linhas_classificadas(arquivo.stream, filename):
            key = (
                row["fornecedor_norm"],
                row["cfop"],
                row["codigo_norm"],
                row["descricao_norm"],
                row["conta"],
            )
            agregadas[key] += 1
            exemplos.setdefault(key, row)

    agora = datetime.now()
    criados = 0
    atualizados = 0
    for key, ocorrencias in agregadas.items():
        row = exemplos[key]
        padrao = ClassificacaoContabilPadrao.query.filter_by(
            fornecedor_norm=row["fornecedor_norm"],
            cfop=row["cfop"],
            codigo_norm=row["codigo_norm"],
            descricao_norm=row["descricao_norm"],
            conta=row["conta"],
        ).first()
        if padrao is None:
            padrao = ClassificacaoContabilPadrao(
                fornecedor_norm=row["fornecedor_norm"],
                cfop=row["cfop"],
                codigo_norm=row["codigo_norm"],
                descricao_norm=row["descricao_norm"],
                conta=row["conta"],
            )
            db.session.add(padrao)
            criados += 1
        else:
            atualizados += 1
        padrao.nome_conta = row["nome_conta"]
        padrao.comentario = row["comentario"][:500]
        padrao.ocorrencias = int(padrao.ocorrencias or 0) + int(ocorrencias)
        padrao.origem = row["origem"][:120]
        padrao.atualizado_em = agora

    db.session.commit()
    return {
        "arquivos_lidos": len(arquivos_lidos),
        "linhas_agregadas": sum(agregadas.values()),
        "padroes_criados": criados,
        "padroes_atualizados": atualizados,
        "padroes_total": ClassificacaoContabilPadrao.query.count(),
    }


def importar_padroes_internos(forcar: bool = False) -> dict:
    if not forcar and ClassificacaoContabilPadrao.query.count() > 0:
        return {
            "origem": "banco",
            "padroes_criados": 0,
            "padroes_atualizados": 0,
            "padroes_total": ClassificacaoContabilPadrao.query.count(),
        }
    if not BUNDLED_PADROES_PATH.exists():
        return {
            "origem": "interno_indisponivel",
            "padroes_criados": 0,
            "padroes_atualizados": 0,
            "padroes_total": ClassificacaoContabilPadrao.query.count(),
        }

    data = json.loads(BUNDLED_PADROES_PATH.read_text(encoding="utf-8"))
    agora = datetime.now()
    criados = 0
    atualizados = 0
    for row in data.get("padroes", []):
        padrao = ClassificacaoContabilPadrao.query.filter_by(
            fornecedor_norm=str(row.get("fornecedor_norm") or ""),
            cfop=str(row.get("cfop") or ""),
            codigo_norm=str(row.get("codigo_norm") or ""),
            descricao_norm=str(row.get("descricao_norm") or ""),
            conta=str(row.get("conta") or ""),
        ).first()
        if padrao is None:
            padrao = ClassificacaoContabilPadrao(
                fornecedor_norm=str(row.get("fornecedor_norm") or ""),
                cfop=str(row.get("cfop") or ""),
                codigo_norm=str(row.get("codigo_norm") or ""),
                descricao_norm=str(row.get("descricao_norm") or ""),
                conta=str(row.get("conta") or ""),
            )
            db.session.add(padrao)
            criados += 1
        else:
            atualizados += 1
        padrao.nome_conta = str(row.get("nome_conta") or "")[:180]
        padrao.comentario = str(row.get("comentario") or "")[:500]
        padrao.ocorrencias = int(row.get("ocorrencias") or 1)
        padrao.origem = str(row.get("origem") or "Base interna 2026")[:120]
        padrao.atualizado_em = agora

    db.session.commit()
    return {
        "origem": "base_interna_2026",
        "padroes_criados": criados,
        "padroes_atualizados": atualizados,
        "padroes_total": ClassificacaoContabilPadrao.query.count(),
    }


def garantir_padroes_internos() -> None:
    if ClassificacaoContabilPadrao.query.count() == 0:
        importar_padroes_internos()


def _conta_dominante(query):
    rows = query.all()
    if not rows:
        return None
    total = sum(int(r.ocorrencias or 1) for r in rows)
    por_conta: dict[tuple[str, str], dict[str, Any]] = {}
    for row in rows:
        key = (row.conta, row.nome_conta)
        slot = por_conta.setdefault(key, {"ocorrencias": 0, "row": row})
        slot["ocorrencias"] += int(row.ocorrencias or 1)
    melhor = max(por_conta.values(), key=lambda entry: entry["ocorrencias"])
    return melhor["row"], melhor["ocorrencias"], total


def _motivo_sem_padrao(item: ItemNota) -> str:
    codigo = _normalizar_codigo(item.codigo)
    descricao = normalizar_texto(item.descricao)
    cfops = _cfop_variantes(item.cfop)
    fornecedor = normalizar_texto(item.fornecedor)
    if not codigo:
        return "Sem codigo do item"
    if not str(item.cfop or "").strip():
        return "Sem CFOP"
    if fornecedor and ClassificacaoContabilPadrao.query.filter_by(fornecedor_norm=fornecedor).count() == 0:
        return "Fornecedor sem historico"
    if codigo and ClassificacaoContabilPadrao.query.filter_by(codigo_norm=codigo).count() == 0:
        return "Codigo sem historico"
    if descricao and ClassificacaoContabilPadrao.query.filter_by(descricao_norm=descricao).count() == 0:
        return "Descricao sem historico"
    if cfops and _filtrar_cfop(ClassificacaoContabilPadrao.query, cfops).count() == 0:
        return "CFOP sem historico"
    return "Padrao inconclusivo"


def _tipo_regra(metodo: str) -> str:
    metodo_norm = normalizar_texto(metodo)
    if "FORNECEDOR" in metodo_norm and "CODIGO" in metodo_norm and "CFOP" in metodo_norm:
        return "Alta"
    if "CODIGO" in metodo_norm:
        return "Media"
    if "DESCRICAO" in metodo_norm:
        return "Media"
    if "CFOP DOMINANTE" in metodo_norm:
        return "Baixa"
    return "Sem regra"


def sugerir_classificacao_item(item: ItemNota) -> dict:
    garantir_padroes_internos()
    fornecedor = normalizar_texto(item.fornecedor)
    codigo_origem = getattr(item, "codigo_grv", None) or item.codigo
    codigo = _normalizar_codigo(codigo_origem)
    descricao = normalizar_texto(item.descricao)
    cfop = str(getattr(item, "cfop_grv", None) or item.cfop or "").strip()
    cfops = _cfop_variantes(cfop)

    tentativas = [
        (
            "Fornecedor + codigo + CFOP",
            98,
            _filtrar_cfop(
                ClassificacaoContabilPadrao.query.filter_by(fornecedor_norm=fornecedor, codigo_norm=codigo),
                cfops,
            ),
        ),
        (
            "Fornecedor + codigo",
            92,
            ClassificacaoContabilPadrao.query.filter_by(fornecedor_norm=fornecedor, codigo_norm=codigo),
        ),
        (
            "Codigo + CFOP",
            86,
            _filtrar_cfop(ClassificacaoContabilPadrao.query.filter_by(codigo_norm=codigo), cfops),
        ),
        (
            "Codigo do item",
            82,
            ClassificacaoContabilPadrao.query.filter_by(codigo_norm=codigo),
        ),
        (
            "Fornecedor + descricao + CFOP",
            82,
            _filtrar_cfop(
                ClassificacaoContabilPadrao.query.filter_by(fornecedor_norm=fornecedor, descricao_norm=descricao),
                cfops,
            ),
        ),
        (
            "CFOP + descricao",
            70,
            _filtrar_cfop(ClassificacaoContabilPadrao.query.filter_by(descricao_norm=descricao), cfops),
        ),
        (
            "Descricao do item",
            66,
            ClassificacaoContabilPadrao.query.filter_by(descricao_norm=descricao),
        ),
        (
            "CFOP dominante",
            55,
            _filtrar_cfop(ClassificacaoContabilPadrao.query, cfops),
        ),
    ]

    for metodo, confianca_base, query in tentativas:
        if "codigo" in metodo.lower() and not codigo:
            continue
        if "descricao" in metodo.lower() and not descricao:
            continue
        if "cfop" in metodo.lower() and not cfop:
            continue
        dominante = _conta_dominante(query)
        if not dominante:
            continue
        regra, ocorrencias, total = dominante
        consenso = ocorrencias / max(total, 1)
        confianca = min(99, round(confianca_base * consenso))
        return {
            "regra": regra,
            "conta": regra.conta,
            "nome_conta": regra.nome_conta,
            "comentario": regra.comentario or "",
            "confianca": int(confianca),
            "metodo": metodo,
            "motivo_pendencia": "",
            "tipo_regra": _tipo_regra(metodo),
            "status": "Classificado" if confianca >= 80 else "Revisar",
        }

    return {
        "regra": None,
        "conta": "",
        "nome_conta": "",
        "comentario": "",
        "confianca": 0,
        "metodo": "Sem padrão",
        "motivo_pendencia": _motivo_sem_padrao(item),
        "tipo_regra": "Sem regra",
        "status": "Pendente",
    }


def classificar_item(item: ItemNota, sobrescrever_manual: bool = False, sincronizar_grv: bool = True) -> ClassificacaoContabilItem:
    existente = ClassificacaoContabilItem.query.filter_by(item_nota_id=item.id).first()
    if existente and existente.status in {"Revisado", "Aprovado"} and not sobrescrever_manual:
        return existente

    if sincronizar_grv and getattr(item, "numero_lancamento", None):
        try:
            from .erp_lancamento_service import sincronizar_codigos_grv_nota

            sincronizar_codigos_grv_nota(
                str(item.numero_nota or ""),
                codigo_lancamento=str(item.numero_lancamento or ""),
                chave=str(item.chave_acesso or ""),
            )
            db.session.refresh(item)
        except Exception:
            pass

    sugestao = sugerir_classificacao_item(item)
    if sugestao["conta"]:
        sugestao["nome_conta"] = buscar_nome_conta(sugestao["conta"]) or sugestao["nome_conta"]
    agora = datetime.now()
    registro = existente or ClassificacaoContabilItem(item_nota_id=item.id, numero_nota=str(item.numero_nota or ""))
    registro.numero_nota = str(item.numero_nota or "")
    registro.fornecedor = item.fornecedor
    registro.codigo_item = str(getattr(item, "codigo_grv", None) or item.codigo or "").strip()
    registro.descricao_item = item.descricao
    registro.cfop = str(getattr(item, "cfop_grv", None) or item.cfop or "").strip()
    registro.conta = sugestao["conta"]
    registro.nome_conta = sugestao["nome_conta"]
    registro.comentario = sugestao["comentario"][:500]
    registro.confianca = sugestao["confianca"]
    registro.metodo = sugestao["metodo"]
    registro.status = sugestao["status"]
    registro.motivo_pendencia = sugestao.get("motivo_pendencia") or ""
    registro.tipo_regra = sugestao.get("tipo_regra") or ""
    registro.regra_id = sugestao["regra"].id if sugestao["regra"] else None
    registro.atualizado_em = agora
    if existente is None:
        registro.criado_em = agora
        db.session.add(registro)
    return registro


def classificar_nota(numero_nota: str, sobrescrever_manual: bool = False) -> int:
    itens = ItemNota.query.filter_by(numero_nota=str(numero_nota), status="Lançado").all()
    try:
        from .erp_lancamento_service import sincronizar_codigos_grv_itens

        sincronizar_codigos_grv_itens([
            item for item in itens
            if getattr(item, "numero_lancamento", None) and getattr(item, "tributos_origem", None) != "GRV"
        ])
        for item in itens:
            db.session.refresh(item)
    except Exception:
        pass
    total = 0
    for item in itens:
        classificar_item(item, sobrescrever_manual=sobrescrever_manual, sincronizar_grv=False)
        total += 1
    db.session.commit()
    return total


def classificar_lancadas_desde_2026(
    limite: int = 500,
    data_inicio: datetime | None = None,
    data_fim: datetime | None = None,
    sobrescrever_manual: bool = False,
) -> int:
    data_inicio = max(data_inicio or datetime(2026, 1, 1), datetime(2026, 1, 1))
    query = (
        ItemNota.query
        .filter(ItemNota.status == "Lançado")
        .filter(ItemNota.data_lancamento >= data_inicio)
        .order_by(ItemNota.data_lancamento.desc())
    )
    if data_fim:
        query = query.filter(ItemNota.data_lancamento <= data_fim)
    if limite:
        query = query.limit(limite)
    itens = query.all()
    try:
        from .erp_lancamento_service import sincronizar_codigos_grv_itens

        sincronizar_codigos_grv_itens([
            item for item in itens
            if getattr(item, "numero_lancamento", None) and getattr(item, "tributos_origem", None) != "GRV"
        ])
        for item in itens:
            db.session.refresh(item)
    except Exception:
        pass
    total = 0
    for item in itens:
        antes = ClassificacaoContabilItem.query.filter_by(item_nota_id=item.id).first()
        if antes and antes.status in {"Revisado", "Aprovado"} and not sobrescrever_manual:
            continue
        classificar_item(item, sobrescrever_manual=sobrescrever_manual, sincronizar_grv=False)
        total += 1
    db.session.commit()
    return total


def classificar_lancadas_sem_registro(
    limite: int = 1000,
    data_inicio: datetime | None = None,
    data_fim: datetime | None = None,
) -> int:
    data_inicio = max(data_inicio or datetime(2026, 1, 1), datetime(2026, 1, 1))
    query = (
        ItemNota.query.outerjoin(ClassificacaoContabilItem, ClassificacaoContabilItem.item_nota_id == ItemNota.id)
        .filter(ItemNota.status == "Lançado")
        .filter(ItemNota.data_lancamento >= data_inicio)
        .filter(ClassificacaoContabilItem.id.is_(None))
        .order_by(ItemNota.data_lancamento.desc())
    )
    if data_fim:
        query = query.filter(ItemNota.data_lancamento <= data_fim)
    if limite:
        query = query.limit(limite)
    itens = query.all()
    try:
        from .erp_lancamento_service import sincronizar_codigos_grv_itens

        sincronizar_codigos_grv_itens([
            item for item in itens
            if getattr(item, "numero_lancamento", None) and getattr(item, "tributos_origem", None) != "GRV"
        ])
        for item in itens:
            db.session.refresh(item)
    except Exception:
        pass
    total = 0
    for item in itens:
        classificar_item(item, sincronizar_grv=False)
        total += 1
    db.session.commit()
    return total


def resumo_classificacoes(data_inicio: datetime | None = None, data_fim: datetime | None = None) -> dict:
    query = ClassificacaoContabilItem.query.join(ItemNota, ItemNota.id == ClassificacaoContabilItem.item_nota_id)
    query = query.filter(ItemNota.data_lancamento >= datetime(2026, 1, 1))
    if data_inicio:
        query = query.filter(ItemNota.data_lancamento >= data_inicio)
    if data_fim:
        query = query.filter(ItemNota.data_lancamento <= data_fim)

    total = query.count()
    revisao = query.filter(ClassificacaoContabilItem.status.in_(["Revisar", "Pendente"])).count()
    revisados = query.filter(ClassificacaoContabilItem.status == "Revisado").count()
    aprovados = query.filter(ClassificacaoContabilItem.status == "Aprovado").count()
    auto = query.filter(ClassificacaoContabilItem.status == "Classificado").count()
    confianca_media = query.with_entities(func.avg(ClassificacaoContabilItem.confianca)).scalar() or 0
    return {
        "total": total,
        "classificados": auto,
        "revisao": revisao,
        "revisados": revisados,
        "aprovados": aprovados,
        "confianca_media": round(float(confianca_media), 1),
        "padroes": ClassificacaoContabilPadrao.query.count(),
    }
