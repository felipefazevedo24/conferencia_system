"""Pagina publica (sem login) para o proprio cliente/fornecedor atualizar o
cadastro. Fluxo:

    1) escolhe se e Cliente ou Fornecedor
    2) informa o CNPJ e busca os dados automaticamente (BrasilAPI)
    3) revisa/edita os campos, escolhe o Regime Tributario (obrigatorio)
    4) confirma que o e-mail e o contato real da empresa
    5) envia -> fica em fila de revisao interna (CadastroAtualizacaoPublica)
"""
from __future__ import annotations

import io
import json

from flask import Blueprint, Response, jsonify, render_template, request

from ..auth import permission_required
from ..extensions import db
from ..models import CadastroAtualizacaoPublica
from ..services import cadastro_workflow_service as cad_svc

atualizacao_cadastral_bp = Blueprint("atualizacao_cadastral", __name__)

REGIMES_VALIDOS = {"Lucro Real", "Lucro Presumido", "Simples Nacional", "MEI"}
TIPOS_VALIDOS = {"cliente", "fornecedor"}
CONTRIBUINTE_ICMS_VALIDOS = {
    "Sim, sou Contribuinte do ICMS",
    "Contribuinte isento de IE",
    "Não contribuinte",
}


@atualizacao_cadastral_bp.route("/atualizacao-cadastral")
def pagina_atualizacao_cadastral():
    return render_template("atualizacao_cadastral.html")


@atualizacao_cadastral_bp.route("/api/atualizacao-cadastral/consultar-cnpj", methods=["POST"])
def consultar_cnpj_publico():
    payload = request.get_json(silent=True) or {}
    cnpj = str(payload.get("cnpj") or "").strip()
    try:
        dados = cad_svc.consultar_cartao_cnpj(cnpj)
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400
    except Exception:  # noqa: BLE001
        return jsonify({"error": "Não foi possível consultar o CNPJ agora. Tente novamente em instantes."}), 502
    return jsonify({"sucesso": True, "dados": dados})


@atualizacao_cadastral_bp.route("/api/atualizacao-cadastral", methods=["POST"])
def enviar_atualizacao_cadastral():
    payload = request.get_json(silent=True) or {}

    tipo = str(payload.get("tipo") or "").strip().lower()
    documento = cad_svc.normalizar_documento(str(payload.get("documento") or ""))
    regime = str(payload.get("regime_tributario") or "").strip()
    contribuinte_icms = str(payload.get("contribuinte_icms") or "").strip()
    possui_beneficios_fiscais = bool(payload.get("possui_beneficios_fiscais"))
    beneficios_fiscais_descricao = str(payload.get("beneficios_fiscais_descricao") or "").strip()
    email = str(payload.get("email") or "").strip()
    email_confirmado = bool(payload.get("email_confirmado"))

    if tipo not in TIPOS_VALIDOS:
        return jsonify({"error": "Selecione se você é Cliente ou Fornecedor."}), 400
    if not cad_svc.cnpj_valido(documento):
        return jsonify({"error": "Informe um CNPJ válido com 14 dígitos."}), 400
    if regime not in REGIMES_VALIDOS:
        return jsonify({"error": "Selecione o Regime Tributário."}), 400
    if contribuinte_icms not in CONTRIBUINTE_ICMS_VALIDOS:
        return jsonify({"error": "Selecione se a empresa é Contribuinte do ICMS."}), 400
    if possui_beneficios_fiscais and not beneficios_fiscais_descricao:
        return jsonify({"error": "Especifique qual(is) benefício(s) fiscal(is) a empresa possui."}), 400
    if not email:
        return jsonify({"error": "Informe o e-mail de contato da empresa."}), 400
    if not email_confirmado:
        return jsonify({"error": "Confirme que o e-mail informado é o contato real da empresa."}), 400

    def _campo(nome, limite=None):
        valor = str(payload.get(nome) or "").strip()
        if limite:
            valor = valor[:limite]
        return valor or None

    registro = CadastroAtualizacaoPublica(
        tipo=tipo,
        documento=cad_svc.formatar_cnpj(documento),
        razao_social=_campo("razao_social", 220),
        nome_fantasia=_campo("nome_fantasia", 220),
        inscricao_estadual=_campo("inscricao_estadual", 40),
        regime_tributario=regime,
        contribuinte_icms=contribuinte_icms,
        possui_beneficios_fiscais=possui_beneficios_fiscais,
        beneficios_fiscais_descricao=(beneficios_fiscais_descricao[:500] if possui_beneficios_fiscais else None),
        endereco=_campo("endereco", 300),
        cep=_campo("cep", 12),
        municipio=_campo("municipio", 120),
        uf=(str(payload.get("uf") or "").strip().upper()[:2] or None),
        telefone=_campo("telefone", 40),
        email=email[:160],
        email_confirmado=email_confirmado,
        contato=_campo("contato", 120),
        observacoes=_campo("observacoes", 500),
        situacao_cadastral=_campo("situacao_cadastral", 60),
        fonte_cnpj=_campo("fonte_cnpj", 40),
        dados_json=json.dumps(payload, ensure_ascii=False),
        status="Pendente de revisão",
        origem_ip=(request.headers.get("X-Forwarded-For") or request.remote_addr or "")[:60],
    )
    db.session.add(registro)
    db.session.commit()
    return jsonify({"sucesso": True, "id": registro.id})


# ---------------------------------------------------------------------------
# Area interna (com login) para revisar/exportar as atualizacoes recebidas
# ---------------------------------------------------------------------------

STATUS_VALIDOS = {"Pendente de revisão", "Em análise", "Concluído", "Descartado"}


def _filtrar_registros():
    tipo = str(request.args.get("tipo") or "").strip().lower()
    status = str(request.args.get("status") or "").strip()
    busca = str(request.args.get("busca") or "").strip()

    query = CadastroAtualizacaoPublica.query
    if tipo in TIPOS_VALIDOS:
        query = query.filter(CadastroAtualizacaoPublica.tipo == tipo)
    if status:
        query = query.filter(CadastroAtualizacaoPublica.status == status)
    if busca:
        termo = f"%{busca}%"
        query = query.filter(
            db.or_(
                CadastroAtualizacaoPublica.documento.ilike(termo),
                CadastroAtualizacaoPublica.razao_social.ilike(termo),
                CadastroAtualizacaoPublica.nome_fantasia.ilike(termo),
                CadastroAtualizacaoPublica.email.ilike(termo),
            )
        )
    return query.order_by(CadastroAtualizacaoPublica.created_at.desc())


def _serializar(registro):
    return {
        "id": registro.id,
        "tipo": registro.tipo,
        "documento": registro.documento,
        "razao_social": registro.razao_social,
        "nome_fantasia": registro.nome_fantasia,
        "inscricao_estadual": registro.inscricao_estadual,
        "regime_tributario": registro.regime_tributario,
        "contribuinte_icms": registro.contribuinte_icms,
        "possui_beneficios_fiscais": registro.possui_beneficios_fiscais,
        "beneficios_fiscais_descricao": registro.beneficios_fiscais_descricao,
        "endereco": registro.endereco,
        "cep": registro.cep,
        "municipio": registro.municipio,
        "uf": registro.uf,
        "telefone": registro.telefone,
        "email": registro.email,
        "email_confirmado": registro.email_confirmado,
        "contato": registro.contato,
        "observacoes": registro.observacoes,
        "situacao_cadastral": registro.situacao_cadastral,
        "status": registro.status,
        "created_at": registro.created_at.strftime("%d/%m/%Y %H:%M") if registro.created_at else "",
    }


@atualizacao_cadastral_bp.route("/api/admin/atualizacoes-cadastrais")
@permission_required("PAGE_ADMIN_ATUALIZACOES_CADASTRAIS")
def listar_atualizacoes_cadastrais():
    registros = _filtrar_registros().limit(2000).all()
    total = len(registros)
    pendentes = sum(1 for r in registros if r.status == "Pendente de revisão")
    return jsonify(
        {
            "total": total,
            "pendentes": pendentes,
            "itens": [_serializar(r) for r in registros],
        }
    )


@atualizacao_cadastral_bp.route("/api/admin/atualizacoes-cadastrais/<int:registro_id>/status", methods=["POST"])
@permission_required("PAGE_ADMIN_ATUALIZACOES_CADASTRAIS")
def atualizar_status_atualizacao(registro_id):
    payload = request.get_json(silent=True) or {}
    novo_status = str(payload.get("status") or "").strip()
    if novo_status not in STATUS_VALIDOS:
        return jsonify({"error": "Status inválido."}), 400
    registro = CadastroAtualizacaoPublica.query.get_or_404(registro_id)
    registro.status = novo_status
    db.session.commit()
    return jsonify({"sucesso": True})


@atualizacao_cadastral_bp.route("/api/admin/atualizacoes-cadastrais/export.xlsx")
@permission_required("PAGE_ADMIN_ATUALIZACOES_CADASTRAIS")
def exportar_atualizacoes_cadastrais_xlsx():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    registros = _filtrar_registros().limit(20000).all()

    colunas = [
        ("Recebido em", "created_at"),
        ("Tipo", "tipo"),
        ("CNPJ", "documento"),
        ("Razão social", "razao_social"),
        ("Nome fantasia", "nome_fantasia"),
        ("Inscrição estadual", "inscricao_estadual"),
        ("Regime tributário", "regime_tributario"),
        ("Contribuinte do ICMS", "contribuinte_icms"),
        ("Possui benefícios fiscais", "possui_beneficios_fiscais"),
        ("Quais benefícios fiscais", "beneficios_fiscais_descricao"),
        ("Endereço", "endereco"),
        ("CEP", "cep"),
        ("Município", "municipio"),
        ("UF", "uf"),
        ("Telefone", "telefone"),
        ("E-mail", "email"),
        ("E-mail confirmado", "email_confirmado"),
        ("Contato", "contato"),
        ("Situação cadastral", "situacao_cadastral"),
        ("Observações", "observacoes"),
        ("Status", "status"),
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "Atualizações cadastrais"

    header_fill = PatternFill(start_color="0F62C9", end_color="0F62C9", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for col_idx, (titulo, _) in enumerate(colunas, start=1):
        cell = ws.cell(row=1, column=col_idx, value=titulo)
        cell.fill = header_fill
        cell.font = header_font

    for row_idx, registro in enumerate(registros, start=2):
        dados = _serializar(registro)
        for col_idx, (_, chave) in enumerate(colunas, start=1):
            valor = dados.get(chave)
            if chave == "tipo" and valor:
                valor = valor.capitalize()
            elif chave in ("email_confirmado", "possui_beneficios_fiscais"):
                valor = "Sim" if valor else "Não"
            ws.cell(row=row_idx, column=col_idx, value=valor)

    for col_idx, (titulo, _) in enumerate(colunas, start=1):
        largura = max(12, min(45, len(titulo) + 6))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = largura

    ws.freeze_panes = "A2"

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return Response(
        buffer.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=atualizacoes_cadastrais.xlsx"},
    )
