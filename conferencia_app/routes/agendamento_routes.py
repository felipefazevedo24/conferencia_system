from __future__ import annotations

from calendar import monthrange
import csv
import io
import json
import os
import re
from datetime import date, datetime, timedelta

from flask import Blueprint, current_app, jsonify, request, session, send_file
from sqlalchemy import or_, true
from werkzeug.utils import secure_filename

from ..auth import is_admin_session, permission_required, permission_required_any
from ..extensions import db
from ..models import (
    AgendamentoCliente,
    AgendamentoFornecedor,
    AgendamentoMotorista,
    AgendamentoSolicitacao,
    AgendamentoSolicitacaoHistorico,
    AgendamentoSolicitacaoItem,
    AgendamentoVeiculo,
    Viagem,
    ViagemParada,
    Usuario,
)
from ..services.email_service import enviar_email_agendamento_update
from ..compras.services import compras_service
from ..services.erp_estoque_service import buscar_estoque_grv
from ..services.agendamento_service import (
    PRIORIDADES_SOLICITACAO,
    STATUS_ATIVOS,
    STATUS_SOLICITACAO,
    TIPOS_SOLICITACAO,
    VEICULOS_KANBAN,
    consultar_nf_agendamento,
    consultar_oc_agendamento,
    estimar_rota_agendamento,
    ensure_cadastros_base_carregados,
    formatar_endereco_logistico,
    importar_cadastros_excel,
    listar_cadastros,
    listar_motoristas_agendamento,
    listar_veiculos_agendamento,
    montar_waze_url_agendamento,
    prioridade_label_agendamento,
    resumo_cadastros,
    salvar_motorista_agendamento,
    sincronizar_motoristas_usuarios,
    serializar_cadastro,
    serializar_motorista,
    status_label_agendamento,
)
from ..services.agendamento_ordem_coleta_pdf import gerar_ordem_coleta_pdf

try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover
    load_workbook = None


agendamento_bp = Blueprint("agendamento", __name__)

PRIORIDADE_ORDEM = {"Critica": 0, "Alta": 1, "Media": 2, "Baixa": 3}
ANEXO_SOLICITACAO_SUBDIR = "agendamento_solicitacoes_anexos"
_RECEBIMENTO_CACHE: dict[str, dict] = {}
_RECEBIMENTO_RISCO_CACHE: dict[str, dict] = {}


def _solicitacao_anexo_dir() -> str:
    pasta = os.path.join(current_app.instance_path, ANEXO_SOLICITACAO_SUBDIR)
    os.makedirs(pasta, exist_ok=True)
    return pasta


def _extrair_anexo_payload(payload_origem: dict) -> dict:
    anexo = payload_origem.get("anexo")
    return anexo if isinstance(anexo, dict) else {}


def _salvar_anexo_solicitacao(arquivo, solicitacao_id: int) -> dict | None:
    if not arquivo or not getattr(arquivo, "filename", ""):
        return None
    nome = secure_filename(arquivo.filename)
    extensao = nome.rsplit(".", 1)[-1].lower() if "." in nome else ""
    if extensao not in {"pdf", "jpg", "jpeg", "png", "webp"}:
        return None
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    nome_final = f"solicitacao_{solicitacao_id}_{stamp}_{nome}"
    caminho = os.path.join(_solicitacao_anexo_dir(), nome_final)
    arquivo.save(caminho)
    return {
        "nome_original": str(arquivo.filename or "").strip(),
        "nome_arquivo": nome_final,
        "caminho_relativo": os.path.join(ANEXO_SOLICITACAO_SUBDIR, nome_final).replace("\\", "/"),
        "mimetype": str(getattr(arquivo, "mimetype", "") or "").strip() or None,
    }


def _atualizar_payload_origem_anexo(registro: AgendamentoSolicitacao, anexo_meta: dict | None) -> None:
    if not anexo_meta:
        return
    payload = {}
    if registro.payload_origem:
        try:
            payload = json.loads(registro.payload_origem) or {}
        except Exception:
            payload = {}
    payload["anexo"] = anexo_meta
    registro.payload_origem = _json_text(payload)


def _notificar_solicitante_agendamento(row: AgendamentoSolicitacao, titulo: str, detalhe: str = "") -> None:
    try:
        usuario = Usuario.query.filter_by(username=row.solicitante).first()
        if not usuario or not usuario.email:
            return
        veiculo = AgendamentoVeiculo.query.get(row.veiculo_id) if row.veiculo_id else None
        url = request.url_root.rstrip("/") + "/logistica/solicitar-transporte"
        enviar_email_agendamento_update(
            usuario.email,
            f"[Transporte] {titulo} - {row.codigo or 'solicitacao'}",
            titulo,
            [
                ("Codigo", row.codigo or f"#{row.id}"),
                ("Tipo", {"COLETA": "Coleta", "ENTREGA": "Entrega", "AVULSA": "Avulsa"}.get(row.tipo, row.tipo or "")),
                ("Status", status_label_agendamento(row.status)),
                ("Documento", f"{row.documento_tipo or ''} {row.documento_numero or ''}".strip()),
                ("Veiculo", veiculo.nome_exibicao if veiculo else ""),
                ("Motorista", row.motorista_nome or ""),
                ("Saida", row.data_hora_saida_prevista.strftime("%d/%m/%Y %H:%M") if row.data_hora_saida_prevista else ""),
                ("Detalhe", detalhe),
            ],
            url,
        )
    except Exception:
        current_app.logger.exception("Falha ao notificar solicitante da solicitacao %s", getattr(row, "id", None))


def _json_text(payload) -> str | None:
    if payload is None:
        return None
    try:
        import json

        return json.dumps(payload, ensure_ascii=False, default=str)[:4000]
    except Exception:
        return str(payload)[:4000]


def _parse_datetime(value, field_name: str, required: bool = False) -> datetime | None:
    raw = str(value or "").strip()
    if not raw:
        if required:
            raise ValueError(f"Informe {field_name}.")
        return None
    try:
        return datetime.fromisoformat(raw)
    except ValueError as exc:
        raise ValueError(f"{field_name} inválido.") from exc


def _origem_documento_label(origem: str | None) -> str:
    mapping = {
        "GoogleSheets": "Google Sheets",
        "Consyste": "Consyste",
        "ERPPostgres": "ERP/Postgres",
        "Manual": "Manual",
        "AutoCIF": "Automático (CIF)",
        "ORDEM_DE_COMPRA": "Ordem de Compra",
        "ROMANEIO": "Romaneio",
    }
    return mapping.get(str(origem or "").strip(), str(origem or "").strip() or "---")


def _is_admin_or_compras() -> bool:
    return session.get("role") in {"Admin", "Compras"}


_FRETE_KEY_TOKENS = (
    "frete_por_conta",
    "fretepconta",
    "frete_conta",
    "conta_frete",
    "modalidade_frete",
    "frete_modalidade",
    "tipo_frete",
    "modfrete",
    "mod_frete",
    "incoterm",
    "frete",
)


def _tokens_cif() -> list[str]:
    raw = current_app.config.get("SOLICITACAO_CIF_VALORES_CIF") or ("CIF", "REMETENTE", "EMITENTE", "FORNECEDOR")
    if isinstance(raw, str):
        vals = [tok.strip().upper() for tok in raw.split(",") if tok and tok.strip()]
    else:
        vals = [str(tok).strip().upper() for tok in raw if str(tok).strip()]
    return vals or ["CIF", "REMETENTE", "EMITENTE", "FORNECEDOR"]


def _oc_modalidade_cif(oc_json) -> bool:
    if isinstance(oc_json, str):
        try:
            oc_json = json.loads(oc_json)
        except Exception:
            oc_json = {}
    if not isinstance(oc_json, dict):
        return False

    lowered = {str(k).lower(): v for k, v in oc_json.items()}
    valor = ""
    for token in _FRETE_KEY_TOKENS:
        if token in lowered:
            valor = str(lowered[token] or "").strip().upper()
            break
    if not valor:
        for key, raw_val in lowered.items():
            if "frete" in key and any(t in key for t in ("conta", "modalidade", "tipo", "cif", "incoterm")):
                valor = str(raw_val or "").strip().upper()
                break
    if not valor:
        return False

    for token in _tokens_cif():
        if valor == token or token in valor:
            return True
    return False


def _to_date(value) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value or "").strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%Y-%m-%d %H:%M:%S", "%d/%m/%Y %H:%M:%S"):
        try:
            return datetime.strptime(text[:19], fmt).date()
        except ValueError:
            continue
    try:
        return datetime.fromisoformat(text.replace("Z", "").split("+")[0]).date()
    except ValueError:
        return None


def _to_float(value) -> float:
    text = str(value or "").strip()
    if not text:
        return 0.0
    text = text.replace(" ", "")
    if "," in text and "." in text:
        text = text.replace(".", "").replace(",", ".")
    elif "," in text:
        text = text.replace(",", ".")
    try:
        return float(text)
    except (TypeError, ValueError):
        return 0.0


def _codigo_item_oc(item: dict) -> str:
    for key in (
        "cod_interno",
        "codigo",
        "codigo_produto",
        "produto_codigo",
        "codigo_item",
        "codigo_material",
        "item_codigo",
        "material",
    ):
        code = re.sub(r"[^A-Z0-9]", "", str(item.get(key) or "").strip().upper())
        if code:
            return code
    return ""


def _mapa_estoque_aliases(estoque_por_codigo: dict[str, dict]) -> dict[str, dict]:
    aliases: dict[str, dict] = {}
    for codigo, payload in (estoque_por_codigo or {}).items():
        bruto = str(codigo or "").strip().upper()
        normalizado = re.sub(r"[^A-Z0-9]", "", bruto)
        if not bruto and not normalizado:
            continue

        chaves = {bruto, normalizado}
        sem_zero = normalizado.lstrip("0")
        if sem_zero:
            chaves.add(sem_zero)

        somente_digitos = re.sub(r"\D", "", normalizado)
        if somente_digitos:
            chaves.add(somente_digitos)
            chaves.add(somente_digitos.lstrip("0") or "0")

        for chave in chaves:
            if chave:
                aliases.setdefault(chave, payload)
    return aliases


def _oc_ja_recebida(status_oc_nome: str) -> bool:
    status = str(status_oc_nome or "").strip().upper()
    if not status:
        return False
    return bool(re.search(r"RECEB|ENCERR|CONCL|FINAL|ATEND", status))


def _calcular_risco_estoque_oc(numero_oc: str, estoque_aliases: dict[str, dict]) -> dict:
    risco_estoque = "sem_dados"
    risco_estoque_label = "Sem dados"
    risco_estoque_detalhe = "Sem itens da OC ou sem saldo disponível para cálculo."

    if not numero_oc or not estoque_aliases:
        return {
            "risco_estoque": risco_estoque,
            "risco_estoque_label": risco_estoque_label,
            "risco_estoque_detalhe": risco_estoque_detalhe,
        }

    try:
        consulta = consultar_oc_agendamento(numero_oc)
    except Exception:
        return {
            "risco_estoque": risco_estoque,
            "risco_estoque_label": risco_estoque_label,
            "risco_estoque_detalhe": risco_estoque_detalhe,
        }

    itens = consulta.get("itens") if isinstance(consulta, dict) else []
    if not isinstance(itens, list):
        itens = []

    resumo_itens: dict[str, dict[str, float]] = {}
    for item in itens:
        if not isinstance(item, dict):
            continue
        codigo = _codigo_item_oc(item)
        if not codigo:
            continue
        qtd = _to_float(item.get("quantidade") or item.get("qtde") or item.get("qtd") or item.get("volume") or 0)
        if qtd <= 0:
            qtd = 1.0
        consumo = _to_float(item.get("consumo_diario") or item.get("consumo_medio_diario") or 0)
        slot = resumo_itens.setdefault(codigo, {"qtd": 0.0, "consumo": 0.0})
        slot["qtd"] += qtd
        if consumo > 0:
            slot["consumo"] = max(slot["consumo"], consumo)

    if not resumo_itens:
        return {
            "risco_estoque": risco_estoque,
            "risco_estoque_label": risco_estoque_label,
            "risco_estoque_detalhe": risco_estoque_detalhe,
        }

    cobertura_critica = False
    cobertura_atencao = False
    faltas = 0
    baixos = 0
    for codigo_item, vals in resumo_itens.items():
        qtd_oc = float(vals.get("qtd") or 0)
        consumo_diario = float(vals.get("consumo") or 0)
        saldo = float((estoque_aliases.get(codigo_item) or {}).get("qtde_total") or 0)
        if consumo_diario > 0:
            dias_cobertura = saldo / consumo_diario if consumo_diario else 0
            if dias_cobertura < 3:
                cobertura_critica = True
                faltas += 1
            elif dias_cobertura < 7:
                cobertura_atencao = True
                baixos += 1
            continue
        if saldo + 0.0001 < qtd_oc:
            cobertura_critica = True
            faltas += 1
        elif saldo < (qtd_oc * 1.5):
            cobertura_atencao = True
            baixos += 1

    if cobertura_critica:
        risco_estoque = "critico"
        risco_estoque_label = "Crítico"
        risco_estoque_detalhe = f"{faltas} item(ns) sem cobertura de saldo para esta OC."
    elif cobertura_atencao:
        risco_estoque = "atencao"
        risco_estoque_label = "Atenção"
        risco_estoque_detalhe = f"{baixos} item(ns) com cobertura baixa de estoque."
    else:
        risco_estoque = "ok"
        risco_estoque_label = "OK"
        risco_estoque_detalhe = "Itens da OC com cobertura de estoque no momento."

    return {
        "risco_estoque": risco_estoque,
        "risco_estoque_label": risco_estoque_label,
        "risco_estoque_detalhe": risco_estoque_detalhe,
    }


def _calcular_risco_estoque_por_ocs(ocs: list[str]) -> dict[str, dict]:
    numeros = []
    vistos: set[str] = set()
    for oc in ocs:
        numero = str(oc or "").strip()
        if not numero or numero in vistos:
            continue
        vistos.add(numero)
        numeros.append(numero)

    riscos: dict[str, dict] = {}
    pendentes: list[str] = []
    for numero in numeros:
        cached = _RECEBIMENTO_RISCO_CACHE.get(numero)
        if isinstance(cached, dict):
            riscos[numero] = dict(cached)
        else:
            pendentes.append(numero)

    if not pendentes:
        return riscos

    try:
        estoque_payload = buscar_estoque_grv(forcar_atualizacao=False)
        estoque_aliases = _mapa_estoque_aliases(estoque_payload.get("por_codigo") or {})
    except Exception:
        return riscos

    for numero in pendentes:
        risco_payload = _calcular_risco_estoque_oc(numero, estoque_aliases)
        riscos[numero] = risco_payload
        _RECEBIMENTO_RISCO_CACHE[numero] = dict(risco_payload)

    return riscos


def _payload_vazio_recebimento(ano: int, mes: int, inicio: date, fim: date, aviso: str = "") -> dict:
    dias = []
    cursor = inicio
    while cursor <= fim:
        dias.append(
            {
                "data": cursor.isoformat(),
                "qtd_total": 0,
                "qtd_coletas": 0,
                "qtd_entregas": 0,
                "qtd_sem_viagem": 0,
                "qtd_atrasadas": 0,
            }
        )
        cursor += timedelta(days=1)
    return {
        "visao": "mes",
        "modo_mensal_habilitado": False,
        "mes": f"{ano:04d}-{mes:02d}",
        "data_ref": inicio.isoformat(),
        "inicio": inicio.isoformat(),
        "fim": fim.isoformat(),
        "resumo": {
            "total": 0,
            "coletas": 0,
            "entregas": 0,
            "coletas_sem_viagem": 0,
            "atrasadas": 0,
            "pico_dia": 0,
            "dias_com_recebimento": 0,
        },
        "dias": dias,
        "eventos": [],
        "aviso": aviso,
    }


def _periodo_recebimento(visao: str, mes_ref: str, data_ref: str) -> tuple[date, date, date]:
    if visao == "mes":
        try:
            mes_base = datetime.strptime(mes_ref or datetime.now().strftime("%Y-%m"), "%Y-%m")
        except ValueError as exc:
            raise ValueError("Parâmetro 'mes' inválido. Use YYYY-MM.") from exc
        ano = mes_base.year
        mes = mes_base.month
        fim_mes = monthrange(ano, mes)[1]
        inicio = date(ano, mes, 1)
        fim = date(ano, mes, fim_mes)
        return inicio, fim, inicio

    texto_ref = str(data_ref or "").strip()
    if texto_ref:
        ref = _to_date(texto_ref)
        if ref is None:
            raise ValueError("Parâmetro 'data_ref' inválido. Use YYYY-MM-DD.")
    else:
        ref = date.today()
    inicio = ref - timedelta(days=ref.weekday())
    fim = inicio + timedelta(days=6)
    return inicio, fim, ref


@agendamento_bp.route("/api/logistica/recebimento/calendario")
@permission_required_any("PAGE_LOGISTICA_AGENDAMENTO", "PAGE_LOGISTICA_SOLICITACAO")
def recebimento_calendario_dados():
    visao_req = str(request.args.get("visao") or "").strip().lower()
    admin_mode = session.get("role") == "Admin"
    visao = "mes" if (visao_req == "mes" and admin_mode) else "semana"
    mes_ref = str(request.args.get("mes") or "").strip()
    data_ref = str(request.args.get("data_ref") or "").strip()
    try:
        inicio, fim, referencia = _periodo_recebimento(visao, mes_ref, data_ref)
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400

    ano = inicio.year
    mes = inicio.month
    hoje = date.today()
    cache_key = f"{visao}:{inicio.isoformat()}:{fim.isoformat()}"

    try:
        ocs = compras_service.ordens_compra_entregas(limite=5000)
        cif_rows = compras_service.ordens_compra_cif_recentes(janela_dias=730, limite=5000)
    except Exception:
        current_app.logger.exception("Falha ao consultar OCs para calendário de recebimento")
        cached = _RECEBIMENTO_CACHE.get(cache_key)
        if cached:
            resp = dict(cached)
            resp["aviso"] = ""
            resp["fonte_dados"] = "cache"
            return jsonify(resp)
        return jsonify(
            _payload_vazio_recebimento(
                ano,
                mes,
                inicio,
                fim,
                "Nao foi possivel consultar o ERP agora. Tente recarregar em alguns instantes.",
            )
        )
    cif_por_oc: dict[str, bool] = {}
    for row in cif_rows:
        oc_num = str(row.get("cod_ordem_compra") or "").strip()
        if not oc_num or oc_num in cif_por_oc:
            continue
        cif_por_oc[oc_num] = _oc_modalidade_cif(row.get("oc_json"))

    eventos_base: list[dict] = []
    for row in ocs:
        previsao = _to_date(row.get("previsao_entrega"))
        if not previsao or previsao < inicio or previsao > fim:
            continue
        numero_oc = str(row.get("cod_ordem_compra") or "").strip()
        if not numero_oc:
            continue
        is_coleta = bool(cif_por_oc.get(numero_oc, False))
        eventos_base.append(
            {
                "numero_oc": numero_oc,
                "fornecedor": str(row.get("fornecedor") or "").strip() or "(Sem fornecedor)",
                "status_oc_nome": str(row.get("status_oc_nome") or "").strip(),
                "previsao": previsao,
                "tipo_logistico": "COLETA" if is_coleta else "ENTREGA",
            }
        )

    ocs_mes = sorted({e["numero_oc"] for e in eventos_base})

    solicitacao_por_oc: dict[str, AgendamentoSolicitacao] = {}
    if ocs_mes:
        try:
            solicitacoes = (
                _query_solicitacoes_visiveis()
                .filter(AgendamentoSolicitacao.tipo == "COLETA")
                .filter(AgendamentoSolicitacao.numero_oc.in_(ocs_mes))
                .filter(AgendamentoSolicitacao.status != "Cancelada")
                .order_by(AgendamentoSolicitacao.id.desc())
                .all()
            )
            for sol in solicitacoes:
                oc_num = str(sol.numero_oc or "").strip()
                if oc_num and oc_num not in solicitacao_por_oc:
                    solicitacao_por_oc[oc_num] = sol
        except Exception:
            current_app.logger.exception("Falha ao consultar solicitacoes de coleta no calendário de recebimento")

    viagem_por_solicitacao: dict[int, dict] = {}
    if solicitacao_por_oc:
        ids = [int(sol.id) for sol in solicitacao_por_oc.values()]
        vinculos = (
            db.session.query(
                ViagemParada.solicitacao_id,
                ViagemParada.status,
                Viagem.id,
                Viagem.codigo,
                Viagem.status,
                Viagem.liberada,
            )
            .join(Viagem, Viagem.id == ViagemParada.viagem_id)
            .filter(ViagemParada.solicitacao_id.in_(ids))
            .filter(Viagem.status != "Cancelada")
            .order_by(Viagem.id.desc())
            .all()
        )
        for solicitacao_id, parada_status, viagem_id, viagem_codigo, viagem_status, viagem_liberada in vinculos:
            sid = int(solicitacao_id)
            if sid in viagem_por_solicitacao:
                continue
            viagem_por_solicitacao[sid] = {
                "id": int(viagem_id),
                "codigo": str(viagem_codigo or "").strip() or f"VG-{viagem_id}",
                "status": str(viagem_status or "").strip(),
                "parada_status": str(parada_status or "").strip(),
                "liberada": bool(viagem_liberada),
            }

    eventos: list[dict] = []
    dias: dict[str, dict] = {}
    total_coletas = total_entregas = sem_viagem = atrasadas = 0
    for row in eventos_base:
        previsao = row["previsao"]
        numero_oc = row["numero_oc"]
        status_oc_nome = row["status_oc_nome"]
        oc_recebida = _oc_ja_recebida(status_oc_nome)
        solicitacao = solicitacao_por_oc.get(numero_oc)
        viagem = viagem_por_solicitacao.get(int(solicitacao.id)) if solicitacao else None
        risco = "normal"
        if previsao < hoje:
            risco = "atrasado"
            atrasadas += 1
        elif previsao == hoje:
            risco = "hoje"
        elif previsao <= (hoje + timedelta(days=3)):
            risco = "proximos_3"

        tipo_logistico = row["tipo_logistico"]
        if tipo_logistico == "COLETA":
            total_coletas += 1
            if not viagem:
                sem_viagem += 1
        else:
            total_entregas += 1

        can_schedule = bool(tipo_logistico == "COLETA" and not viagem and not oc_recebida and _is_admin_or_compras())
        data_key = previsao.isoformat()

        cached_risco = _RECEBIMENTO_RISCO_CACHE.get(numero_oc) if numero_oc else None
        if isinstance(cached_risco, dict):
            risco_estoque = str(cached_risco.get("risco_estoque") or "sem_dados")
            risco_estoque_label = str(cached_risco.get("risco_estoque_label") or "Sem dados")
            risco_estoque_detalhe = str(
                cached_risco.get("risco_estoque_detalhe") or "Sem itens da OC ou sem saldo disponível para cálculo."
            )
        else:
            risco_estoque = "pendente"
            risco_estoque_label = "Calculando..."
            risco_estoque_detalhe = "Risco de estoque será carregado ao abrir o dia."
        dia_ref = dias.setdefault(
            data_key,
            {"data": data_key, "qtd_total": 0, "qtd_coletas": 0, "qtd_entregas": 0, "qtd_sem_viagem": 0, "qtd_atrasadas": 0},
        )
        dia_ref["qtd_total"] += 1
        if tipo_logistico == "COLETA":
            dia_ref["qtd_coletas"] += 1
        else:
            dia_ref["qtd_entregas"] += 1
        if tipo_logistico == "COLETA" and not viagem:
            dia_ref["qtd_sem_viagem"] += 1
        if risco == "atrasado":
            dia_ref["qtd_atrasadas"] += 1

        eventos.append(
            {
                "data": data_key,
                "numero_oc": numero_oc,
                "fornecedor": row["fornecedor"],
                "status_oc": status_oc_nome,
                "oc_recebida": oc_recebida,
                "oc_recebida_label": "Recebida" if oc_recebida else "Pendente",
                "tipo_logistico": tipo_logistico,
                "tipo_logistico_label": "Coleta (nossa frota)" if tipo_logistico == "COLETA" else "Entrega do fornecedor",
                "risco": risco,
                "risco_estoque": risco_estoque,
                "risco_estoque_label": risco_estoque_label,
                "risco_estoque_detalhe": risco_estoque_detalhe,
                "solicitacao": {
                    "id": int(solicitacao.id),
                    "codigo": str(solicitacao.codigo or f"LOG-{solicitacao.id}").strip(),
                    "status": str(solicitacao.status or "").strip(),
                } if solicitacao else None,
                "viagem": viagem,
                "pode_programar_coleta": can_schedule,
                "programar_hint": "Cria solicitação de coleta na Central de Viagens" if can_schedule else "",
            }
        )

    pico_dia = max((d["qtd_total"] for d in dias.values()), default=0)
    payload = {
        "visao": visao,
        "modo_mensal_habilitado": bool(admin_mode),
        "mes": f"{ano:04d}-{mes:02d}",
        "data_ref": referencia.isoformat(),
        "inicio": inicio.isoformat(),
        "fim": fim.isoformat(),
        "resumo": {
            "total": len(eventos),
            "coletas": total_coletas,
            "entregas": total_entregas,
            "coletas_sem_viagem": sem_viagem,
            "atrasadas": atrasadas,
            "pico_dia": pico_dia,
            "dias_com_recebimento": sum(1 for d in dias.values() if int(d.get("qtd_total") or 0) > 0),
        },
        "dias": sorted(dias.values(), key=lambda d: d["data"]),
        "eventos": sorted(eventos, key=lambda e: (e["data"], e["tipo_logistico"], e["numero_oc"])),
        "aviso": "",
    }
    _RECEBIMENTO_CACHE[cache_key] = dict(payload)
    return jsonify(payload)


@agendamento_bp.route("/api/logistica/recebimento/calendario/risco-estoque", methods=["POST"])
@permission_required_any("PAGE_LOGISTICA_AGENDAMENTO", "PAGE_LOGISTICA_SOLICITACAO")
def recebimento_calendario_risco_estoque():
    payload = request.get_json(silent=True) or {}
    ocs = payload.get("ocs") if isinstance(payload, dict) else None
    if not isinstance(ocs, list):
        return jsonify({"error": "Informe uma lista de OCs em 'ocs'."}), 400

    ocs_limpo = [str(oc or "").strip() for oc in ocs if str(oc or "").strip()][:300]
    riscos = _calcular_risco_estoque_por_ocs(ocs_limpo)
    return jsonify({"sucesso": True, "riscos": riscos})


@agendamento_bp.route("/api/logistica/recebimento/calendario/relatorio-dia")
@permission_required_any("PAGE_LOGISTICA_AGENDAMENTO", "PAGE_LOGISTICA_SOLICITACAO")
def recebimento_calendario_relatorio_dia():
    data_ref = str(request.args.get("data") or "").strip()
    alvo = _to_date(data_ref)
    if not alvo:
        return jsonify({"error": "Parâmetro 'data' inválido. Use YYYY-MM-DD."}), 400

    inicio = alvo
    fim = alvo
    ocs = compras_service.ordens_compra_entregas(limite=5000)
    cif_rows = compras_service.ordens_compra_cif_recentes(janela_dias=730, limite=5000)
    cif_por_oc: dict[str, bool] = {}
    for row in cif_rows:
        oc_num = str(row.get("cod_ordem_compra") or "").strip()
        if not oc_num or oc_num in cif_por_oc:
            continue
        cif_por_oc[oc_num] = _oc_modalidade_cif(row.get("oc_json"))

    eventos = []
    for row in ocs:
        previsao = _to_date(row.get("previsao_entrega"))
        if not previsao or previsao < inicio or previsao > fim:
            continue
        numero_oc = str(row.get("cod_ordem_compra") or "").strip()
        if not numero_oc:
            continue
        is_coleta = bool(cif_por_oc.get(numero_oc, False))
        eventos.append(
            {
                "data": previsao.isoformat(),
                "numero_oc": numero_oc,
                "fornecedor": str(row.get("fornecedor") or "").strip() or "(Sem fornecedor)",
                "status_oc": str(row.get("status_oc_nome") or "").strip(),
                "programacao": "Coleta" if is_coleta else "Entrega fornecedor",
            }
        )

    nome_arquivo = f"recebimento_{alvo.strftime('%Y%m%d')}.csv"
    stream = io.StringIO()
    writer = csv.writer(stream, delimiter=";")
    writer.writerow(["Data", "OC", "Fornecedor", "Programacao", "Status OC"])
    for evento in sorted(eventos, key=lambda e: e["numero_oc"]):
        writer.writerow(
            [
                evento.get("data") or "",
                evento.get("numero_oc") or "",
                evento.get("fornecedor") or "",
                evento.get("programacao") or "",
                evento.get("status_oc") or "",
            ]
        )

    payload = ("\ufeff" + stream.getvalue()).encode("utf-8")
    return send_file(
        io.BytesIO(payload),
        mimetype="text/csv; charset=utf-8",
        as_attachment=True,
        download_name=nome_arquivo,
    )


@agendamento_bp.route("/api/logistica/recebimento/calendario/programar-coleta", methods=["POST"])
@permission_required_any("PAGE_LOGISTICA_AGENDAMENTO", "PAGE_LOGISTICA_SOLICITACAO")
def recebimento_calendario_programar_coleta():
    if not _is_admin_or_compras():
        return jsonify({"error": "Somente Compras e Administrador podem programar coleta."}), 403

    payload = request.get_json(silent=True) or {}
    numero_oc = str(payload.get("numero_oc") or "").strip()
    if not numero_oc:
        return jsonify({"error": "Informe o número da OC."}), 400

    existente = (
        _query_solicitacoes_visiveis()
        .filter(
            AgendamentoSolicitacao.tipo == "COLETA",
            AgendamentoSolicitacao.numero_oc == numero_oc,
            AgendamentoSolicitacao.status != "Cancelada",
        )
        .first()
    )
    if existente:
        return jsonify(
            {
                "sucesso": True,
                "ja_existia": True,
                "mensagem": f"A OC {numero_oc} já está na Central ({existente.codigo or ('LOG-' + str(existente.id))}).",
                "solicitacao_id": int(existente.id),
                "central_url": f"/logistica/viagens?tab=COLETA&q={numero_oc}",
            }
        )

    consulta = consultar_oc_agendamento(numero_oc)
    if not consulta.get("encontrada"):
        return jsonify({"error": consulta.get("error") or "OC não encontrada na bridge."}), 404

    parceiro = consulta.get("fornecedor") if isinstance(consulta.get("fornecedor"), dict) else {}
    parceiro_payload = {
        "codigo": str(parceiro.get("codigo") or "").strip(),
        "nome": str(parceiro.get("nome") or parceiro.get("razao_social") or "").strip(),
        "razao_social": str(parceiro.get("razao_social") or parceiro.get("nome") or "").strip(),
        "cnpj_cpf": str(parceiro.get("cnpj_cpf") or "").strip(),
        "contato": str(parceiro.get("contato") or "").strip(),
        "telefone": str(parceiro.get("telefone") or "").strip(),
        "email": str(parceiro.get("email") or "").strip(),
        "logradouro": str(parceiro.get("logradouro") or "").strip() or "Endereço de coleta a confirmar",
        "numero": str(parceiro.get("numero") or "").strip(),
        "complemento": str(parceiro.get("complemento") or "").strip(),
        "bairro": str(parceiro.get("bairro") or "").strip(),
        "cidade": str(parceiro.get("cidade") or "").strip() or "A confirmar",
        "uf": (str(parceiro.get("uf") or "").strip()[:2] or "SP").upper(),
        "cep": str(parceiro.get("cep") or "").strip(),
        "observacoes": str(consulta.get("warning") or "").strip(),
    }

    usuario = session.get("username", "sistema")
    agora = datetime.now()
    fonte = consulta.get("fonte") if isinstance(consulta.get("fonte"), dict) else {}
    sol = AgendamentoSolicitacao(
        tipo="COLETA",
        status="Pendente",
        prioridade="Media",
        prazo_limite=None,
        data_desejada=None,
        solicitante=usuario,
        criado_em=agora,
        atualizado_em=agora,
        documento_tipo="OC",
        documento_numero=numero_oc,
        numero_oc=numero_oc,
        origem_documento="ORDEM_DE_COMPRA",
        observacoes_solicitante="Coleta gerada automaticamente pelo calendário de recebimento.",
        observacoes_logistica=(f"Bridge: {fonte.get('label')}" if fonte.get("label") else "Gerada via bridge de compras."),
        payload_origem=_json_text(
            {
                "origem": "calendario_recebimento",
                "numero_oc": numero_oc,
                "fonte": fonte,
                "warning": consulta.get("warning") or "",
            }
        ),
    )
    ok, msg = _aplicar_parceiro(sol, parceiro_payload, "Fornecedor")
    if not ok:
        return jsonify({"error": msg or "Não foi possível aplicar os dados do fornecedor retornados pela bridge."}), 409

    db.session.add(sol)
    db.session.flush()
    sol.codigo = f"LOG-{agora.strftime('%Y%m%d')}-{sol.id:04d}"

    itens = consulta.get("itens") if isinstance(consulta.get("itens"), list) else []
    if not _sincronizar_itens(sol, itens):
        _sincronizar_itens(
            sol,
            [
                {
                    "descricao": f"Coleta referente à OC {numero_oc}",
                    "quantidade": 1,
                    "unidade": "UN",
                    "volumes": 1,
                    "observacoes": "Item padrão criado automaticamente.",
                }
            ],
        )

    _registrar_historico(
        sol.id,
        evento="CRIADA_OC_BRIDGE",
        usuario=usuario,
        status_novo="Pendente",
        detalhe=f"Solicitação criada a partir da OC {numero_oc} via calendário de recebimento.",
        payload={"numero_oc": numero_oc, "fonte": fonte},
    )
    db.session.commit()
    return jsonify(
        {
            "sucesso": True,
            "ja_existia": False,
            "mensagem": f"Coleta da OC {numero_oc} criada na Central.",
            "solicitacao_id": int(sol.id),
            "solicitacao_codigo": str(sol.codigo or "").strip(),
            "central_url": f"/logistica/viagens?tab=COLETA&q={numero_oc}",
        }
    )


def _is_origem_automatica(row: AgendamentoSolicitacao) -> bool:
    origem = str(row.origem_documento or "").strip()
    if str(row.tipo or "").strip() == "COLETA":
        return origem in {"ORDEM_DE_COMPRA", "AutoCIF"}
    if str(row.tipo or "").strip() == "ENTREGA":
        return origem in {"AutoCIF", "ROMANEIO"}
    return False


def _filtro_solicitacao_visivel():
    return true()


def _query_solicitacoes_visiveis():
    return AgendamentoSolicitacao.query.filter(_filtro_solicitacao_visivel())


def _get_solicitacao_visivel(solicitacao_id: int) -> AgendamentoSolicitacao | None:
    return _query_solicitacoes_visiveis().filter(AgendamentoSolicitacao.id == solicitacao_id).first()


def _normalizar_cabecalho(texto: str) -> str:
    return re.sub(r"[^a-z0-9]", "", str(texto or "").strip().lower())


def _parse_float_br(valor) -> float | None:
    texto = str(valor or "").strip()
    if not texto:
        return None
    texto = texto.replace(".", "").replace(",", ".") if "," in texto else texto
    try:
        return float(texto)
    except (TypeError, ValueError):
        return None


def _parse_data_livre(valor) -> datetime | None:
    if isinstance(valor, datetime):
        return valor
    texto = str(valor or "").strip()
    if not texto:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d/%m/%Y %H:%M", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(texto[:19], fmt)
        except ValueError:
            continue
    return None


def _extrair_linhas_oc_upload(file_storage) -> list[dict]:
    nome = str(getattr(file_storage, "filename", "") or "").strip()
    ext = nome.rsplit(".", 1)[-1].lower() if "." in nome else ""
    if ext not in {"csv", "xlsx", "xls"}:
        raise ValueError("Formato inválido. Envie arquivo XLSX, XLS ou CSV.")

    rows: list[list] = []
    if ext == "csv":
        raw = file_storage.read()
        try:
            file_storage.seek(0)
        except Exception:
            pass
        text = raw.decode("utf-8-sig", errors="ignore")
        reader = csv.reader(io.StringIO(text), delimiter=";")
        rows = [list(r) for r in reader]
        if len(rows) <= 1 or len(rows[0]) == 1:
            reader = csv.reader(io.StringIO(text), delimiter=",")
            rows = [list(r) for r in reader]
    else:
        if load_workbook is None:
            raise ValueError("Dependência openpyxl indisponível para leitura de XLSX/XLS.")
        data = file_storage.read()
        try:
            file_storage.seek(0)
        except Exception:
            pass
        wb = load_workbook(filename=io.BytesIO(data), read_only=True, data_only=True)
        try:
            ws = wb.active
            rows = [list(r) for r in ws.iter_rows(values_only=True)]
        finally:
            wb.close()

    if not rows:
        raise ValueError("Arquivo vazio.")

    header = [_normalizar_cabecalho(c) for c in rows[0]]
    idx = {h: i for i, h in enumerate(header) if h}

    def col(*aliases):
        for alias in aliases:
            if alias in idx:
                return idx[alias]
        return None

    c_oc = col("oc", "numerooc", "ordemdecompra", "ordemcompra", "nrooc")
    c_forn = col("fornecedor", "nomefornecedor", "razaosocial", "parceiro")
    c_cid = col("cidade", "cidadeorigem")
    c_uf = col("uf", "estado")
    c_val = col("valor", "valoroc", "valortotal", "total")
    c_data = col("dataprevista", "previsao", "previsaodeentrega", "dataentrega")
    c_transp = col("transportadora", "transporte")
    c_end = col("endereco", "logradouro", "localcoleta")

    parsed: list[dict] = []
    for line_no, row in enumerate(rows[1:], start=2):
        if not row or not any(str(v or "").strip() for v in row):
            continue
        numero_oc = str(row[c_oc] if c_oc is not None and c_oc < len(row) else "").strip()
        fornecedor = str(row[c_forn] if c_forn is not None and c_forn < len(row) else "").strip()
        cidade = str(row[c_cid] if c_cid is not None and c_cid < len(row) else "").strip()
        uf = str(row[c_uf] if c_uf is not None and c_uf < len(row) else "").strip().upper()[:2]
        valor_raw = row[c_val] if c_val is not None and c_val < len(row) else ""
        data_raw = row[c_data] if c_data is not None and c_data < len(row) else ""
        transportadora = str(row[c_transp] if c_transp is not None and c_transp < len(row) else "").strip()
        endereco = str(row[c_end] if c_end is not None and c_end < len(row) else "").strip()

        valor = _parse_float_br(valor_raw)
        data_prevista = _parse_data_livre(data_raw)
        parsed.append(
            {
                "line": line_no,
                "numero_oc": numero_oc,
                "fornecedor": fornecedor,
                "cidade": cidade,
                "uf": uf,
                "valor": valor,
                "valor_raw": str(valor_raw or "").strip(),
                "data_prevista": data_prevista,
                "data_prevista_raw": str(data_raw or "").strip(),
                "transportadora": transportadora,
                "endereco": endereco,
            }
        )
    return parsed


def _validar_linhas_oc(rows: list[dict]) -> tuple[list[dict], list[str]]:
    validas: list[dict] = []
    erros: list[str] = []
    for row in rows:
        faltas = []
        if not row.get("numero_oc"):
            faltas.append("Número da OC")
        if not row.get("fornecedor"):
            faltas.append("Fornecedor")
        if not row.get("cidade"):
            faltas.append("Cidade")
        if not row.get("uf"):
            faltas.append("UF")
        if row.get("valor") is None:
            faltas.append("Valor")
        if row.get("data_prevista") is None:
            faltas.append("Data Prevista")
        if faltas:
            erros.append(f"Linha {row.get('line')}: faltando/invalidos -> {', '.join(faltas)}")
            continue
        validas.append(row)
    return validas, erros


def _kanban_coluna(registro: AgendamentoSolicitacao, veiculo: AgendamentoVeiculo | None) -> str:
    status = str(registro.status or "").strip()
    if status == "Concluida":
        return "concluido"
    if status == "Cancelada":
        return "cancelada"
    if status in {"Alocada", "EmAndamento", "EmRota"} and veiculo and str(veiculo.codigo or "").strip().upper() in VEICULOS_KANBAN:
        return str(veiculo.codigo or "").strip().upper()
    return "pendentes"


def _intervalo_planejado(
    registro: AgendamentoSolicitacao,
    *,
    veiculo: AgendamentoVeiculo | None = None,
    inicio_override: datetime | None = None,
    fim_override: datetime | None = None,
) -> tuple[datetime | None, datetime | None]:
    inicio = inicio_override or registro.data_hora_saida_prevista
    if not inicio:
        return None, None
    fim = fim_override or registro.data_hora_retorno_prevista
    if fim and fim > inicio:
        return inicio, fim
    veiculo_ref = veiculo or (AgendamentoVeiculo.query.get(registro.veiculo_id) if registro.veiculo_id else None)
    duracao = int(
        getattr(veiculo_ref, "duracao_padrao_min", 0)
        or current_app.config.get("AGENDAMENTO_DURACAO_PADRAO_MINUTOS", 120)
    )
    return inicio, inicio + timedelta(minutes=max(30, duracao))


def _serializar_item(item: AgendamentoSolicitacaoItem) -> dict:
    return {
        "id": item.id,
        "sequencia": int(item.sequencia or 0),
        "codigo_item": str(item.codigo_item or "").strip(),
        "descricao": str(item.descricao or "").strip(),
        "quantidade": float(item.quantidade or 0.0),
        "unidade": str(item.unidade or "").strip(),
        "volumes": float(item.volumes or 0.0),
        "observacoes": str(item.observacoes or "").strip(),
    }


def _serializar_historico(item: AgendamentoSolicitacaoHistorico) -> dict:
    return {
        "id": item.id,
        "evento": str(item.evento or "").strip(),
        "status_anterior": str(item.status_anterior or "").strip(),
        "status_novo": str(item.status_novo or "").strip(),
        "usuario": str(item.usuario or "").strip(),
        "detalhe": str(item.detalhe or "").strip(),
        "criado_em": item.criado_em.strftime("%d/%m/%Y %H:%M") if item.criado_em else "",
    }


def _serializar_solicitacao(
    registro: AgendamentoSolicitacao,
    *,
    veiculo: AgendamentoVeiculo | None = None,
    itens: list[AgendamentoSolicitacaoItem] | None = None,
    historico: list[AgendamentoSolicitacaoHistorico] | None = None,
) -> dict:
    payload_origem = {}
    if registro.payload_origem:
        try:
            payload_origem = json.loads(registro.payload_origem)
        except Exception:
            payload_origem = {}
    anexo = _extrair_anexo_payload(payload_origem)
    endereco = {
        "logradouro": str(registro.logradouro or "").strip(),
        "numero": str(registro.numero or "").strip(),
        "complemento": str(registro.complemento or "").strip(),
        "bairro": str(registro.bairro or "").strip(),
        "cidade": str(registro.cidade or "").strip(),
        "uf": str(registro.uf or "").strip(),
        "cep": str(registro.cep or "").strip(),
        "observacoes": str(registro.observacoes_endereco or "").strip(),
    }
    prazo = registro.prazo_limite
    atrasada = bool(prazo and prazo < datetime.now() and str(registro.status or "").strip() not in {"Concluida", "Cancelada"})
    return {
        "id": registro.id,
        "codigo": str(registro.codigo or f"LOG-{registro.id}").strip(),
        "tipo": str(registro.tipo or "").strip(),
        "tipo_label": {"COLETA": "Coleta", "ENTREGA": "Entrega", "AVULSA": "Avulsa"}.get(str(registro.tipo or "").strip(), str(registro.tipo or "").strip()),
        "status": str(registro.status or "").strip(),
        "status_label": status_label_agendamento(registro.status),
        "prioridade": str(registro.prioridade or "").strip(),
        "prioridade_label": prioridade_label_agendamento(registro.prioridade),
        "solicitante": str(registro.solicitante or "").strip(),
        "criado_em": registro.criado_em.strftime("%d/%m/%Y %H:%M") if registro.criado_em else "",
        "documento_tipo": str(registro.documento_tipo or "").strip(),
        "documento_numero": str(registro.documento_numero or "").strip(),
        "numero_oc": str(registro.numero_oc or "").strip(),
        "numero_nf": str(registro.numero_nf or "").strip(),
        "orcamento": str(getattr(registro, "orcamento", "") or "").strip(),
        "romaneio_numero": str(payload_origem.get("romaneio_numero") or "").strip(),
        "origem_documento": str(registro.origem_documento or "").strip(),
        "origem_documento_label": _origem_documento_label(registro.origem_documento),
        "parceiro_tipo": str(registro.parceiro_tipo or "").strip(),
        "parceiro_codigo": str(registro.parceiro_codigo or "").strip(),
        "parceiro_nome": str(registro.parceiro_nome or "").strip(),
        "parceiro_razao_social": str(registro.parceiro_razao_social or "").strip(),
        "parceiro_documento": str(registro.parceiro_documento or "").strip(),
        "contato": str(registro.contato or "").strip(),
        "telefone": str(registro.telefone or "").strip(),
        "email": str(registro.email or "").strip(),
        "endereco": endereco,
        "endereco_formatado": formatar_endereco_logistico(endereco),
        "observacoes_solicitante": str(registro.observacoes_solicitante or "").strip(),
        "observacoes_logistica": str(registro.observacoes_logistica or "").strip(),
        "motivo_cancelamento": str(registro.motivo_cancelamento or "").strip(),
        "data_desejada": registro.data_desejada.isoformat(timespec="minutes") if registro.data_desejada else "",
        "data_desejada_label": registro.data_desejada.strftime("%d/%m/%Y %H:%M") if registro.data_desejada else "",
        "cancelamento_pendente": bool(registro.cancelamento_pendente),
        "cancelamento_solicitado_por": str(registro.cancelamento_solicitado_por or "").strip(),
        "cancelamento_motivo_pendente": str(registro.cancelamento_motivo_pendente or "").strip(),
        "departamento_solicitante": str(registro.departamento_solicitante or "").strip(),
        "tempo_estimado_min": int(registro.tempo_estimado_min) if registro.tempo_estimado_min else None,
        "prazo_limite": prazo.isoformat(timespec="minutes") if prazo else "",
        "prazo_limite_label": prazo.strftime("%d/%m/%Y %H:%M") if prazo else "",
        "data_hora_saida_prevista": registro.data_hora_saida_prevista.isoformat(timespec="minutes") if registro.data_hora_saida_prevista else "",
        "data_hora_saida_prevista_label": registro.data_hora_saida_prevista.strftime("%d/%m/%Y %H:%M") if registro.data_hora_saida_prevista else "",
        "data_hora_retorno_prevista": registro.data_hora_retorno_prevista.isoformat(timespec="minutes") if registro.data_hora_retorno_prevista else "",
        "data_hora_retorno_prevista_label": registro.data_hora_retorno_prevista.strftime("%d/%m/%Y %H:%M") if registro.data_hora_retorno_prevista else "",
        "data_hora_saida_real": registro.data_hora_saida_real.isoformat(timespec="minutes") if registro.data_hora_saida_real else "",
        "data_hora_saida_real_label": registro.data_hora_saida_real.strftime("%d/%m/%Y %H:%M") if registro.data_hora_saida_real else "",
        "data_hora_retorno_real": registro.data_hora_retorno_real.isoformat(timespec="minutes") if registro.data_hora_retorno_real else "",
        "data_hora_retorno_real_label": registro.data_hora_retorno_real.strftime("%d/%m/%Y %H:%M") if registro.data_hora_retorno_real else "",
        "qtd_itens": int(registro.qtd_itens or 0),
        "qtd_volumes": float(registro.qtd_volumes or 0.0),
        "resumo_itens": str(registro.resumo_itens or "").strip(),
        "motorista": {
            "id": registro.motorista_id,
            "nome": str(registro.motorista_nome or "").strip(),
        } if registro.motorista_id or str(registro.motorista_nome or "").strip() else None,
        "km_estimado": float(registro.km_estimado) if registro.km_estimado is not None else None,
        "km_estimado_retorno": float(registro.km_estimado_retorno) if registro.km_estimado_retorno is not None else None,
        "waze_url": montar_waze_url_agendamento(
            {
                **endereco,
                "latitude": registro.destino_latitude,
                "longitude": registro.destino_longitude,
            }
        ),
        "google_maps_url": montar_waze_url_agendamento(
            {
                **endereco,
                "latitude": registro.destino_latitude,
                "longitude": registro.destino_longitude,
            }
        ),
        "veiculo": {
            "id": veiculo.id,
            "codigo": str(veiculo.codigo or "").strip(),
            "nome": str(veiculo.nome_exibicao or veiculo.codigo or "").strip(),
            "cor": str(veiculo.cor_kanban or "").strip(),
        } if veiculo else None,
        "kanban_coluna": _kanban_coluna(registro, veiculo),
        "atrasada": atrasada,
        "itens": [_serializar_item(item) for item in itens or []],
        "historico": [_serializar_historico(item) for item in historico or []],
        "anexo": {
            "nome_original": str(anexo.get("nome_original") or "").strip(),
            "nome_arquivo": str(anexo.get("nome_arquivo") or "").strip(),
            "caminho_relativo": str(anexo.get("caminho_relativo") or "").strip(),
            "mimetype": str(anexo.get("mimetype") or "").strip(),
            "url": f"/api/logistica/central-viagens/solicitacoes/{registro.id}/anexo" if anexo.get("caminho_relativo") else "",
        },
    }


def _ordenar_cards(cards: list[dict]) -> list[dict]:
    return sorted(
        cards,
        key=lambda card: (
            PRIORIDADE_ORDEM.get(str(card.get("prioridade") or ""), 99),
            str(card.get("prazo_limite") or "9999-12-31T23:59"),
            str(card.get("data_hora_saida_prevista") or "9999-12-31T23:59"),
            str(card.get("codigo") or ""),
        ),
    )


@agendamento_bp.route("/api/logistica/agendamento-veiculos/dashboard")
@permission_required("PAGE_LOGISTICA_AGENDAMENTO")
def dashboard_agendamento_veiculos():
    ensure_cadastros_base_carregados()
    sincronizar_motoristas_usuarios(commit=True)
    _auto_transicao_em_andamento()
    termo = str(request.args.get("q") or "").strip()
    status = str(request.args.get("status") or "").strip()
    tipo = str(request.args.get("tipo") or "").strip().upper()
    prioridade = str(request.args.get("prioridade") or "").strip()
    incluir_canceladas = str(request.args.get("incluir_canceladas") or "").strip().lower() in {"1", "true", "sim"}

    query = _query_solicitacoes_visiveis()
    if not incluir_canceladas:
        query = query.filter(AgendamentoSolicitacao.status != "Cancelada")
    if status:
        query = query.filter(AgendamentoSolicitacao.status == status)
    if tipo:
        query = query.filter(AgendamentoSolicitacao.tipo == tipo)
    if prioridade:
        query = query.filter(AgendamentoSolicitacao.prioridade == prioridade)
    if termo:
        like = f"%{termo}%"
        query = query.filter(
            or_(
                AgendamentoSolicitacao.codigo.ilike(like),
                AgendamentoSolicitacao.documento_numero.ilike(like),
                AgendamentoSolicitacao.numero_oc.ilike(like),
                AgendamentoSolicitacao.numero_nf.ilike(like),
                AgendamentoSolicitacao.orcamento.ilike(like),
                AgendamentoSolicitacao.parceiro_nome.ilike(like),
                AgendamentoSolicitacao.parceiro_razao_social.ilike(like),
                AgendamentoSolicitacao.parceiro_documento.ilike(like),
                AgendamentoSolicitacao.solicitante.ilike(like),
                AgendamentoSolicitacao.cidade.ilike(like),
            )
        )

    veiculos = {row.id: row for row in listar_veiculos_agendamento()}
    motoristas = listar_motoristas_agendamento()
    cards = [_serializar_solicitacao(row, veiculo=veiculos.get(row.veiculo_id)) for row in query.order_by(AgendamentoSolicitacao.criado_em.desc()).all()]
    colunas = {"pendentes": [], "IVECO": [], "SAVEIRO": [], "concluido": []}
    for card in cards:
        coluna = card.get("kanban_coluna")
        if coluna in colunas:
            colunas[coluna].append(card)
    return jsonify(
        {
            "resumo": {
                "total": len(cards),
                "pendentes": sum(1 for card in cards if card.get("status") in {"Pendente", "EmAnalise"}),
                "alocadas": sum(1 for card in cards if card.get("status") == "Alocada"),
                "em_andamento": sum(1 for card in cards if card.get("status") == "EmAndamento"),
                "em_rota": sum(1 for card in cards if card.get("status") == "EmRota"),
                "concluidas": sum(1 for card in cards if card.get("status") == "Concluida"),
                "cancelamentos_pendentes": sum(1 for card in cards if card.get("cancelamento_pendente")),
                "atrasadas": sum(1 for card in cards if card.get("atrasada")),
                "motoristas_ativos": len(motoristas),
            },
            "veiculos": [
                {
                    "id": row.id,
                    "codigo": str(row.codigo or "").strip(),
                    "nome": str(row.nome_exibicao or row.codigo or "").strip(),
                    "cor": str(row.cor_kanban or "").strip(),
                    "janela_conflito_min": int(row.janela_conflito_min or 0),
                    "duracao_padrao_min": int(row.duracao_padrao_min or 0),
                }
                for row in veiculos.values()
            ],
            "motoristas": motoristas,
            "cadastros": resumo_cadastros(),
            "colunas": {key: _ordenar_cards(value) for key, value in colunas.items()},
            "filtros": {"status": list(STATUS_SOLICITACAO), "tipos": list(TIPOS_SOLICITACAO), "prioridades": list(PRIORIDADES_SOLICITACAO)},
        }
    )


@agendamento_bp.route("/api/logistica/central-viagens/dashboard")
@permission_required("PAGE_LOGISTICA_AGENDAMENTO")
def dashboard_central_viagens():
    termo = str(request.args.get("q") or "").strip().lower()
    status = str(request.args.get("status") or "").strip()

    query = _query_solicitacoes_visiveis().filter(AgendamentoSolicitacao.tipo.in_(["COLETA", "ENTREGA"]))
    if status:
        query = query.filter(AgendamentoSolicitacao.status == status)
    rows = query.order_by(AgendamentoSolicitacao.criado_em.desc()).limit(800).all()

    # Regra operacional: coleta na Central deve representar apenas fluxo de OC
    # (importada/gerada pelo Compras). Entregas continuam no fluxo automatico
    # por romaneio/expedicao.
    automaticas = []
    for row in rows:
        tipo = str(row.tipo or "").strip()
        status_row = str(row.status or "").strip()
        origem = str(row.origem_documento or "").strip()

        if tipo == "ENTREGA":
            automaticas.append(row)
            continue

        if tipo != "COLETA":
            continue

        # Regra operacional:
        # - pendentes de coleta devem vir apenas do fluxo de OC
        # - se a coleta ja estiver alocada/em rota/em andamento/concluida,
        #   ela precisa continuar visivel na Central para gestao.
        if status_row in {"Alocada", "EmRota", "EmAndamento", "Concluida"}:
            automaticas.append(row)
            continue
        if origem == "ORDEM_DE_COMPRA":
            automaticas.append(row)
    if termo:
        def match(row: AgendamentoSolicitacao) -> bool:
            haystack = " ".join(
                [
                    str(row.numero_oc or ""),
                    str(row.numero_nf or ""),
                    str(row.documento_numero or ""),
                    str(row.parceiro_nome or ""),
                    str(row.cidade or ""),
                    str(row.uf or ""),
                    str(row.status or ""),
                    str(row.observacoes_logistica or ""),
                ]
            ).lower()
            return termo in haystack

        automaticas = [row for row in automaticas if match(row)]

    veiculos = {row.id: row for row in listar_veiculos_agendamento()}
    motoristas = listar_motoristas_agendamento()
    cards = [_serializar_solicitacao(row, veiculo=veiculos.get(row.veiculo_id)) for row in automaticas]

    viagem_por_solicitacao: dict[int, dict] = {}
    try:
        ids_solicitacao = [int(c.get("id")) for c in cards if c.get("id") is not None]
        if ids_solicitacao:
            vinculos = (
                db.session.query(
                    ViagemParada.solicitacao_id,
                    ViagemParada.status,
                    Viagem.id,
                    Viagem.codigo,
                    Viagem.status,
                    Viagem.liberada,
                )
                .join(Viagem, Viagem.id == ViagemParada.viagem_id)
                .filter(ViagemParada.solicitacao_id.in_(ids_solicitacao))
                .filter(Viagem.status != "Cancelada")
                .order_by(Viagem.id.desc())
                .all()
            )
            for solicitacao_id, parada_status, viagem_id, viagem_codigo, viagem_status, viagem_liberada in vinculos:
                sid = int(solicitacao_id)
                if sid in viagem_por_solicitacao:
                    continue
                viagem_por_solicitacao[sid] = {
                    "id": int(viagem_id),
                    "codigo": str(viagem_codigo or "").strip() or f"VG-{viagem_id}",
                    "status": str(viagem_status or "").strip(),
                    "parada_status": str(parada_status or "").strip(),
                    "liberada": bool(viagem_liberada),
                }
    except Exception:
        # Compatibilidade defensiva: em ambientes legados, não quebrar dashboard.
        viagem_por_solicitacao = {}

    rows_por_id = {int(r.id): r for r in automaticas}
    alterou_status = False

    for card in cards:
        sid = int(card.get("id")) if card.get("id") is not None else None
        viagem = viagem_por_solicitacao.get(sid) if sid is not None else None
        card["viagem"] = viagem
        card["viagem_codigo"] = viagem.get("codigo") if viagem else ""
        if viagem and card.get("status") in {"Pendente", "EmAnalise", "Alocada", "EmAndamento", "EmRota"}:
            status_viagem = str(viagem.get("status") or "").strip()
            status_parada = str(viagem.get("parada_status") or "").strip()
            status_novo = None
            if status_viagem == "Concluida":
                if status_parada == "Nao_realizada":
                    status_novo = "Pendente"
                else:
                    status_novo = "Concluida"
            elif status_viagem == "EmAndamento":
                status_novo = "EmRota"
            elif status_viagem == "Planejada" and card.get("status") in {"Pendente", "EmAnalise", "Alocada"}:
                status_novo = "Alocada"

            if status_novo and status_novo != card.get("status"):
                card["status"] = status_novo
                card["status_label"] = status_label_agendamento(status_novo)

                if sid is not None and sid in rows_por_id:
                    row = rows_por_id[sid]
                    row.status = status_novo
                    row.atualizado_em = datetime.now()
                    alterou_status = True

    if alterou_status:
        db.session.commit()

    coletas = [c for c in cards if c.get("tipo") == "COLETA"]
    entregas = [c for c in cards if c.get("tipo") == "ENTREGA"]

    def _pendentes(arr: list[dict]) -> int:
        return sum(1 for c in arr if c.get("status") in {"Pendente", "EmAnalise", "Alocada"})

    resumo = {
        "coletas_pendentes": _pendentes(coletas),
        "entregas_pendentes": _pendentes(entregas),
        "em_andamento": sum(1 for c in cards if c.get("status") in {"EmAndamento", "EmRota", "Alocada"}),
        "finalizadas": sum(1 for c in cards if c.get("status") == "Concluida"),
        "canceladas": sum(1 for c in cards if c.get("status") == "Cancelada"),
        "total": len(cards),
    }

    return jsonify(
        {
            "resumo": resumo,
            "coletas": coletas,
            "entregas": entregas,
            "counts": {"coletas": len(coletas), "entregas": len(entregas)},
            "veiculos": [
                {
                    "id": row.id,
                    "codigo": str(row.codigo or "").strip(),
                    "nome": str(row.nome_exibicao or row.codigo or "").strip(),
                }
                for row in veiculos.values()
            ],
            "motoristas": motoristas,
        }
    )


@agendamento_bp.route("/api/logistica/central-viagens/oc/<numero_oc>")
@permission_required_any("PAGE_LOGISTICA_AGENDAMENTO", "PAGE_LOGISTICA_SOLICITACAO")
def central_viagens_consultar_oc(numero_oc: str):
    numero_oc_limpo = str(numero_oc or "").strip()
    if not numero_oc_limpo:
        return jsonify({"error": "Informe o número da OC."}), 400
    resultado = consultar_oc_agendamento(numero_oc_limpo)
    if not resultado.get("encontrada"):
        return jsonify(resultado), 404

    parceiro = resultado.get("fornecedor") if isinstance(resultado.get("fornecedor"), dict) else {}
    return jsonify(
        {
            "encontrada": True,
            "numero_oc": resultado.get("numero_oc") or numero_oc_limpo,
            "fornecedor": parceiro,
            "itens": resultado.get("itens") or [],
            "warning": resultado.get("warning") or "",
            "fonte": resultado.get("fonte") or {},
        }
    )


@agendamento_bp.route("/api/logistica/central-viagens/oc/criar", methods=["POST"])
@permission_required_any("PAGE_LOGISTICA_AGENDAMENTO", "PAGE_LOGISTICA_SOLICITACAO")
def central_viagens_criar_coleta_por_oc():
    if not _is_admin_or_compras():
        return jsonify({"error": "Somente Compras e Administrador podem gerar coleta por OC."}), 403

    payload = request.get_json(silent=True) or {}
    form = request.form if request.form else {}
    numero_oc = str((form.get("numero_oc") if form else None) or payload.get("numero_oc") or "").strip()
    prioridade = str((form.get("prioridade") if form else None) or payload.get("prioridade") or "Media").strip()
    if prioridade not in PRIORIDADES_SOLICITACAO:
        prioridade = "Media"
    if not numero_oc:
        return jsonify({"error": "Informe o número da OC."}), 400

    existente = (
        _query_solicitacoes_visiveis()
        .filter(
            AgendamentoSolicitacao.tipo == "COLETA",
            AgendamentoSolicitacao.numero_oc == numero_oc,
            AgendamentoSolicitacao.status != "Cancelada",
        )
        .first()
    )
    if existente:
        return jsonify({"error": f"A OC {numero_oc} já possui coleta aberta na Central."}), 409

    consulta = consultar_oc_agendamento(numero_oc)
    if not consulta.get("encontrada"):
        return jsonify({"error": consulta.get("error") or "OC não encontrada na bridge."}), 404

    parceiro = consulta.get("fornecedor") if isinstance(consulta.get("fornecedor"), dict) else {}
    parceiro_payload = {
        "codigo": str(parceiro.get("codigo") or "").strip(),
        "nome": str(parceiro.get("nome") or parceiro.get("razao_social") or "").strip(),
        "razao_social": str(parceiro.get("razao_social") or parceiro.get("nome") or "").strip(),
        "cnpj_cpf": str(parceiro.get("cnpj_cpf") or "").strip(),
        "contato": str(parceiro.get("contato") or "").strip(),
        "telefone": str(parceiro.get("telefone") or "").strip(),
        "email": str(parceiro.get("email") or "").strip(),
        "logradouro": str(parceiro.get("logradouro") or "").strip() or "Endereço de coleta a confirmar",
        "numero": str(parceiro.get("numero") or "").strip(),
        "complemento": str(parceiro.get("complemento") or "").strip(),
        "bairro": str(parceiro.get("bairro") or "").strip(),
        "cidade": str(parceiro.get("cidade") or "").strip() or "A confirmar",
        "uf": (str(parceiro.get("uf") or "").strip()[:2] or "SP").upper(),
        "cep": str(parceiro.get("cep") or "").strip(),
        "observacoes": str(consulta.get("warning") or "").strip(),
    }

    usuario = session.get("username", "sistema")
    agora = datetime.now()
    fonte = consulta.get("fonte") if isinstance(consulta.get("fonte"), dict) else {}
    sol = AgendamentoSolicitacao(
        tipo="COLETA",
        status="Pendente",
        prioridade=prioridade,
        prazo_limite=None,
        data_desejada=None,
        solicitante=usuario,
        criado_em=agora,
        atualizado_em=agora,
        documento_tipo="OC",
        documento_numero=numero_oc,
        numero_oc=numero_oc,
        origem_documento="ORDEM_DE_COMPRA",
        observacoes_solicitante="Coleta gerada automaticamente pela Central de Viagens a partir de OC digitada.",
        observacoes_logistica=(f"Bridge: {fonte.get('label')}" if fonte.get("label") else "Gerada via bridge de compras."),
        payload_origem=_json_text(
            {
                "origem": "oc_digitada_central",
                "numero_oc": numero_oc,
                "fonte": fonte,
                "warning": consulta.get("warning") or "",
            }
        ),
    )

    ok, msg = _aplicar_parceiro(sol, parceiro_payload, "Fornecedor")
    if not ok:
        return jsonify({"error": msg or "Não foi possível aplicar os dados do fornecedor retornados pela bridge."}), 409

    db.session.add(sol)
    db.session.flush()
    sol.codigo = f"LOG-{agora.strftime('%Y%m%d')}-{sol.id:04d}"

    itens = consulta.get("itens") if isinstance(consulta.get("itens"), list) else []
    if not _sincronizar_itens(sol, itens):
        _sincronizar_itens(
            sol,
            [
                {
                    "descricao": f"Coleta referente à OC {numero_oc}",
                    "quantidade": 1,
                    "unidade": "UN",
                    "volumes": 1,
                    "observacoes": "Item padrão criado automaticamente.",
                }
            ],
        )

    anexo_meta = _salvar_anexo_solicitacao(request.files.get("anexo"), sol.id)
    if anexo_meta:
        _atualizar_payload_origem_anexo(sol, anexo_meta)

    _registrar_historico(
        sol.id,
        evento="CRIADA_OC_BRIDGE",
        usuario=usuario,
        status_novo="Pendente",
        detalhe=f"Solicitação criada a partir da OC {numero_oc} via bridge.",
        payload={"numero_oc": numero_oc, "fonte": fonte},
    )
    db.session.commit()

    veiculo = AgendamentoVeiculo.query.get(sol.veiculo_id) if sol.veiculo_id else None
    return jsonify(
        {
            "sucesso": True,
            "solicitacao": _serializar_solicitacao(sol, veiculo=veiculo),
            "pdf_url": f"/api/logistica/central-viagens/solicitacoes/{sol.id}/ordem-coleta.pdf",
        }
    )


@agendamento_bp.route("/api/logistica/central-viagens/solicitacoes/<int:solicitacao_id>/anexo")
@permission_required_any("PAGE_LOGISTICA_AGENDAMENTO", "PAGE_LOGISTICA_SOLICITACAO")
def central_viagens_anexo_solicitacao(solicitacao_id: int):
    row = AgendamentoSolicitacao.query.get(solicitacao_id)
    if not row:
        return jsonify({"error": "Solicitação não encontrada."}), 404

    payload_origem = {}
    if row.payload_origem:
        try:
            payload_origem = json.loads(row.payload_origem) or {}
        except Exception:
            payload_origem = {}
    anexo = _extrair_anexo_payload(payload_origem)
    caminho_relativo = str(anexo.get("caminho_relativo") or "").strip()
    if not caminho_relativo:
        return jsonify({"error": "Esta solicitação não possui anexo."}), 404

    caminho = os.path.join(current_app.instance_path, caminho_relativo)
    if not os.path.isfile(caminho):
        return jsonify({"error": "Arquivo do anexo não encontrado."}), 404

    return send_file(
        caminho,
        mimetype=str(anexo.get("mimetype") or "application/octet-stream") or "application/octet-stream",
        as_attachment=False,
        download_name=str(anexo.get("nome_original") or os.path.basename(caminho)),
    )


@agendamento_bp.route("/api/logistica/central-viagens/solicitacoes/<int:solicitacao_id>/ordem-coleta.pdf")
@permission_required_any("PAGE_LOGISTICA_AGENDAMENTO", "PAGE_LOGISTICA_SOLICITACAO")
def central_viagens_ordem_coleta_pdf(solicitacao_id: int):
    row = AgendamentoSolicitacao.query.get(solicitacao_id)
    if not row:
        return jsonify({"error": "Solicitação não encontrada."}), 404
    if str(row.tipo or "").strip() != "COLETA":
        return jsonify({"error": "Ordem de coleta disponível apenas para solicitações de coleta."}), 409

    itens = (
        AgendamentoSolicitacaoItem.query
        .filter_by(solicitacao_id=row.id)
        .order_by(AgendamentoSolicitacaoItem.sequencia.asc())
        .all()
    )
    pdf_bytes = gerar_ordem_coleta_pdf(row, itens)
    nome = f"ordem_coleta_{str(row.numero_oc or row.id).replace(' ', '_')}.pdf"
    return send_file(
        io.BytesIO(pdf_bytes),
        mimetype="application/pdf",
        as_attachment=True,
        download_name=nome,
    )


@agendamento_bp.route("/api/logistica/central-viagens/importar-oc/preview", methods=["POST"])
@permission_required("PAGE_LOGISTICA_AGENDAMENTO")
def central_viagens_importar_oc_preview():
    if not _is_admin_or_compras():
        return jsonify({"error": "Somente Compras e Administrador podem importar OC."}), 403
    arquivo = request.files.get("arquivo")
    if not arquivo or not getattr(arquivo, "filename", ""):
        return jsonify({"error": "Selecione um arquivo para importar."}), 400
    try:
        linhas = _extrair_linhas_oc_upload(arquivo)
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400

    validas, erros = _validar_linhas_oc(linhas)
    preview = [
        {
            "line": row["line"],
            "numero_oc": row["numero_oc"],
            "fornecedor": row["fornecedor"],
            "cidade": row["cidade"],
            "uf": row["uf"],
            "valor": row["valor"],
            "data_prevista": row["data_prevista"].strftime("%Y-%m-%d") if row.get("data_prevista") else "",
            "transportadora": row.get("transportadora") or "",
            "endereco": row.get("endereco") or "",
        }
        for row in validas[:300]
    ]
    return jsonify(
        {
            "preview": preview,
            "totais": {
                "lidas": len(linhas),
                "validas": len(validas),
                "invalidas": len(erros),
            },
            "erros": erros[:200],
        }
    )


@agendamento_bp.route("/api/logistica/central-viagens/importar-oc/confirmar", methods=["POST"])
@permission_required("PAGE_LOGISTICA_AGENDAMENTO")
def central_viagens_importar_oc_confirmar():
    if not _is_admin_or_compras():
        return jsonify({"error": "Somente Compras e Administrador podem importar OC."}), 403

    payload = request.get_json(silent=True) or {}
    rows = payload.get("rows") if isinstance(payload.get("rows"), list) else []
    if not rows:
        return jsonify({"error": "Nenhum registro enviado para importação."}), 400

    parsed = []
    for idx, raw in enumerate(rows, start=1):
        parsed.append(
            {
                "line": int(raw.get("line") or idx),
                "numero_oc": str(raw.get("numero_oc") or "").strip(),
                "fornecedor": str(raw.get("fornecedor") or "").strip(),
                "cidade": str(raw.get("cidade") or "").strip(),
                "uf": str(raw.get("uf") or "").strip().upper()[:2],
                "valor": _parse_float_br(raw.get("valor")),
                "data_prevista": _parse_data_livre(raw.get("data_prevista")),
                "transportadora": str(raw.get("transportadora") or "").strip(),
                "endereco": str(raw.get("endereco") or "").strip(),
            }
        )
    validas, erros = _validar_linhas_oc(parsed)
    if not validas:
        return jsonify({"error": "Nenhuma linha válida para importar.", "erros": erros}), 400

    usuario = session.get("username", "sistema")
    criadas = 0
    ignoradas = 0
    for row in validas:
        numero_oc = row["numero_oc"]
        existente = (
            _query_solicitacoes_visiveis()
            .filter(
                AgendamentoSolicitacao.tipo == "COLETA",
                AgendamentoSolicitacao.numero_oc == numero_oc,
                AgendamentoSolicitacao.status != "Cancelada",
            )
            .first()
        )
        if existente:
            ignoradas += 1
            continue

        agora = datetime.now()
        sol = AgendamentoSolicitacao(
            tipo="COLETA",
            status="Pendente",
            prioridade="Media",
            prazo_limite=row.get("data_prevista"),
            data_desejada=row.get("data_prevista"),
            solicitante=usuario,
            criado_em=agora,
            atualizado_em=agora,
            documento_tipo="OC",
            documento_numero=numero_oc,
            numero_oc=numero_oc,
            origem_documento="ORDEM_DE_COMPRA",
            observacoes_solicitante=f"Coleta gerada automaticamente pela importação de OC. Valor informado: {row.get('valor')}",
            observacoes_logistica=(f"Transportadora: {row.get('transportadora')}" if row.get("transportadora") else None),
            payload_origem=_json_text({"origem": "importacao_oc", "valor": row.get("valor"), "line": row.get("line")}),
        )
        ok, msg = _aplicar_parceiro(
            sol,
            {
                "nome": row.get("fornecedor"),
                "logradouro": row.get("endereco") or "A definir",
                "cidade": row.get("cidade"),
                "uf": row.get("uf"),
            },
            "Fornecedor",
        )
        if not ok:
            ignoradas += 1
            continue

        db.session.add(sol)
        db.session.flush()
        sol.codigo = f"LOG-{agora.strftime('%Y%m%d')}-{sol.id:04d}"
        _sincronizar_itens(
            sol,
            [
                {
                    "descricao": f"Coleta referente à OC {numero_oc}",
                    "quantidade": 1,
                    "unidade": "UN",
                    "volumes": 1,
                    "observacoes": f"Fornecedor: {row.get('fornecedor')}",
                }
            ],
        )
        _registrar_historico(
            sol.id,
            evento="CRIADA_IMPORT_OC",
            usuario=usuario,
            status_novo="Pendente",
            detalhe=f"Solicitação criada automaticamente pela importação da OC {numero_oc}.",
            payload=row,
        )
        criadas += 1

    db.session.commit()
    return jsonify({"sucesso": True, "criadas": criadas, "ignoradas": ignoradas, "erros": erros[:100]})


@agendamento_bp.route("/api/logistica/central-viagens/solicitacoes/<int:solicitacao_id>/prioridade", methods=["POST"])
@permission_required("PAGE_LOGISTICA_AGENDAMENTO")
def central_viagens_alterar_prioridade(solicitacao_id: int):
    row = _get_solicitacao_visivel(solicitacao_id)
    if not row or not _is_origem_automatica(row):
        return jsonify({"error": "Viagem não encontrada."}), 404
    payload = request.get_json(silent=True) or {}
    prioridade = str(payload.get("prioridade") or "").strip()
    if prioridade not in PRIORIDADES_SOLICITACAO:
        return jsonify({"error": "Prioridade inválida."}), 400
    anterior = row.prioridade
    row.prioridade = prioridade
    row.atualizado_em = datetime.now()
    _registrar_historico(
        row.id,
        evento="PRIORIDADE_ALTERADA",
        usuario=session.get("username", "sistema"),
        detalhe=f"Prioridade alterada de {anterior} para {prioridade}.",
    )
    db.session.commit()
    return jsonify({"sucesso": True})


@agendamento_bp.route("/api/logistica/central-viagens/solicitacoes/<int:solicitacao_id>/reorganizar", methods=["POST"])
@permission_required("PAGE_LOGISTICA_AGENDAMENTO")
def central_viagens_reorganizar(solicitacao_id: int):
    row = _get_solicitacao_visivel(solicitacao_id)
    if not row or not _is_origem_automatica(row):
        return jsonify({"error": "Viagem não encontrada."}), 404
    payload = request.get_json(silent=True) or {}
    regra = str(payload.get("regra") or "").strip()
    agrupamento = str(payload.get("agrupamento") or "").strip()
    prioridade = str(payload.get("prioridade") or "").strip()
    if prioridade and prioridade in PRIORIDADES_SOLICITACAO:
        row.prioridade = prioridade
    row.atualizado_em = datetime.now()
    _registrar_historico(
        row.id,
        evento="REORGANIZADA_CENTRAL",
        usuario=session.get("username", "sistema"),
        detalhe=f"Reorganização aplicada. Regra: {regra or 'manual'} · Agrupamento: {agrupamento or 'nenhum'}.",
        payload=payload,
    )
    db.session.commit()
    return jsonify({"sucesso": True})


@agendamento_bp.route("/api/logistica/central-viagens/solicitacoes/<int:solicitacao_id>/cancelar", methods=["POST"])
@permission_required("PAGE_LOGISTICA_AGENDAMENTO")
def central_viagens_cancelar(solicitacao_id: int):
    row = _get_solicitacao_visivel(solicitacao_id)
    if not row or not _is_origem_automatica(row):
        return jsonify({"error": "Viagem não encontrada."}), 404
    if row.status in {"Concluida", "Cancelada"}:
        return jsonify({"error": "Viagem já finalizada."}), 409
    payload = request.get_json(silent=True) or {}
    motivo = str(payload.get("motivo") or "").strip()
    if not motivo:
        return jsonify({"error": "Informe o motivo do cancelamento."}), 400
    row.status = "Cancelada"
    row.cancelado_por = session.get("username", "sistema")
    row.cancelado_em = datetime.now()
    row.motivo_cancelamento = motivo
    row.cancelamento_pendente = False
    row.atualizado_em = datetime.now()
    _registrar_historico(
        row.id,
        evento="CANCELAMENTO_CENTRAL",
        usuario=session.get("username", "sistema"),
        status_anterior="",
        status_novo="Cancelada",
        detalhe=motivo,
    )
    db.session.commit()
    return jsonify({"sucesso": True})


@agendamento_bp.route("/api/logistica/central-viagens/solicitacoes/<int:solicitacao_id>/excluir", methods=["DELETE"])
@permission_required("PAGE_LOGISTICA_AGENDAMENTO")
def central_viagens_excluir(solicitacao_id: int):
    if not is_admin_session():
        return jsonify({"error": "Somente administrador pode excluir viagens."}), 403
    row = _get_solicitacao_visivel(solicitacao_id)
    if not row or not _is_origem_automatica(row):
        return jsonify({"error": "Viagem não encontrada."}), 404
    if row.status == "EmRota":
        return jsonify({"error": "Não é possível excluir viagem em rota."}), 409
    payload = request.get_json(silent=True) or {}
    usuario = session.get("username", "sistema")
    motivo = str(payload.get("motivo") or "").strip() or "Excluida na Central de Viagens."
    status_anterior = str(row.status or "").strip()
    row.atualizado_em = datetime.now()
    if status_anterior not in {"Concluida", "Cancelada"}:
        row.status = "Cancelada"
    if not row.cancelado_em:
        row.cancelado_em = datetime.now()
    if not row.cancelado_por:
        row.cancelado_por = usuario
    if not row.motivo_cancelamento:
        row.motivo_cancelamento = motivo
    _registrar_historico(
        row.id,
        evento="EXCLUIDA_CENTRAL",
        usuario=usuario,
        status_anterior=status_anterior,
        status_novo=str(row.status or "").strip(),
        detalhe=motivo,
        payload=payload,
    )
    db.session.commit()
    return jsonify({"sucesso": True})


@agendamento_bp.route("/api/logistica/central-viagens/solicitacoes/excluir-lote", methods=["POST", "DELETE"])
@permission_required("PAGE_LOGISTICA_AGENDAMENTO")
def central_viagens_excluir_lote():
    if not is_admin_session():
        return jsonify({"error": "Somente administrador pode excluir viagens."}), 403

    payload = request.get_json(silent=True)
    if not isinstance(payload, dict):
        payload = {}

    ids_raw = payload.get("ids") if isinstance(payload.get("ids"), list) else []
    if not ids_raw:
        ids_query = str(request.args.get("ids") or "").strip()
        if ids_query:
            ids_raw = [item.strip() for item in ids_query.split(",") if item.strip()]

    ids: list[int] = []
    for item in ids_raw:
        try:
            sid = int(str(item).strip())
        except (TypeError, ValueError):
            continue
        if sid > 0 and sid not in ids:
            ids.append(sid)

    if not ids:
        return jsonify({"error": "Informe ao menos uma solicitação para excluir."}), 400

    usuario = session.get("username", "sistema")
    motivo = str(payload.get("motivo") or "").strip() or "Excluida em massa na Central de Viagens."
    excluidas = 0
    falhas: list[dict] = []

    for sid in ids:
        row = _get_solicitacao_visivel(sid)
        if not row or not _is_origem_automatica(row):
            falhas.append({"id": sid, "erro": "Viagem não encontrada."})
            continue
        if row.status == "EmRota":
            falhas.append({"id": sid, "erro": "Não é possível excluir viagem em rota."})
            continue

        status_anterior = str(row.status or "").strip()
        row.atualizado_em = datetime.now()
        if status_anterior not in {"Concluida", "Cancelada"}:
            row.status = "Cancelada"
        if not row.cancelado_em:
            row.cancelado_em = datetime.now()
        if not row.cancelado_por:
            row.cancelado_por = usuario
        if not row.motivo_cancelamento:
            row.motivo_cancelamento = motivo

        _registrar_historico(
            row.id,
            evento="EXCLUIDA_CENTRAL_LOTE",
            usuario=usuario,
            status_anterior=status_anterior,
            status_novo=str(row.status or "").strip(),
            detalhe=motivo,
            payload={"origem": "lote", "ids_total": len(ids)},
        )
        excluidas += 1

    db.session.commit()
    return jsonify({
        "sucesso": True,
        "total": len(ids),
        "excluidas": excluidas,
        "falhas": falhas,
    })


@agendamento_bp.route("/api/logistica/agendamento-veiculos/minhas-solicitacoes")
@permission_required("PAGE_LOGISTICA_SOLICITACAO")
def minhas_solicitacoes_agendamento():
    usuario = session.get("username", "desconhecido")
    rows = (
        _query_solicitacoes_visiveis()
        .filter(AgendamentoSolicitacao.solicitante == usuario)
        .order_by(AgendamentoSolicitacao.criado_em.desc())
        .limit(30)
        .all()
    )
    veiculos = {row.id: row for row in listar_veiculos_agendamento()}
    return jsonify({"rows": [_serializar_solicitacao(row, veiculo=veiculos.get(row.veiculo_id)) for row in rows]})


@agendamento_bp.route("/api/logistica/agendamento-veiculos/agenda")
@permission_required("PAGE_LOGISTICA_AGENDAMENTO")
def agenda_agendamento_veiculos():
    modo = str(request.args.get("modo") or "semana").strip().lower()
    if modo not in {"dia", "semana"}:
        modo = "semana"
    data_raw = str(request.args.get("data") or "").strip()
    try:
        data_base = datetime.fromisoformat(f"{data_raw}T00:00" if data_raw else datetime.now().strftime("%Y-%m-%dT00:00"))
    except ValueError:
        data_base = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
    inicio = data_base.replace(hour=0, minute=0, second=0, microsecond=0)
    fim = inicio + timedelta(days=1 if modo == "dia" else 7)
    veiculos = {row.id: row for row in listar_veiculos_agendamento()}
    rows = (
        _query_solicitacoes_visiveis()
        .filter(
            AgendamentoSolicitacao.status.in_(list(STATUS_ATIVOS) + ["Concluida"]),
            AgendamentoSolicitacao.data_hora_saida_prevista.isnot(None),
            AgendamentoSolicitacao.data_hora_saida_prevista >= inicio,
            AgendamentoSolicitacao.data_hora_saida_prevista < fim,
        )
        .order_by(AgendamentoSolicitacao.data_hora_saida_prevista.asc())
        .all()
    )
    agenda = []
    for row in rows:
        veiculo = veiculos.get(row.veiculo_id)
        if not veiculo:
            continue
        agenda.append(
            {
                "id": row.id,
                "codigo": str(row.codigo or f"LOG-{row.id}").strip(),
                "veiculo_codigo": str(veiculo.codigo or "").strip(),
                "veiculo_nome": str(veiculo.nome_exibicao or veiculo.codigo or "").strip(),
                "tipo": str(row.tipo or "").strip(),
                "status": str(row.status or "").strip(),
                "status_label": status_label_agendamento(row.status),
                "documento": str(row.documento_numero or "").strip(),
                "parceiro_nome": str(row.parceiro_nome or "").strip(),
                "motorista_nome": str(row.motorista_nome or "").strip(),
                "km_estimado": float(row.km_estimado) if row.km_estimado is not None else None,
                "data": row.data_hora_saida_prevista.strftime("%Y-%m-%d"),
                "saida_label": row.data_hora_saida_prevista.strftime("%d/%m/%Y %H:%M"),
                "retorno_label": row.data_hora_retorno_prevista.strftime("%d/%m/%Y %H:%M") if row.data_hora_retorno_prevista else "",
                "resumo_itens": str(row.resumo_itens or "").strip(),
            }
        )
    return jsonify({"modo": modo, "data_base": inicio.strftime("%Y-%m-%d"), "agenda": agenda})


@agendamento_bp.route("/api/logistica/agendamento-veiculos/cadastros")
@permission_required("PAGE_LOGISTICA_AGENDAMENTO")
def listar_cadastros_agendamento():
    tipo = str(request.args.get("tipo") or "").strip().lower()
    if tipo not in {"fornecedor", "cliente"}:
        return jsonify({"error": "Informe o tipo do cadastro."}), 400
    q = str(request.args.get("q") or "").strip()
    limit = max(1, min(int(request.args.get("limit") or 30), 100))
    return jsonify({"rows": listar_cadastros(tipo, q=q, limit=limit)})


@agendamento_bp.route("/api/logistica/agendamento-veiculos/cadastros/importar", methods=["POST"])
@permission_required("PAGE_LOGISTICA_AGENDAMENTO")
def importar_cadastros_agendamento():
    tipo = str(request.form.get("tipo") or "").strip().lower()
    if tipo not in {"fornecedor", "cliente"}:
        return jsonify({"error": "Tipo de cadastro inválido."}), 400
    arquivo = request.files.get("arquivo")
    try:
        resumo = importar_cadastros_excel(tipo, arquivo=arquivo, nome_arquivo=str(getattr(arquivo, "filename", "") or ""))
    except Exception as exc:
        db.session.rollback()
        return jsonify({"error": str(exc)}), 400
    return jsonify({"sucesso": True, "resumo": resumo, "cadastros": resumo_cadastros()})


@agendamento_bp.route("/api/logistica/agendamento-veiculos/motoristas")
@permission_required("PAGE_LOGISTICA_AGENDAMENTO")
def listar_motoristas_endpoint():
    sincronizar_motoristas_usuarios(commit=True)
    q = str(request.args.get("q") or "").strip()
    incluir_inativos = str(request.args.get("incluir_inativos") or "").strip().lower() in {"1", "true", "sim"}
    return jsonify({"rows": listar_motoristas_agendamento(q=q, incluir_inativos=incluir_inativos)})


@agendamento_bp.route("/api/logistica/agendamento-veiculos/motoristas", methods=["POST"])
@permission_required("PAGE_LOGISTICA_AGENDAMENTO")
def salvar_motoristas_endpoint():
    payload = request.get_json(silent=True) or {}
    try:
        row = salvar_motorista_agendamento(payload)
        db.session.commit()
    except ValueError as exc:
        db.session.rollback()
        return jsonify({"error": str(exc)}), 400
    except Exception as exc:
        db.session.rollback()
        return jsonify({"error": str(exc)}), 500
    return jsonify({"sucesso": True, "motorista": serializar_motorista(row)})


@agendamento_bp.route("/api/logistica/agendamento-veiculos/referencia/oc/<numero_oc>")
def consultar_oc_agendamento_endpoint(numero_oc: str):
    from conferencia_app.auth import has_permission
    if not (has_permission("PAGE_LOGISTICA_SOLICITACAO") or has_permission("PAGE_LOGISTICA_AGENDAMENTO")):
        return jsonify({"error": "Acesso negado."}), 403
    resultado = consultar_oc_agendamento(numero_oc)
    return jsonify(resultado), (200 if resultado.get("encontrada") else 404)


@agendamento_bp.route("/api/logistica/agendamento-veiculos/referencia/nf/<numero_nf>")
def consultar_nf_agendamento_endpoint(numero_nf: str):
    from conferencia_app.auth import has_permission
    if not (has_permission("PAGE_LOGISTICA_SOLICITACAO") or has_permission("PAGE_LOGISTICA_AGENDAMENTO")):
        return jsonify({"error": "Acesso negado."}), 403
    resultado = consultar_nf_agendamento(numero_nf)
    return jsonify(resultado), (200 if resultado.get("encontrada") else 404)


@agendamento_bp.route("/api/logistica/agendamento-veiculos/cif/reprocessar", methods=["POST"])
@permission_required("PAGE_LOGISTICA_AGENDAMENTO")
def reprocessar_solicitacoes_cif():
    """Reprocessamento manual da automacao de Solicitacoes Logisticas CIF.

    Executa as duas regras (Coleta a partir de OC CIF e Entrega a partir de
    Romaneio CIF). Util quando o scheduler falhou ou para forcar a geracao.
    Aceita opcional {"regra": "coleta"|"entrega"} para rodar apenas uma.
    """
    from ..services.solicitacao_logistica_cif_service import (
        executar_ciclo,
        gerar_solicitacoes_coleta_cif,
        gerar_solicitacoes_entrega_cif,
    )

    regra = str((request.get_json(silent=True) or {}).get("regra") or "").strip().lower()
    try:
        if regra == "coleta":
            resultado = {"coleta": gerar_solicitacoes_coleta_cif()}
        elif regra == "entrega":
            resultado = {"entrega": gerar_solicitacoes_entrega_cif()}
        else:
            resultado = executar_ciclo()
    except Exception as exc:
        current_app.logger.exception("Falha no reprocessamento manual de Solicitacoes CIF")
        return jsonify({"error": f"Falha ao reprocessar: {exc}"}), 500
    return jsonify({"sucesso": True, "resultado": resultado})


def _registrar_historico(
    solicitacao_id: int,
    *,
    evento: str,
    usuario: str,
    status_anterior: str | None = None,
    status_novo: str | None = None,
    detalhe: str | None = None,
    payload=None,
) -> None:
    db.session.add(
        AgendamentoSolicitacaoHistorico(
            solicitacao_id=solicitacao_id,
            evento=evento,
            status_anterior=status_anterior,
            status_novo=status_novo,
            usuario=usuario,
            detalhe=str(detalhe or "").strip() or None,
            payload_json=_json_text(payload),
            criado_em=datetime.now(),
        )
    )


def _aplicar_parceiro(registro: AgendamentoSolicitacao, parceiro: dict, parceiro_tipo: str) -> tuple[bool, str]:
    nome = str(parceiro.get("nome") or parceiro.get("razao_social") or "").strip()
    logradouro = str(parceiro.get("logradouro") or "").strip()
    cidade = str(parceiro.get("cidade") or "").strip()
    uf = str(parceiro.get("uf") or "").strip()
    if not nome:
        return False, f"Informe o nome do {parceiro_tipo.lower()}."
    if not logradouro or not cidade or not uf:
        return False, "Preencha o endereco completo antes de salvar."
    registro.parceiro_tipo = parceiro_tipo
    registro.parceiro_codigo = str(parceiro.get("codigo") or "").strip() or None
    registro.parceiro_nome = nome
    registro.parceiro_razao_social = str(parceiro.get("razao_social") or nome).strip() or None
    registro.parceiro_documento = re.sub(r"\D", "", str(parceiro.get("cnpj_cpf") or parceiro.get("documento") or "")) or None
    registro.contato = str(parceiro.get("contato") or "").strip() or None
    registro.telefone = str(parceiro.get("telefone") or "").strip() or None
    registro.email = str(parceiro.get("email") or "").strip() or None
    registro.logradouro = logradouro
    registro.numero = str(parceiro.get("numero") or "").strip() or None
    registro.complemento = str(parceiro.get("complemento") or "").strip() or None
    registro.bairro = str(parceiro.get("bairro") or "").strip() or None
    registro.cidade = cidade
    registro.uf = uf[:2].upper()
    registro.cep = re.sub(r"\D", "", str(parceiro.get("cep") or "")) or None
    registro.observacoes_endereco = str(parceiro.get("observacoes") or "").strip() or None
    return True, ""


def _resolver_parceiro_payload(payload: dict, tipo_cadastro: str) -> dict:
    cadastro_id = payload.get("cadastro_id")
    model = AgendamentoFornecedor if tipo_cadastro == "fornecedor" else AgendamentoCliente
    if cadastro_id not in (None, ""):
        try:
            cadastro = model.query.get(int(cadastro_id))
        except (TypeError, ValueError):
            cadastro = None
        if cadastro:
            return serializar_cadastro(cadastro, tipo_cadastro)
    parceiro = payload.get("parceiro")
    return parceiro if isinstance(parceiro, dict) else {}


def _sincronizar_itens(registro: AgendamentoSolicitacao, itens_payload: list[dict]) -> bool:
    AgendamentoSolicitacaoItem.query.filter_by(solicitacao_id=registro.id).delete()
    validos = 0
    total_volumes = 0.0
    for idx, item in enumerate(itens_payload, start=1):
        descricao = str(item.get("descricao") or "").strip()
        try:
            quantidade = float(str(item.get("quantidade") or 0).replace(",", "."))
        except (TypeError, ValueError):
            quantidade = 0.0
        if not descricao or quantidade <= 0:
            continue
        try:
            volumes = float(str(item.get("volumes") or 0).replace(",", "."))
        except (TypeError, ValueError):
            volumes = 0.0
        total_volumes += volumes
        validos += 1
        db.session.add(
            AgendamentoSolicitacaoItem(
                solicitacao_id=registro.id,
                sequencia=idx,
                codigo_item=str(item.get("codigo_item") or "").strip() or None,
                descricao=descricao,
                quantidade=quantidade,
                unidade=str(item.get("unidade") or "").strip() or None,
                volumes=volumes,
                observacoes=str(item.get("observacoes") or "").strip() or None,
            )
        )
    registro.qtd_itens = validos
    registro.qtd_volumes = total_volumes
    volumes_label = int(total_volumes) if float(total_volumes).is_integer() else round(total_volumes, 2)
    registro.resumo_itens = f"{validos} item(ns) / {volumes_label} volume(s)"
    return validos > 0


@agendamento_bp.route("/api/logistica/agendamento-veiculos/solicitacoes/<int:solicitacao_id>")
@permission_required("PAGE_LOGISTICA_AGENDAMENTO")
def obter_solicitacao_agendamento(solicitacao_id: int):
    row = AgendamentoSolicitacao.query.get(solicitacao_id)
    if not row:
        return jsonify({"error": "Solicitação não encontrada."}), 404
    itens = AgendamentoSolicitacaoItem.query.filter_by(solicitacao_id=row.id).order_by(AgendamentoSolicitacaoItem.sequencia.asc()).all()
    historico = AgendamentoSolicitacaoHistorico.query.filter_by(solicitacao_id=row.id).order_by(AgendamentoSolicitacaoHistorico.criado_em.desc()).all()
    veiculo = AgendamentoVeiculo.query.get(row.veiculo_id) if row.veiculo_id else None
    return jsonify({"solicitacao": _serializar_solicitacao(row, veiculo=veiculo, itens=itens, historico=historico)})


@agendamento_bp.route("/api/logistica/agendamento-veiculos/solicitacoes/<int:solicitacao_id>/documento", methods=["PATCH"])
@permission_required("PAGE_LOGISTICA_AGENDAMENTO")
def atualizar_documento_solicitacao(solicitacao_id: int):
    """Gestor informa/corrige OC ou NF — salva e enriquece automaticamente com dados do pedido."""
    row = _get_solicitacao_visivel(solicitacao_id)
    if not row:
        return jsonify({"error": "Solicitação não encontrada."}), 404
    if row.status in ("Cancelada",):
        return jsonify({"error": "Não é possível editar uma solicitação cancelada."}), 400

    payload = request.get_json(silent=True) or {}
    numero_oc = str(payload.get("numero_oc") or "").strip()
    numero_nf = re.sub(r"\D", "", str(payload.get("numero_nf") or ""))

    if not numero_oc and not numero_nf:
        return jsonify({"error": "Informe a OC ou a NF."}), 400

    usuario = session.get("username", "sistema")

    # ── 1. Salva o número no registro ──────────────────────────────────────
    if numero_oc and row.tipo == "COLETA":
        row.numero_oc = numero_oc
        row.documento_numero = numero_oc
        row.documento_tipo = "OC"
    elif numero_nf and row.tipo == "ENTREGA":
        row.numero_nf = numero_nf
        row.documento_numero = numero_nf
        row.documento_tipo = "NF"
    else:
        # AVULSA ou campo não corresponde ao tipo — aceita o que vier
        if numero_oc:
            row.numero_oc = numero_oc
            if not row.documento_numero:
                row.documento_numero = numero_oc
                row.documento_tipo = "OC"
        if numero_nf:
            row.numero_nf = numero_nf
            if not row.documento_numero or row.documento_tipo == "OC":
                row.documento_numero = numero_nf
                row.documento_tipo = "NF"

    # ── 2. Enriquece com dados do pedido/NF ───────────────────────────────
    enriquecido = False
    aviso_enriquecimento = ""
    parceiro_nome_aplicado = ""
    itens_aplicados = 0
    try:
        consulta = (
            consultar_oc_agendamento(numero_oc) if numero_oc
            else consultar_nf_agendamento(numero_nf)
        )
        if consulta.get("encontrada"):
            # OC retorna "fornecedor"; NF retorna "cliente"
            parceiro = consulta.get("fornecedor") or consulta.get("cliente") or {}

            if parceiro.get("nome"):
                row.parceiro_nome = parceiro["nome"]
                parceiro_nome_aplicado = parceiro["nome"]
            if parceiro.get("razao_social"):
                row.parceiro_razao_social = parceiro["razao_social"]
            if parceiro.get("cnpj_cpf"):
                row.parceiro_documento = re.sub(r"\D", "", str(parceiro["cnpj_cpf"]))
            if parceiro.get("codigo"):
                row.parceiro_codigo = str(parceiro["codigo"])
            if parceiro.get("logradouro"):
                row.logradouro = parceiro["logradouro"]
            if parceiro.get("numero"):
                row.numero = parceiro["numero"]
            if parceiro.get("complemento"):
                row.complemento = parceiro["complemento"]
            if parceiro.get("bairro"):
                row.bairro = parceiro["bairro"]
            if parceiro.get("cidade"):
                row.cidade = parceiro["cidade"]
            if parceiro.get("uf"):
                row.uf = str(parceiro["uf"])[:2].upper()
            if parceiro.get("cep"):
                row.cep = parceiro["cep"]
            if parceiro.get("telefone") and not row.telefone:
                row.telefone = parceiro["telefone"]
            if parceiro.get("email") and not row.email:
                row.email = parceiro["email"]

            # Substitui itens existentes pelos do pedido/NF
            itens_novos = consulta.get("itens") or []
            if itens_novos:
                AgendamentoSolicitacaoItem.query.filter_by(solicitacao_id=row.id).delete()
                for seq, it in enumerate(itens_novos, start=1):
                    db.session.add(AgendamentoSolicitacaoItem(
                        solicitacao_id=row.id,
                        sequencia=seq,
                        codigo_item=str(it.get("codigo_item") or "").strip(),
                        descricao=str(it.get("descricao") or f"Item {seq}").strip(),
                        quantidade=float(it.get("quantidade") or 0.0),
                        unidade=str(it.get("unidade") or "").strip(),
                        volumes=float(it.get("volumes") or 0.0),
                        observacoes=str(it.get("observacoes") or "").strip() or None,
                    ))
                row.qtd_itens = len(itens_novos)
                total_vol = sum(float(i.get("volumes") or 0.0) for i in itens_novos)
                row.resumo_itens = f"{len(itens_novos)} item(ns) / {total_vol:.0f} vol."
                itens_aplicados = len(itens_novos)

            if consulta.get("warning"):
                aviso_enriquecimento = consulta["warning"]
            enriquecido = True
        else:
            aviso_enriquecimento = consulta.get("error") or "Documento não encontrado na base de dados."
    except Exception as exc:
        current_app.logger.warning("Enriquecimento de documento falhou: %s", exc)
        aviso_enriquecimento = "Número salvo, mas não foi possível buscar os dados automaticamente."

    detalhe_hist = (
        f"{'OC' if numero_oc else 'NF'} {numero_oc or numero_nf} informada pelo gestor."
        + (" Dados aplicados automaticamente." if enriquecido else f" {aviso_enriquecimento}")
    )
    _registrar_historico(row.id, evento="DOCUMENTO_ATUALIZADO", usuario=usuario, detalhe=detalhe_hist)

    row.atualizado_em = datetime.now()
    db.session.commit()

    itens = (
        AgendamentoSolicitacaoItem.query
        .filter_by(solicitacao_id=row.id)
        .order_by(AgendamentoSolicitacaoItem.sequencia.asc())
        .all()
    )
    veiculo = AgendamentoVeiculo.query.get(row.veiculo_id) if row.veiculo_id else None
    return jsonify({
        "sucesso": True,
        "enriquecido": enriquecido,
        "parceiro_nome": parceiro_nome_aplicado,
        "itens_aplicados": itens_aplicados,
        "aviso": aviso_enriquecimento,
        "solicitacao": _serializar_solicitacao(row, veiculo=veiculo, itens=itens),
    })


@agendamento_bp.route("/api/logistica/agendamento-veiculos/solicitacoes", methods=["POST"])
@permission_required("PAGE_LOGISTICA_SOLICITACAO")
def criar_solicitacao_agendamento():
    return jsonify({"error": "Criação manual de solicitação de viagem foi descontinuada. Use a Central de Viagens."}), 410


@agendamento_bp.route("/api/logistica/agendamento-veiculos/solicitacoes/<int:solicitacao_id>/alocar", methods=["POST"])
@permission_required("PAGE_LOGISTICA_AGENDAMENTO")
def alocar_solicitacao_agendamento(solicitacao_id: int):
    row = _get_solicitacao_visivel(solicitacao_id)
    if not row:
        return jsonify({"error": "Solicitação não encontrada."}), 404
    if str(row.status or "").strip() in {"Concluida", "Cancelada"}:
        return jsonify({"error": "Não é possível alocar uma solicitação finalizada."}), 409

    payload = request.get_json(silent=True) or {}
    try:
        veiculo_id = int(payload.get("veiculo_id")) if payload.get("veiculo_id") not in (None, "") else None
    except (TypeError, ValueError):
        veiculo_id = None
    veiculo = AgendamentoVeiculo.query.get(veiculo_id) if veiculo_id else None
    if not veiculo:
        codigo = str(payload.get("veiculo_codigo") or "").strip().upper()
        veiculo = AgendamentoVeiculo.query.filter_by(codigo=codigo, ativo=True).first() if codigo else None
    if not veiculo:
        return jsonify({"error": "Selecione um veículo válido."}), 400

    try:
        saida = _parse_datetime(payload.get("data_hora_saida_prevista"), "a data e hora de saída", required=True)
        retorno = _parse_datetime(payload.get("data_hora_retorno_prevista"), "a previsão de retorno")
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400
    if retorno and retorno <= saida:
        return jsonify({"error": "A previsão de retorno deve ser maior que a saída."}), 400

    motorista = None
    try:
        motorista_id = int(payload.get("motorista_id")) if payload.get("motorista_id") not in (None, "") else None
    except (TypeError, ValueError):
        motorista_id = None
    if not motorista_id and str(row.tipo or "").strip() != "AVULSA":
        return jsonify({"error": "Selecione um motorista para a viagem."}), 400
    motorista = AgendamentoMotorista.query.get(motorista_id) if motorista_id else None
    if motorista_id and not motorista:
        return jsonify({"error": "Selecione um motorista válido."}), 400

    departamento = str(payload.get("departamento_solicitante") or "").strip().upper()
    departamentos_validos = ["COMPRAS", "ASSISTÊNCIA TÉCNICA", "ENGENHARIA/PCP", "LOGÍSTICA", "FACILITIES"]
    if not departamento or departamento not in departamentos_validos:
        return jsonify({"error": "Selecione o departamento solicitante."}), 400

    buffer_min = int(
        getattr(veiculo, "janela_conflito_min", 0)
        or current_app.config.get("AGENDAMENTO_CONFLITO_MINUTOS", 30)
    )
    query = _query_solicitacoes_visiveis().filter(
        AgendamentoSolicitacao.veiculo_id == veiculo.id,
        AgendamentoSolicitacao.status.in_(["Alocada", "EmRota"]),
        AgendamentoSolicitacao.id != row.id,
    )
    inicio_atual, fim_atual = _intervalo_planejado(AgendamentoSolicitacao(), veiculo=veiculo, inicio_override=saida, fim_override=retorno)
    for existente in query.all():
        outro_inicio, outro_fim = _intervalo_planejado(existente)
        if not outro_inicio or not outro_fim:
            continue
        mesma_janela = (
            inicio_atual == outro_inicio
            and (
                (fim_atual is None and outro_fim is None)
                or (fim_atual is not None and outro_fim is not None and fim_atual == outro_fim)
            )
        )
        if motorista and mesma_janela and int(existente.motorista_id or 0) == int(motorista.id):
            # Permite consolidar multiplas solicitacoes na mesma saida operacional.
            continue
        if (inicio_atual - timedelta(minutes=buffer_min)) < (outro_fim + timedelta(minutes=buffer_min)) and (
            fim_atual + timedelta(minutes=buffer_min)
        ) > (outro_inicio - timedelta(minutes=buffer_min)):
            return jsonify({"error": f"{veiculo.nome_exibicao} já possui uma saída programada para {outro_inicio.strftime('%d/%m/%Y %H:%M')}."}), 409

    if motorista:
        query_motorista = _query_solicitacoes_visiveis().filter(
            AgendamentoSolicitacao.motorista_id == motorista.id,
            AgendamentoSolicitacao.status.in_(["Alocada", "EmAndamento", "EmRota"]),
            AgendamentoSolicitacao.id != row.id,
        )
        for existente in query_motorista.all():
            outro_inicio, outro_fim = _intervalo_planejado(existente)
            if not outro_inicio or not outro_fim:
                continue
            mesma_janela = (
                inicio_atual == outro_inicio
                and (
                    (fim_atual is None and outro_fim is None)
                    or (fim_atual is not None and outro_fim is not None and fim_atual == outro_fim)
                )
            )
            if mesma_janela and int(existente.veiculo_id or 0) == int(veiculo.id):
                continue
            if (inicio_atual - timedelta(minutes=buffer_min)) < (outro_fim + timedelta(minutes=buffer_min)) and (
                fim_atual + timedelta(minutes=buffer_min)
            ) > (outro_inicio - timedelta(minutes=buffer_min)):
                return jsonify({"error": f"{motorista.nome} ja possui uma viagem programada para {outro_inicio.strftime('%d/%m/%Y %H:%M')}."}), 409

    status_anterior = str(row.status or "").strip()
    row.veiculo_id = veiculo.id
    row.motorista_id = motorista.id if motorista else None
    row.motorista_nome = (str(motorista.nome or "").strip() or None) if motorista else None
    row.data_hora_saida_prevista = saida
    row.data_hora_retorno_prevista = retorno
    row.alocado_por = session.get("username", "desconhecido")
    row.alocado_em = datetime.now()
    row.status = "Alocada"
    row.atualizado_em = datetime.now()
    row.departamento_solicitante = departamento
    observacao = str(payload.get("observacoes_logistica") or "").strip()
    if observacao:
        row.observacoes_logistica = observacao

    rota = estimar_rota_agendamento(
        {
            "logradouro": row.logradouro,
            "numero": row.numero,
            "bairro": row.bairro,
            "cidade": row.cidade,
            "uf": row.uf,
            "cep": row.cep,
            "latitude": row.destino_latitude,
            "longitude": row.destino_longitude,
        },
        origem_latitude=payload.get("origem_latitude"),
        origem_longitude=payload.get("origem_longitude"),
    )
    row.origem_latitude = rota.get("origem_latitude")
    row.origem_longitude = rota.get("origem_longitude")
    row.destino_latitude = rota.get("destino_latitude")
    row.destino_longitude = rota.get("destino_longitude")
    row.km_estimado = rota.get("km_estimado")
    row.km_estimado_retorno = rota.get("km_estimado_retorno")

    detalhe = f"Alocada no veículo {veiculo.nome_exibicao} para {saida.strftime('%d/%m/%Y %H:%M')}."
    if motorista:
        detalhe += f" Motorista: {motorista.nome}."
    if rota.get("km_estimado") is not None:
        detalhe += f" Estimativa: {rota['km_estimado']:.1f} km ida / {rota['km_estimado_retorno']:.1f} km ida e volta."
    _registrar_historico(
        row.id,
        evento="ALOCADA",
        usuario=session.get("username", "desconhecido"),
        status_anterior=status_anterior,
        status_novo="Alocada",
        detalhe=detalhe,
        payload=payload,
    )
    db.session.commit()
    _notificar_solicitante_agendamento(row, "Solicitacao de transporte alocada", detalhe)
    return jsonify({"sucesso": True, "solicitacao": _serializar_solicitacao(row, veiculo=veiculo)})


@agendamento_bp.route("/api/logistica/agendamento-veiculos/solicitacoes/alocar-lote", methods=["POST"])
@permission_required("PAGE_LOGISTICA_AGENDAMENTO")
def alocar_solicitacoes_lote_agendamento():
    payload = request.get_json(silent=True) or {}
    ids_raw = payload.get("ids") if isinstance(payload.get("ids"), list) else []
    ids: list[int] = []
    for item in ids_raw:
        try:
            sid = int(item)
        except (TypeError, ValueError):
            continue
        if sid > 0 and sid not in ids:
            ids.append(sid)
    if not ids:
        return jsonify({"error": "Informe as solicitações para alocação em lote."}), 400

    try:
        veiculo_id = int(payload.get("veiculo_id")) if payload.get("veiculo_id") not in (None, "") else None
    except (TypeError, ValueError):
        veiculo_id = None
    veiculo = AgendamentoVeiculo.query.get(veiculo_id) if veiculo_id else None
    if not veiculo:
        codigo = str(payload.get("veiculo_codigo") or "").strip().upper()
        veiculo = AgendamentoVeiculo.query.filter_by(codigo=codigo, ativo=True).first() if codigo else None
    if not veiculo:
        return jsonify({"error": "Selecione um veículo válido."}), 400

    try:
        motorista_id = int(payload.get("motorista_id")) if payload.get("motorista_id") not in (None, "") else None
    except (TypeError, ValueError):
        motorista_id = None
    motorista = AgendamentoMotorista.query.get(motorista_id) if motorista_id else None
    if not motorista:
        return jsonify({"error": "Selecione um motorista válido."}), 400

    try:
        saida = _parse_datetime(payload.get("data_hora_saida_prevista"), "a data e hora de saída", required=True)
        retorno = _parse_datetime(payload.get("data_hora_retorno_prevista"), "a previsão de retorno")
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400
    if retorno and retorno <= saida:
        return jsonify({"error": "A previsão de retorno deve ser maior que a saída."}), 400

    departamento = str(payload.get("departamento_solicitante") or "").strip().upper()
    departamentos_validos = ["COMPRAS", "ASSISTÊNCIA TÉCNICA", "ENGENHARIA/PCP", "LOGÍSTICA", "FACILITIES"]
    if not departamento or departamento not in departamentos_validos:
        return jsonify({"error": "Selecione o departamento solicitante."}), 400

    rows = _query_solicitacoes_visiveis().filter(AgendamentoSolicitacao.id.in_(ids)).all()
    found_ids = {int(r.id) for r in rows}
    missing = [sid for sid in ids if sid not in found_ids]
    if missing:
        return jsonify({"error": f"Solicitações não encontradas: {', '.join(str(x) for x in missing)}."}), 404

    for row in rows:
        if str(row.status or "").strip() in {"Concluida", "Cancelada"}:
            return jsonify({"error": f"Solicitação {row.codigo or row.id} já está finalizada e não pode ser alocada."}), 409

    buffer_min = int(
        getattr(veiculo, "janela_conflito_min", 0)
        or current_app.config.get("AGENDAMENTO_CONFLITO_MINUTOS", 30)
    )
    inicio_atual, fim_atual = _intervalo_planejado(AgendamentoSolicitacao(), veiculo=veiculo, inicio_override=saida, fim_override=retorno)

    query_veiculo = _query_solicitacoes_visiveis().filter(
        AgendamentoSolicitacao.veiculo_id == veiculo.id,
        AgendamentoSolicitacao.status.in_(["Alocada", "EmRota"]),
        AgendamentoSolicitacao.id.notin_(ids),
    )
    for existente in query_veiculo.all():
        outro_inicio, outro_fim = _intervalo_planejado(existente)
        if not outro_inicio or not outro_fim:
            continue
        mesma_janela = (
            inicio_atual == outro_inicio
            and (
                (fim_atual is None and outro_fim is None)
                or (fim_atual is not None and outro_fim is not None and fim_atual == outro_fim)
            )
        )
        if mesma_janela and int(existente.motorista_id or 0) == int(motorista.id):
            continue
        if (inicio_atual - timedelta(minutes=buffer_min)) < (outro_fim + timedelta(minutes=buffer_min)) and (
            fim_atual + timedelta(minutes=buffer_min)
        ) > (outro_inicio - timedelta(minutes=buffer_min)):
            return jsonify({"error": f"{veiculo.nome_exibicao} já possui uma saída programada para {outro_inicio.strftime('%d/%m/%Y %H:%M')}."}), 409

    query_motorista = _query_solicitacoes_visiveis().filter(
        AgendamentoSolicitacao.motorista_id == motorista.id,
        AgendamentoSolicitacao.status.in_(["Alocada", "EmAndamento", "EmRota"]),
        AgendamentoSolicitacao.id.notin_(ids),
    )
    for existente in query_motorista.all():
        outro_inicio, outro_fim = _intervalo_planejado(existente)
        if not outro_inicio or not outro_fim:
            continue
        mesma_janela = (
            inicio_atual == outro_inicio
            and (
                (fim_atual is None and outro_fim is None)
                or (fim_atual is not None and outro_fim is not None and fim_atual == outro_fim)
            )
        )
        if mesma_janela and int(existente.veiculo_id or 0) == int(veiculo.id):
            continue
        if (inicio_atual - timedelta(minutes=buffer_min)) < (outro_fim + timedelta(minutes=buffer_min)) and (
            fim_atual + timedelta(minutes=buffer_min)
        ) > (outro_inicio - timedelta(minutes=buffer_min)):
            return jsonify({"error": f"{motorista.nome} ja possui uma viagem programada para {outro_inicio.strftime('%d/%m/%Y %H:%M')}."}), 409

    observacao = str(payload.get("observacoes_logistica") or "").strip()
    usuario = session.get("username", "desconhecido")
    atualizados = 0
    for row in rows:
        status_anterior = str(row.status or "").strip()
        row.veiculo_id = veiculo.id
        row.motorista_id = motorista.id
        row.motorista_nome = str(motorista.nome or "").strip() or None
        row.data_hora_saida_prevista = saida
        row.data_hora_retorno_prevista = retorno
        row.alocado_por = usuario
        row.alocado_em = datetime.now()
        row.status = "Alocada"
        row.atualizado_em = datetime.now()
        row.departamento_solicitante = departamento
        if observacao:
            row.observacoes_logistica = observacao

        rota = estimar_rota_agendamento(
            {
                "logradouro": row.logradouro,
                "numero": row.numero,
                "bairro": row.bairro,
                "cidade": row.cidade,
                "uf": row.uf,
                "cep": row.cep,
                "latitude": row.destino_latitude,
                "longitude": row.destino_longitude,
            },
            origem_latitude=payload.get("origem_latitude"),
            origem_longitude=payload.get("origem_longitude"),
        )
        row.origem_latitude = rota.get("origem_latitude")
        row.origem_longitude = rota.get("origem_longitude")
        row.destino_latitude = rota.get("destino_latitude")
        row.destino_longitude = rota.get("destino_longitude")
        row.km_estimado = rota.get("km_estimado")
        row.km_estimado_retorno = rota.get("km_estimado_retorno")

        detalhe = f"Alocada em lote no veículo {veiculo.nome_exibicao} para {saida.strftime('%d/%m/%Y %H:%M')}. Motorista: {motorista.nome}."
        _registrar_historico(
            row.id,
            evento="ALOCADA",
            usuario=usuario,
            status_anterior=status_anterior,
            status_novo="Alocada",
            detalhe=detalhe,
            payload={
                "lote": True,
                "ids": ids,
                "veiculo_id": veiculo.id,
                "motorista_id": motorista.id,
                "data_hora_saida_prevista": payload.get("data_hora_saida_prevista"),
                "data_hora_retorno_prevista": payload.get("data_hora_retorno_prevista"),
            },
        )
        atualizados += 1

    db.session.commit()
    return jsonify({"sucesso": True, "total": atualizados})


@agendamento_bp.route("/api/logistica/agendamento-veiculos/solicitacoes/<int:solicitacao_id>/status", methods=["POST"])
@permission_required("PAGE_LOGISTICA_AGENDAMENTO")
def atualizar_status_agendamento(solicitacao_id: int):
    row = _get_solicitacao_visivel(solicitacao_id)
    if not row:
        return jsonify({"error": "Solicitação não encontrada."}), 404
    status_atual = str(row.status or "").strip()
    if status_atual in {"Cancelada", "Concluida"}:
        return jsonify({"error": "A solicitação já está encerrada."}), 409

    payload = request.get_json(silent=True) or {}
    novo_status = str(payload.get("status") or "").strip()
    if novo_status not in STATUS_SOLICITACAO:
        return jsonify({"error": "Status inválido."}), 400
    if novo_status == "Cancelada":
        return jsonify({"error": "Use a ação de cancelamento."}), 400
    if novo_status == "Alocada":
        return jsonify({"error": "Use a ação de alocação para definir veículo e horário."}), 400
    if novo_status == "EmAndamento" and status_atual not in {"Alocada", "EmAndamento"}:
        return jsonify({"error": "A solicitação precisa estar alocada antes de entrar em andamento."}), 409
    if novo_status == "EmRota" and status_atual not in {"Alocada", "EmAndamento", "EmRota"}:
        return jsonify({"error": "A solicitação precisa estar alocada antes de entrar em rota."}), 409
    if novo_status == "Concluida" and status_atual not in {"Alocada", "EmAndamento", "EmRota"}:
        return jsonify({"error": "A solicitação precisa estar alocada ou em andamento antes da conclusão."}), 409
    if novo_status == "Concluida":
        username = session.get("username", "")
        role = session.get("role", "")
        is_motorista_desta_viagem = False
        if row.motorista_id:
            mot = AgendamentoMotorista.query.get(row.motorista_id)
            if mot and mot.usuario_username == username:
                is_motorista_desta_viagem = True
        if role != "Admin" and not is_motorista_desta_viagem:
            return jsonify({"error": "Somente o motorista da viagem ou admin pode concluir."}), 403
    if novo_status == "Pendente" and status_atual in {"EmRota", "EmAndamento"}:
        return jsonify({"error": "Não é possível voltar uma solicitação em andamento para pendente."}), 409

    row.status = novo_status
    row.atualizado_em = datetime.now()
    observacao = str(payload.get("observacoes_logistica") or "").strip()
    if observacao:
        row.observacoes_logistica = observacao
    if novo_status == "Pendente":
        row.veiculo_id = None
        row.motorista_id = None
        row.motorista_nome = None
        row.data_hora_saida_prevista = None
        row.data_hora_retorno_prevista = None
        row.origem_latitude = None
        row.origem_longitude = None
        row.destino_latitude = None
        row.destino_longitude = None
        row.km_estimado = None
        row.km_estimado_retorno = None
        row.alocado_por = None
        row.alocado_em = None
    if novo_status == "Concluida":
        try:
            row.data_hora_saida_real = _parse_datetime(payload.get("data_hora_saida_real"), "a saída real")
            row.data_hora_retorno_real = _parse_datetime(payload.get("data_hora_retorno_real"), "o retorno real")
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400
        row.concluido_por = session.get("username", "desconhecido")
        row.concluido_em = datetime.now()

    _registrar_historico(row.id, evento="STATUS_ALTERADO", usuario=session.get("username", "desconhecido"), status_anterior=status_atual, status_novo=novo_status, detalhe=observacao or f"Status alterado para {status_label_agendamento(novo_status)}.", payload=payload)
    db.session.commit()
    _notificar_solicitante_agendamento(row, "Status da solicitacao atualizado", observacao or f"Status alterado para {status_label_agendamento(novo_status)}.")
    veiculo = AgendamentoVeiculo.query.get(row.veiculo_id) if row.veiculo_id else None
    return jsonify({"sucesso": True, "solicitacao": _serializar_solicitacao(row, veiculo=veiculo)})


@agendamento_bp.route("/api/logistica/agendamento-veiculos/solicitacoes/<int:solicitacao_id>/cancelar", methods=["POST"])
@permission_required("PAGE_LOGISTICA_AGENDAMENTO")
def cancelar_solicitacao_agendamento(solicitacao_id: int):
    """Logística solicita cancelamento — o solicitante precisa aprovar."""
    row = _get_solicitacao_visivel(solicitacao_id)
    if not row:
        return jsonify({"error": "Solicitação não encontrada."}), 404
    if str(row.status or "").strip() == "Concluida":
        return jsonify({"error": "Não é possível cancelar uma solicitação concluída."}), 409
    if str(row.status or "").strip() == "Cancelada":
        return jsonify({"error": "Solicitação já cancelada."}), 409
    if str(row.status or "").strip() == "EmRota" and session.get("role") != "Admin":
        return jsonify({"error": "Somente admin pode cancelar uma solicitação em rota."}), 403

    payload = request.get_json(silent=True) or {}
    motivo = str(payload.get("motivo") or "").strip()
    if not motivo:
        return jsonify({"error": "Informe o motivo do cancelamento."}), 400

    usuario = session.get("username", "desconhecido")
    status_anterior = str(row.status or "").strip()
    if session.get("role") == "Admin":
        row.status = "Cancelada"
        row.cancelado_por = usuario
        row.cancelado_em = datetime.now()
        row.motivo_cancelamento = motivo
        row.cancelamento_pendente = False
        row.cancelamento_solicitado_por = None
        row.cancelamento_motivo_pendente = None
        row.atualizado_em = datetime.now()
        _registrar_historico(
            row.id,
            evento="CANCELAMENTO_ADMIN",
            usuario=usuario,
            status_anterior=status_anterior,
            status_novo="Cancelada",
            detalhe=f"Cancelada diretamente por admin. Motivo: {motivo}",
            payload=payload,
        )
        db.session.commit()
        _notificar_solicitante_agendamento(row, "Solicitacao cancelada", f"Cancelada diretamente por admin. Motivo: {motivo}")
        veiculo = AgendamentoVeiculo.query.get(row.veiculo_id) if row.veiculo_id else None
        return jsonify({"sucesso": True, "mensagem": "Solicitacao cancelada diretamente pelo admin.", "solicitacao": _serializar_solicitacao(row, veiculo=veiculo)})

    row.cancelamento_pendente = True
    row.cancelamento_solicitado_por = usuario
    row.cancelamento_motivo_pendente = motivo
    row.atualizado_em = datetime.now()
    _registrar_historico(row.id, evento="CANCELAMENTO_SOLICITADO", usuario=usuario, status_anterior=status_anterior, detalhe=f"Cancelamento solicitado. Motivo: {motivo}", payload=payload)
    db.session.commit()
    _notificar_solicitante_agendamento(row, "Aprovacao de cancelamento solicitada", f"A logistica solicitou cancelamento. Motivo: {motivo}")
    veiculo = AgendamentoVeiculo.query.get(row.veiculo_id) if row.veiculo_id else None
    return jsonify({"sucesso": True, "mensagem": "Cancelamento solicitado. Aguardando aprovação do solicitante.", "solicitacao": _serializar_solicitacao(row, veiculo=veiculo)})


@agendamento_bp.route("/api/logistica/agendamento-veiculos/solicitacoes/<int:solicitacao_id>/aprovar-cancelamento", methods=["POST"])
@permission_required("PAGE_LOGISTICA_SOLICITACAO")
def aprovar_cancelamento_solicitacao(solicitacao_id: int):
    """Solicitante aprova o cancelamento pedido pela logistica."""
    row = _get_solicitacao_visivel(solicitacao_id)
    if not row:
        return jsonify({"error": "Solicitação não encontrada."}), 404
    if not row.cancelamento_pendente:
        return jsonify({"error": "Não há cancelamento pendente para esta solicitação."}), 409
    username = session.get("username", "")
    if row.solicitante != username and session.get("role") != "Admin":
        return jsonify({"error": "Somente o solicitante original ou admin pode aprovar."}), 403

    status_anterior = str(row.status or "").strip()
    motivo_aprovado = row.cancelamento_motivo_pendente
    row.status = "Cancelada"
    row.cancelado_por = row.cancelamento_solicitado_por
    row.cancelado_em = datetime.now()
    row.motivo_cancelamento = motivo_aprovado
    row.cancelamento_pendente = False
    row.cancelamento_solicitado_por = None
    row.cancelamento_motivo_pendente = None
    row.atualizado_em = datetime.now()
    _registrar_historico(row.id, evento="CANCELAMENTO_APROVADO", usuario=username, status_anterior=status_anterior, status_novo="Cancelada", detalhe=f"Cancelamento aprovado por {username}. Motivo: {motivo_aprovado}")
    db.session.commit()
    _notificar_solicitante_agendamento(row, "Cancelamento aprovado", f"Cancelamento aprovado por {username}.")
    veiculo = AgendamentoVeiculo.query.get(row.veiculo_id) if row.veiculo_id else None
    return jsonify({"sucesso": True, "solicitacao": _serializar_solicitacao(row, veiculo=veiculo)})


@agendamento_bp.route("/api/logistica/agendamento-veiculos/solicitacoes/<int:solicitacao_id>/rejeitar-cancelamento", methods=["POST"])
@permission_required("PAGE_LOGISTICA_SOLICITACAO")
def rejeitar_cancelamento_solicitacao(solicitacao_id: int):
    """Solicitante rejeita o cancelamento pedido pela logistica."""
    row = _get_solicitacao_visivel(solicitacao_id)
    if not row:
        return jsonify({"error": "Solicitação não encontrada."}), 404
    if not row.cancelamento_pendente:
        return jsonify({"error": "Não há cancelamento pendente para esta solicitação."}), 409
    username = session.get("username", "")
    if row.solicitante != username and session.get("role") != "Admin":
        return jsonify({"error": "Somente o solicitante original ou admin pode rejeitar."}), 403

    row.cancelamento_pendente = False
    row.cancelamento_solicitado_por = None
    row.cancelamento_motivo_pendente = None
    row.atualizado_em = datetime.now()
    _registrar_historico(row.id, evento="CANCELAMENTO_REJEITADO", usuario=username, detalhe=f"Cancelamento rejeitado por {username}.")
    db.session.commit()
    _notificar_solicitante_agendamento(row, "Cancelamento rejeitado", f"Cancelamento rejeitado por {username}.")
    veiculo = AgendamentoVeiculo.query.get(row.veiculo_id) if row.veiculo_id else None
    return jsonify({"sucesso": True, "solicitacao": _serializar_solicitacao(row, veiculo=veiculo)})


@agendamento_bp.route("/api/logistica/agendamento-veiculos/motoristas-km")
@permission_required("PAGE_ADMIN_DASHBOARD")
def motoristas_km_dashboard():
    """Retorna km acumulado por motorista para exibir no dashboard principal."""
    dias = int(request.args.get("dias", 30))
    desde = datetime.now() - timedelta(days=dias)
    rows = (
        db.session.query(
            AgendamentoSolicitacao.motorista_id,
            AgendamentoSolicitacao.motorista_nome,
            db.func.count(AgendamentoSolicitacao.id).label("viagens"),
            db.func.sum(AgendamentoSolicitacao.km_estimado).label("km_total"),
        )
        .filter(
            AgendamentoSolicitacao.motorista_id.isnot(None),
            AgendamentoSolicitacao.km_estimado.isnot(None),
            AgendamentoSolicitacao.status.in_(["Alocada", "EmRota", "Concluida"]),
            _filtro_solicitacao_visivel(),
            AgendamentoSolicitacao.criado_em >= desde,
        )
        .group_by(AgendamentoSolicitacao.motorista_id, AgendamentoSolicitacao.motorista_nome)
        .all()
    )
    motoristas = []
    for r in rows:
        km_total = float(r.km_total or 0)
        viagens = int(r.viagens or 0)
        motoristas.append({
            "motorista_id": r.motorista_id,
            "motorista_nome": str(r.motorista_nome or "").strip(),
            "viagens": viagens,
            "km_total": round(km_total, 1),
            "km_medio": round(km_total / viagens, 1) if viagens else 0,
        })
    motoristas.sort(key=lambda x: x["km_total"], reverse=True)
    return jsonify({"motoristas": motoristas, "dias": dias})


@agendamento_bp.route("/api/logistica/agendamento-veiculos/disponibilidade")
@permission_required("PAGE_LOGISTICA_SOLICITACAO")
def disponibilidade_veiculos():
    """Retorna blocos ocupados por veículo para um intervalo de datas."""
    data_inicio_str = request.args.get("data_inicio", "")
    data_fim_str = request.args.get("data_fim", "")
    try:
        data_inicio = datetime.strptime(data_inicio_str, "%Y-%m-%d") if data_inicio_str else datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
        data_fim = datetime.strptime(data_fim_str, "%Y-%m-%d").replace(hour=23, minute=59, second=59) if data_fim_str else data_inicio + timedelta(days=7)
    except ValueError:
        data_inicio = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
        data_fim = data_inicio + timedelta(days=7)

    veiculos = AgendamentoVeiculo.query.filter_by(ativo=True).order_by(AgendamentoVeiculo.ordem_exibicao).all()
    solicitacoes = AgendamentoSolicitacao.query.filter(
        AgendamentoSolicitacao.veiculo_id.isnot(None),
        AgendamentoSolicitacao.data_hora_saida_prevista.isnot(None),
        AgendamentoSolicitacao.status.in_(["Alocada", "EmAndamento", "EmRota"]),
        _filtro_solicitacao_visivel(),
        AgendamentoSolicitacao.data_hora_saida_prevista >= data_inicio,
        AgendamentoSolicitacao.data_hora_saida_prevista <= data_fim,
    ).all()

    blocos_por_veiculo = {}
    for v in veiculos:
        blocos_por_veiculo[v.id] = {
            "veiculo_id": v.id,
            "nome": v.nome_exibicao,
            "placa": v.placa or "",
            "duracao_padrao_min": v.duracao_padrao_min,
            "blocos": [],
        }

    for s in solicitacoes:
        vid = s.veiculo_id
        if vid not in blocos_por_veiculo:
            continue
        inicio = s.data_hora_saida_prevista
        duracao = s.tempo_estimado_min or blocos_por_veiculo[vid]["duracao_padrao_min"] or 120
        fim = inicio + timedelta(minutes=duracao)
        blocos_por_veiculo[vid]["blocos"].append({
            "solicitacao_id": s.id,
            "codigo": s.codigo or "",
            "tipo": s.tipo or "",
            "parceiro": s.parceiro_nome or "",
            "status": s.status,
            "inicio": inicio.strftime("%Y-%m-%d %H:%M"),
            "fim": fim.strftime("%Y-%m-%d %H:%M"),
            "duracao_min": duracao,
        })

    for vid in blocos_por_veiculo:
        blocos_por_veiculo[vid]["blocos"].sort(key=lambda b: b["inicio"])

    return jsonify({
        "veiculos": list(blocos_por_veiculo.values()),
        "data_inicio": data_inicio.strftime("%Y-%m-%d"),
        "data_fim": data_fim.strftime("%Y-%m-%d"),
    })


@agendamento_bp.route("/api/logistica/agendamento-veiculos/sugestao-rota")
@permission_required("PAGE_LOGISTICA_AGENDAMENTO")
def sugestao_rota():
    """Sugere a ordem das coletas/entregas de um dia usando nearest-neighbor."""
    from math import radians, sin, cos, sqrt, atan2
    data_str = request.args.get("data", "")
    veiculo_id = request.args.get("veiculo_id", "")
    try:
        data = datetime.strptime(data_str, "%Y-%m-%d") if data_str else datetime.now()
    except ValueError:
        data = datetime.now()
    dia_inicio = data.replace(hour=0, minute=0, second=0, microsecond=0)
    dia_fim = dia_inicio + timedelta(days=1)

    q = AgendamentoSolicitacao.query.filter(
        AgendamentoSolicitacao.status.in_(["Alocada", "EmAndamento"]),
        _filtro_solicitacao_visivel(),
        AgendamentoSolicitacao.data_hora_saida_prevista.isnot(None),
        AgendamentoSolicitacao.data_hora_saida_prevista >= dia_inicio,
        AgendamentoSolicitacao.data_hora_saida_prevista < dia_fim,
    )
    if veiculo_id:
        q = q.filter(AgendamentoSolicitacao.veiculo_id == int(veiculo_id))
    rows = q.all()

    if not rows:
        return jsonify({"sugestao": [], "mensagem": "Nenhuma solicitação alocada para este dia."})

    from conferencia_app.config import Config
    base_lat = Config.AGENDAMENTO_BASE_LAT
    base_lng = Config.AGENDAMENTO_BASE_LNG

    def haversine(lat1, lon1, lat2, lon2):
        R = 6371
        dlat = radians(lat2 - lat1)
        dlon = radians(lon2 - lon1)
        a = sin(dlat / 2) ** 2 + cos(radians(lat1)) * cos(radians(lat2)) * sin(dlon / 2) ** 2
        return R * 2 * atan2(sqrt(a), sqrt(1 - a))

    pontos = []
    for r in rows:
        lat = r.destino_latitude or 0
        lng = r.destino_longitude or 0
        pontos.append({
            "id": r.id,
            "codigo": r.codigo or "",
            "tipo": r.tipo or "",
            "parceiro": r.parceiro_nome or "",
            "cidade": r.cidade_destino or "",
            "uf": r.uf_destino or "",
            "lat": float(lat),
            "lng": float(lng),
            "km_estimado": float(r.km_estimado or 0),
            "saida_prevista": r.data_hora_saida_prevista.strftime("%Y-%m-%d %H:%M") if r.data_hora_saida_prevista else "",
        })

    # Nearest-neighbor a partir da base
    ordenados = []
    restantes = list(pontos)
    cur_lat, cur_lng = base_lat, base_lng
    velocidade_media_kmh = 50
    tempo_parada_min = 30

    while restantes:
        melhor = None
        melhor_dist = float("inf")
        for p in restantes:
            if p["lat"] == 0 and p["lng"] == 0:
                d = 9999
            else:
                d = haversine(cur_lat, cur_lng, p["lat"], p["lng"])
            if d < melhor_dist:
                melhor_dist = d
                melhor = p
        restantes.remove(melhor)
        km_trecho = round(melhor_dist * 1.28, 1) if melhor_dist < 9000 else 0
        tempo_trecho_min = round((km_trecho / velocidade_media_kmh) * 60) if km_trecho > 0 else 0
        melhor["ordem"] = len(ordenados) + 1
        melhor["km_trecho"] = km_trecho
        melhor["tempo_trecho_min"] = tempo_trecho_min
        melhor["tempo_parada_min"] = tempo_parada_min
        melhor["tempo_total_min"] = tempo_trecho_min + tempo_parada_min
        ordenados.append(melhor)
        if melhor["lat"] != 0 or melhor["lng"] != 0:
            cur_lat, cur_lng = melhor["lat"], melhor["lng"]

    tempo_acumulado = 0
    for p in ordenados:
        tempo_acumulado += p["tempo_total_min"]
        p["tempo_acumulado_min"] = tempo_acumulado

    return jsonify({"sugestao": ordenados, "data": data.strftime("%Y-%m-%d"), "veiculo_id": veiculo_id or None})


def _auto_transicao_em_andamento():
    """Muda Alocada -> EmAndamento quando a data de saída prevista chegou."""
    hoje = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
    rows = AgendamentoSolicitacao.query.filter(
        AgendamentoSolicitacao.status == "Alocada",
        _filtro_solicitacao_visivel(),
        AgendamentoSolicitacao.data_hora_saida_prevista.isnot(None),
        AgendamentoSolicitacao.data_hora_saida_prevista < hoje + timedelta(days=1),
    ).all()
    for row in rows:
        row.status = "EmAndamento"
        row.atualizado_em = datetime.now()
        _registrar_historico(
            row.id,
            evento="AUTO_EM_ANDAMENTO",
            usuario="sistema",
            status_anterior="Alocada",
            status_novo="EmAndamento",
            detalhe="Status alterado automaticamente: dia da viagem chegou.",
        )
    if rows:
        db.session.commit()


@agendamento_bp.route("/api/logistica/motorista/minhas-viagens")
@permission_required("PAGE_LOGISTICA_MOTORISTA")
def minhas_viagens_motorista():
    """Retorna viagens do motorista logado (usa usuario_username)."""
    username = session.get("username", "")
    sincronizar_motoristas_usuarios(commit=True)
    motorista = AgendamentoMotorista.query.filter_by(usuario_username=username).first()
    if not motorista:
        return jsonify({"viagens": [], "motorista": None})
    rows = (
        _query_solicitacoes_visiveis()
        .filter(
            AgendamentoSolicitacao.motorista_id == motorista.id,
            AgendamentoSolicitacao.status.in_(STATUS_ATIVOS | {"Concluida"}),
        )
        .order_by(AgendamentoSolicitacao.data_hora_saida_prevista.asc())
        .all()
    )
    viagens = []
    for row in rows:
        veiculo = AgendamentoVeiculo.query.get(row.veiculo_id) if row.veiculo_id else None
        viagens.append(_serializar_solicitacao(row, veiculo=veiculo))
    return jsonify({
        "viagens": viagens,
        "motorista": serializar_motorista(motorista),
    })


@agendamento_bp.route("/api/logistica/motorista/finalizar/<int:solicitacao_id>", methods=["POST"])
@permission_required("PAGE_LOGISTICA_MOTORISTA")
def motorista_finalizar_viagem(solicitacao_id: int):
    """Motorista marca viagem como Concluída."""
    username = session.get("username", "")
    sincronizar_motoristas_usuarios(commit=True)
    motorista = AgendamentoMotorista.query.filter_by(usuario_username=username).first()
    if not motorista:
        return jsonify({"error": "Motorista não encontrado para este usuário."}), 404
    row = _get_solicitacao_visivel(solicitacao_id)
    if not row:
        return jsonify({"error": "Solicitação não encontrada."}), 404
    if row.motorista_id != motorista.id:
        return jsonify({"error": "Esta viagem não pertence a você."}), 403
    status_atual = str(row.status or "").strip()
    if status_atual in {"Concluida", "Cancelada"}:
        return jsonify({"error": "Viagem já encerrada."}), 409
    if status_atual not in {"EmRota"}:
        return jsonify({"error": "Você precisa iniciar a viagem antes de finalizar."}), 409
    row.status = "Concluida"
    row.concluido_por = username
    row.concluido_em = datetime.now()
    row.data_hora_retorno_real = datetime.now()
    row.atualizado_em = datetime.now()
    _registrar_historico(
        row.id,
        evento="CONCLUIDA_MOTORISTA",
        usuario=username,
        status_anterior=status_atual,
        status_novo="Concluida",
        detalhe=f"Viagem finalizada pelo motorista {motorista.nome}.",
    )
    db.session.commit()
    veiculo = AgendamentoVeiculo.query.get(row.veiculo_id) if row.veiculo_id else None
    return jsonify({"sucesso": True, "solicitacao": _serializar_solicitacao(row, veiculo=veiculo)})


@agendamento_bp.route("/api/logistica/motorista/iniciar/<int:solicitacao_id>", methods=["POST"])
@permission_required("PAGE_LOGISTICA_MOTORISTA")
def motorista_iniciar_viagem(solicitacao_id: int):
    """Motorista confirma que iniciou a viagem (Alocada/EmAndamento -> EmRota)."""
    username = session.get("username", "")
    sincronizar_motoristas_usuarios(commit=True)
    motorista = AgendamentoMotorista.query.filter_by(usuario_username=username).first()
    if not motorista:
        return jsonify({"error": "Motorista não encontrado para este usuário."}), 404
    row = _get_solicitacao_visivel(solicitacao_id)
    if not row:
        return jsonify({"error": "Solicitação não encontrada."}), 404
    if row.motorista_id != motorista.id:
        return jsonify({"error": "Esta viagem não pertence a você."}), 403
    status_atual = str(row.status or "").strip()
    if status_atual in {"Concluida", "Cancelada"}:
        return jsonify({"error": "Viagem já encerrada."}), 409
    if status_atual not in {"Alocada", "EmAndamento"}:
        return jsonify({"error": "Viagem precisa estar alocada para iniciar."}), 409
    row.status = "EmRota"
    row.data_hora_saida_real = datetime.now()
    row.atualizado_em = datetime.now()
    _registrar_historico(
        row.id,
        evento="INICIADA_MOTORISTA",
        usuario=username,
        status_anterior=status_atual,
        status_novo="EmRota",
        detalhe=f"Viagem iniciada pelo motorista {motorista.nome}.",
    )
    db.session.commit()
    veiculo = AgendamentoVeiculo.query.get(row.veiculo_id) if row.veiculo_id else None
    return jsonify({"sucesso": True, "solicitacao": _serializar_solicitacao(row, veiculo=veiculo)})
