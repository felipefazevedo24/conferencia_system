"""Rotas da funcionalidade de Inventario Inicial da Logistica."""

from __future__ import annotations

import json
import math
import os
import re
import unicodedata
from datetime import datetime
from io import BytesIO
from pathlib import Path

from flask import Blueprint, current_app, jsonify, render_template, request, session, send_file
from openpyxl import Workbook
import requests

from ..auth import has_permission, permission_required, permission_required_any, roles_required
from ..extensions import db
from ..models import (
    ChapaCalculo,
    ChapaCalculoLog,
    ItemNota,
    LogisticaInventarioAjuste,
    LogisticaInventarioAnaliseCausa,
    LogisticaInventarioInicial,
    LogisticaInventarioRelatorioAjuste,
    RELATORIO_AJUSTE_DEPOSITO_TIPOS,
    RELATORIO_AJUSTE_MAX_ITENS,
    RELATORIO_AJUSTE_MOTIVOS,
    RELATORIO_AJUSTE_TIPOS,
)
from ..services.logistica_inventario_relatorio_pdf import gerar_relatorio_ajuste_pdf
from ..services.erp_estoque_service import (
    LocalizacaoEstoqueNaoEncontrada,
    atualizar_localizacao_estoque,
    buscar_consumo_kardex_grv,
    buscar_estoque_grv,
    buscar_ordens_compra_abertas_grv,
    buscar_reservas_produto_acabado_grv,
    custo_medio_para,
    qtde_grv_para,
)
from ..services import logistica_inventario_ajuste_service as ajuste_svc


logistica_inventario_bp = Blueprint("logistica_inventario", __name__)

PERMISSION = "PAGE_LOGISTICA_INVENTARIO"
INVENTARIO_EXPORT_JSON_REL_PATH = "inventario_material_local.json"
UNIDADES_PADRAO = [
    "UN", "PC", "CX", "PCT", "RL", "KG", "G", "MG", "L", "ML", "M", "CM", "MM", "M2", "M3",
]
ESTOQUE_VISOES = {
    "materia_prima": {"label": "Matéria-prima", "familia": "N - 01 - MATÉRIA-PRIMA"},
    "revenda": {"label": "Material para revenda", "familia": "N - 00 - MERCADORIA PARA REVENDA"},
    "produto_acabado": {"label": "Produto acabado", "familia": "N - 04 - PRODUTOS"},
}


def _normalizar_busca(texto: str | None) -> str:
    valor = unicodedata.normalize("NFKD", str(texto or "").strip().upper())
    valor = "".join(ch for ch in valor if not unicodedata.combining(ch))
    return re.sub(r"\s+", " ", valor)


def _unidade_estoque_discreta(unidade: str | None) -> bool:
    codigo = re.sub(r"[^A-Z]", "", str(unidade or "").upper())
    return codigo in {"UN", "UND", "UNIDADE", "PC", "PCA", "PECA", "CX", "CAIXA", "KIT"}


def _nivel_cobertura(saldo_disponivel: float | None, consumo_diario: float) -> tuple[int | None, str, str]:
    if saldo_disponivel is None:
        return None, "sem_estoque", "Sem registro no GRV"
    if consumo_diario <= 0:
        return None, "sem_consumo", "Sem consumo no periodo"
    dias = int(math.ceil(float(saldo_disponivel or 0) / consumo_diario))
    critico = int(current_app.config.get("ESTOQUE_MATERIA_PRIMA_DIAS_CRITICOS", 7))
    atencao = int(current_app.config.get("ESTOQUE_MATERIA_PRIMA_DIAS_ATENCAO", 15))
    if dias <= critico:
        return dias, "critico", "Cobertura critica"
    if dias <= atencao:
        return dias, "atencao", "Monitorar cobertura"
    return dias, "normal", "Cobertura adequada"


def _fmt_registro(row: LogisticaInventarioInicial, incluir_grv: bool = False) -> dict:
    dados = {
        "id": row.id,
        "local_codigo": row.local_codigo,
        "codigo_produto": row.codigo_produto,
        "unidade_medida": row.unidade_medida,
        "quantidade": row.quantidade,
        "lote": row.lote or "",
        "observacao": row.observacao or "",
        "criado_por": row.criado_por,
        "criado_em": row.criado_em.isoformat() if row.criado_em else None,
        "atualizado_em": row.atualizado_em.isoformat() if row.atualizado_em else None,
    }
    # So incluido quando explicitamente pedido (tela de consulta/revisao).
    # A tela de contagem (Novo Inventario) NUNCA pede isso, pra nao vesar a
    # conferencia com o saldo esperado - mesma logica da conferencia cega
    # usada no resto do sistema.
    #
    # O saldo do GRV usado aqui e' o SNAPSHOT gravado no momento da contagem
    # (ver criar_inventario_inicial), nao uma consulta em tempo real - senao
    # uma contagem que batia com o GRV na hora em que foi feita passaria a
    # aparecer como "divergente" so porque o estoque mudou depois (giro
    # normal), sem relacao nenhuma com erro de contagem.
    if incluir_grv:
        qtde_grv = row.qtde_grv_no_momento
        dados["qtde_grv"] = qtde_grv
        dados["grv_consultado_em"] = row.grv_consultado_em.isoformat() if row.grv_consultado_em else None
        dados["divergente"] = (qtde_grv is not None) and (abs(float(row.quantidade or 0) - qtde_grv) > ajuste_svc.TOLERANCIA_DIVERGENCIA)
    return dados


def _build_query(local: str = "", codigo: str = ""):
    query = LogisticaInventarioInicial.query
    if local:
        query = query.filter(LogisticaInventarioInicial.local_codigo.ilike(f"%{local}%"))
    if codigo:
        query = query.filter(LogisticaInventarioInicial.codigo_produto.ilike(f"%{codigo}%"))
    return query


def _token_integracao_inventario_recebido() -> str:
    token = str(request.headers.get("X-Integracao-Token") or "").strip()
    if token:
        return token
    auth = str(request.headers.get("Authorization") or "").strip()
    if auth.lower().startswith("bearer "):
        return auth[7:].strip()
    return str(request.args.get("token") or "").strip()


def _token_integracao_inventario_valido() -> bool:
    # Usa token proprio do inventario; se ausente, reaproveita token da integracao de expedicao.
    esperado = str(
        current_app.config.get("INVENTARIO_INTEGRACAO_TOKEN")
        or current_app.config.get("EXPEDICAO_INTEGRACAO_TOKEN")
        or ""
    ).strip()
    if not esperado:
        return True
    return _token_integracao_inventario_recebido() == esperado


def _gerar_dados_material_local(rows: list[LogisticaInventarioInicial]) -> list[dict]:
    # Mantem apenas a versao mais recente por par (material, local).
    mais_recente_por_chave: dict[tuple[str, str], LogisticaInventarioInicial] = {}
    for row in rows:
        codigo = str(row.codigo_produto or "").strip().upper()
        local = str(row.local_codigo or "").strip().upper()
        if not codigo or not local:
            continue
        chave = (codigo, local)
        atual = mais_recente_por_chave.get(chave)
        if not atual or (row.atualizado_em or row.criado_em or datetime.min) > (atual.atualizado_em or atual.criado_em or datetime.min):
            mais_recente_por_chave[chave] = row

    saida = []
    for (codigo, local), row in sorted(mais_recente_por_chave.items()):
        saida.append(
            {
                "codigo_material": codigo,
                "local": local,
                "quantidade": float(row.quantidade or 0),
                "unidade": str(row.unidade_medida or "UN"),
                "atualizado_em": (row.atualizado_em or row.criado_em).isoformat() if (row.atualizado_em or row.criado_em) else None,
            }
        )
    return saida


def _montar_payload_material_local() -> dict:
    rows = (
        LogisticaInventarioInicial.query
        .order_by(LogisticaInventarioInicial.atualizado_em.desc(), LogisticaInventarioInicial.id.desc())
        .all()
    )
    dados = _gerar_dados_material_local(rows)
    return {
        "sucesso": True,
        "gerado_em": datetime.now().isoformat(),
        "total": len(dados),
        "dados": dados,
    }


def _salvar_snapshot_inventario_json(payload: dict) -> Path:
    static_dir = Path(__file__).resolve().parents[2] / "static"
    static_dir.mkdir(parents=True, exist_ok=True)
    out_path = static_dir / INVENTARIO_EXPORT_JSON_REL_PATH
    out_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    return out_path


def _enviar_para_grv(payload: dict) -> dict:
    api_url = str(
        os.environ.get("INVENTARIO_GRV_API_URL")
        or os.environ.get("ERP_LANCAMENTO_API_URL")
        or current_app.config.get("INVENTARIO_GRV_API_URL")
        or current_app.config.get("ERP_LANCAMENTO_API_URL")
        or ""
    ).strip()
    token = str(
        os.environ.get("INVENTARIO_GRV_API_TOKEN")
        or os.environ.get("ERP_LANCAMENTO_API_TOKEN")
        or current_app.config.get("INVENTARIO_GRV_API_TOKEN")
        or current_app.config.get("ERP_LANCAMENTO_API_TOKEN")
        or ""
    ).strip()
    timeout = int(
        os.environ.get("INVENTARIO_GRV_API_TIMEOUT")
        or os.environ.get("ERP_LANCAMENTO_API_TIMEOUT")
        or current_app.config.get("INVENTARIO_GRV_API_TIMEOUT")
        or current_app.config.get("ERP_LANCAMENTO_API_TIMEOUT")
        or 30
    )

    if not api_url:
        return {"enviado": False, "motivo": "INVENTARIO_GRV_API_URL nao configurada."}

    headers = {
        "Accept": "application/json",
        "Content-Type": "application/json",
        "ngrok-skip-browser-warning": "true",
        "User-Agent": "ColumbiaSync/Inventario-GRV",
    }
    if token:
        headers["Authorization"] = f"Bearer {token}"

    resp = requests.post(api_url, headers=headers, json=payload, timeout=timeout)
    resp.raise_for_status()
    body = resp.json() if (resp.headers.get("content-type", "").lower().find("json") >= 0) else {}
    return {
        "enviado": True,
        "status_code": resp.status_code,
        "resposta": body if isinstance(body, dict) else {},
    }


@logistica_inventario_bp.route("/logistica/inventario")
@permission_required(PERMISSION)
def inventario_home_page():
    return render_template(
        "logistica_inventario_home.html",
        user=session["username"],
        user_role=session.get("role", ""),
    )


@logistica_inventario_bp.route("/logistica/inventario/novo")
@permission_required(PERMISSION)
def inventario_novo_page():
    return render_template(
        "logistica_inventario_inicial.html",
        user=session["username"],
        user_role=session.get("role", ""),
    )


@logistica_inventario_bp.route("/logistica/inventario/consulta")
@permission_required(PERMISSION)
def inventario_consulta_page():
    return render_template(
        "logistica_inventario_consulta.html",
        user=session["username"],
        user_role=session.get("role", ""),
    )


@logistica_inventario_bp.route("/logistica/estoque")
@logistica_inventario_bp.route("/logistica/estoque-materia-prima")
@permission_required(PERMISSION)
def estoque_page():
    return render_template(
        "logistica_estoque_materia_prima.html",
        user=session["username"],
        user_role=session.get("role", ""),
    )


@logistica_inventario_bp.route("/api/logistica/estoque-materia-prima", methods=["GET"])
@logistica_inventario_bp.route("/api/logistica/estoque", methods=["GET"])
@permission_required(PERMISSION)
def estoque_materia_prima_api():
    visao = str(request.args.get("visao") or "materia_prima").strip()
    if visao not in ESTOQUE_VISOES:
        visao = "materia_prima"
    visao_cfg = ESTOQUE_VISOES[visao]
    termo = _normalizar_busca(request.args.get("q"))
    incluir_sem_saldo = str(request.args.get("incluir_sem_saldo") or "").strip().lower() in {"1", "true", "sim", "yes"}
    refresh = str(request.args.get("refresh") or "").strip() == "1"
    try:
        limite = max(1, min(int(request.args.get("limit") or 300), 800))
    except (TypeError, ValueError):
        limite = 300

    nivel_filtro = str(request.args.get("nivel") or "").strip().lower()
    somente_estoque_minimo = str(request.args.get("estoque_minimo") or "").strip().lower() in {"1", "true", "sim", "yes"}
    estoque = buscar_estoque_grv(forcar_atualizacao=refresh)
    familia_alvo = _normalizar_busca(visao_cfg["familia"])
    candidatos = []
    for codigo, item in (estoque.get("por_codigo") or {}).items():
        if _normalizar_busca(item.get("familia")) != familia_alvo:
            continue
        saldo_total = float(item.get("qtde_total") or 0)
        saldo_disponivel = float(item.get("qtde_disponivel") or 0)
        saldo_reservado = float(item.get("qtde_reservada") or 0)
        estoque_minimo = float(item.get("estoque_minimo") or 0)
        lote_economico = float(item.get("lote_economico") or 0)
        haystack = _normalizar_busca(" ".join([codigo, str(item.get("item") or ""), str(item.get("localizacoes") or "")]))
        if termo and termo not in haystack:
            continue
        if not incluir_sem_saldo and saldo_total <= 0 and saldo_disponivel <= 0 and saldo_reservado <= 0:
            continue
        candidatos.append({
            "codigo": codigo,
            "descricao": str(item.get("item") or "").strip(),
            "familia": str(item.get("familia") or "").strip(),
            "grupo": str(item.get("grupo") or "").strip(),
            "unidade": str(item.get("unidade") or "UN").strip() or "UN",
            "saldo_disponivel": saldo_disponivel,
            "saldo_reservado": saldo_reservado,
            "saldo_total": saldo_total,
            "estoque_minimo": estoque_minimo,
            "lote_economico": lote_economico,
            "localizacoes": sorted(set(str(x or "").strip() for x in item.get("localizacoes") or [] if str(x or "").strip())),
        })

    rows = candidatos
    codigos = [row["codigo"] for row in rows]
    consumo_por_codigo = {}
    reservas_por_codigo = {}
    ordens_compra_por_codigo = {}
    if codigos:
        try:
            consumo = buscar_consumo_kardex_grv(codigos=codigos, forcar_atualizacao=refresh)
            consumo_por_codigo = consumo.get("por_codigo") or {}
        except Exception:
            current_app.logger.warning("Nao foi possivel consultar consumo de materia-prima.", exc_info=True)
            consumo_por_codigo = {}
        try:
            reservas_por_codigo = buscar_reservas_produto_acabado_grv(codigos=codigos)
        except Exception:
            current_app.logger.warning("Nao foi possivel consultar reservas de estoque.", exc_info=True)
            reservas_por_codigo = {}

    for row in rows:
        codigo_key = re.sub(r"[^A-Z0-9]", "", row["codigo"].upper())
        consumo_diario = float((consumo_por_codigo.get(codigo_key) or {}).get("consumo_medio_diario") or 0)
        if consumo_diario > 0 and _unidade_estoque_discreta(row["unidade"]):
            consumo_diario = float(math.ceil(consumo_diario))
        cobertura_dias, nivel, nivel_label = _nivel_cobertura(row["saldo_disponivel"], consumo_diario)
        reservas = reservas_por_codigo.get(codigo_key, [])
        row["consumo_diario"] = consumo_diario
        row["cobertura_dias"] = cobertura_dias
        row["nivel"] = nivel
        row["nivel_label"] = nivel_label
        row["reservas"] = [
            {
                "quantidade": float(reserva.get("qtde") or 0),
                "descricao": str(reserva.get("descricao") or "").strip(),
                "orcamento": str(reserva.get("numero_orcamento") or reserva.get("orcamento_descricao") or reserva.get("cod_orcamento") or "").strip(),
                "versao": str(reserva.get("versao_orcamento") or "").strip(),
                "cliente": str(reserva.get("cliente_orcamento") or reserva.get("cliente_os_orcamento") or reserva.get("cliente_os_descricao") or "").strip(),
                "os": str(reserva.get("os_descricao") or reserva.get("os_por_orcamento") or "").strip(),
                "cod_os": str(reserva.get("cod_os_descricao") or reserva.get("cod_os_por_orcamento") or "").strip(),
                "titulo_os": str(reserva.get("titulo_os_descricao") or "").strip(),
            }
            for reserva in reservas[:20]
        ]

    if visao in {"materia_prima", "revenda"}:
        codigos_reposicao = [row["codigo"] for row in rows if row["nivel"] in ("critico", "atencao")]
        if codigos_reposicao:
            try:
                ordens_compra_por_codigo = buscar_ordens_compra_abertas_grv(codigos=codigos_reposicao)
            except Exception:
                current_app.logger.warning("Nao foi possivel consultar OCs abertas para estoque em reposicao.", exc_info=True)
    for row in rows:
        codigo_key = re.sub(r"[^A-Z0-9]", "", row["codigo"].upper())
        row["ordens_compra_abertas"] = ordens_compra_por_codigo.get(codigo_key, [])

    if nivel_filtro:
        rows = [row for row in rows if str(row.get("nivel") or "") == nivel_filtro]
    if somente_estoque_minimo:
        rows = [row for row in rows if float(row.get("estoque_minimo") or 0) > 0]

    prioridade_nivel = {"critico": 0, "atencao": 1, "sem_consumo": 2, "sem_estoque": 3, "normal": 4}
    rows.sort(key=lambda row: (
        0 if visao == "produto_acabado" and float(row.get("saldo_reservado") or 0) > 0 else 1,
        prioridade_nivel.get(str(row.get("nivel") or ""), 9),
        row.get("cobertura_dias") is None,
        row.get("cobertura_dias") or 999999,
        row["codigo"],
    ))
    rows = rows[:limite]

    resumo = {
        "itens": len(rows),
        "total_filtrado": len(rows),
        "criticos": sum(1 for row in rows if row.get("nivel") == "critico"),
        "atencao": sum(1 for row in rows if row.get("nivel") == "atencao"),
        "sem_consumo": sum(1 for row in rows if row.get("nivel") == "sem_consumo"),
        "saldo_disponivel": sum(float(row.get("saldo_disponivel") or 0) for row in rows),
        "saldo_total": sum(float(row.get("saldo_total") or 0) for row in rows),
    }
    return jsonify({
        "visao": visao,
        "visao_label": visao_cfg["label"],
        "familia": visao_cfg["familia"],
        "visoes": ESTOQUE_VISOES,
        "resumo": resumo,
        "items": rows,
        "limit": limite,
    })


@logistica_inventario_bp.route("/logistica/inventario-inicial")
@permission_required(PERMISSION)
def inventario_inicial_legacy_redirect():
    return inventario_novo_page()


@logistica_inventario_bp.route("/api/logistica/inventario-inicial/unidades", methods=["GET"])
@permission_required(PERMISSION)
def listar_unidades_padrao():
    return jsonify({"unidades": UNIDADES_PADRAO})


@logistica_inventario_bp.route("/api/logistica/inventario-inicial", methods=["GET"])
@permission_required(PERMISSION)
def listar_inventario_inicial():
    limite = request.args.get("limit", type=int) or 100
    limite = max(1, min(limite, 500))
    local = (request.args.get("local") or "").strip().lower()
    codigo = (request.args.get("codigo") or "").strip().lower()

    query = _build_query(local=local, codigo=codigo)

    rows = query.order_by(LogisticaInventarioInicial.criado_em.desc()).limit(limite).all()

    # A comparacao com o GRV so e incluida quando explicitamente pedida
    # (tela de consulta) - a tela de contagem nunca passa esse parametro,
    # entao nunca recebe o saldo esperado antes de fechar a contagem.
    # O valor usado e' o snapshot gravado no momento de cada contagem (ver
    # criar_inventario_inicial), nao uma consulta em tempo real - o
    # "forcar_grv"/cache do buscar_estoque_grv() nao se aplica mais aqui.
    incluir_grv = request.args.get("comparar_grv") == "1"

    resposta = {"registros": [_fmt_registro(row, incluir_grv) for row in rows]}
    if incluir_grv:
        # Quantas contagens retornadas nao conseguiram capturar o saldo do
        # GRV no momento em que foram feitas (API fora do ar naquela hora,
        # ou codigo nao encontrado no GRV) - a coluna "Qtd GRV" fica vazia
        # pra essas, sem tentar reconsultar agora.
        resposta["sem_grv_no_momento"] = sum(1 for row in rows if row.qtde_grv_no_momento is None)
    return jsonify(resposta)


@logistica_inventario_bp.route("/api/logistica/inventario-inicial/exportar", methods=["GET"])
@permission_required(PERMISSION)
def exportar_inventario_inicial_excel():
    local = (request.args.get("local") or "").strip().lower()
    codigo = (request.args.get("codigo") or "").strip().lower()
    comparar_grv = request.args.get("comparar_grv") == "1"

    query = _build_query(local=local, codigo=codigo)
    rows = query.order_by(LogisticaInventarioInicial.criado_em.desc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Inventario"

    headers = [
        "Data",
        "Local",
        "Codigo Produto",
        "Unidade",
        "Quantidade",
        "Lote",
        "Observacao",
        "Criado Por",
    ]
    if comparar_grv:
        headers += ["Qtde GRV", "Divergente"]
    ws.append(headers)

    for row in rows:
        linha = [
            row.criado_em.strftime("%d/%m/%Y %H:%M:%S") if row.criado_em else "",
            row.local_codigo,
            row.codigo_produto,
            row.unidade_medida,
            float(row.quantidade or 0),
            row.lote or "",
            row.observacao or "",
            row.criado_por,
        ]
        if comparar_grv:
            qtde_grv = row.qtde_grv_no_momento
            divergente = (qtde_grv is not None) and (abs(float(row.quantidade or 0) - qtde_grv) > ajuste_svc.TOLERANCIA_DIVERGENCIA)
            linha += [qtde_grv if qtde_grv is not None else "N/D", "SIM" if divergente else "NAO"]
        ws.append(linha)

    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            val = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(val))
        ws.column_dimensions[col_letter].width = min(max(12, max_len + 2), 60)

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)

    nome_arquivo = f"inventario_logistica_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(
        stream,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=nome_arquivo,
    )


@logistica_inventario_bp.route("/api/logistica/inventario-inicial", methods=["POST"])
@permission_required(PERMISSION)
def criar_inventario_inicial():
    payload = request.get_json(silent=True) or {}

    local_codigo = str(payload.get("local_codigo") or "").strip().upper()
    codigo_produto = str(payload.get("codigo_produto") or "").strip().upper()
    unidade_medida = str(payload.get("unidade_medida") or "UN").strip().upper() or "UN"
    lote = str(payload.get("lote") or "").strip()
    observacao = str(payload.get("observacao") or "").strip()

    if not local_codigo:
        return jsonify({"error": "Local e obrigatorio."}), 400
    if not codigo_produto:
        return jsonify({"error": "Codigo do produto e obrigatorio."}), 400

    try:
        quantidade = float(str(payload.get("quantidade") or "0").replace(",", "."))
    except (TypeError, ValueError):
        return jsonify({"error": "Quantidade invalida."}), 400

    if quantidade <= 0:
        return jsonify({"error": "Quantidade deve ser maior que zero."}), 400

    row = LogisticaInventarioInicial(
        local_codigo=local_codigo,
        codigo_produto=codigo_produto,
        unidade_medida=unidade_medida[:20],
        quantidade=quantidade,
        lote=lote[:120] if lote else None,
        observacao=observacao[:800] if observacao else None,
        criado_por=session.get("username", "sistema"),
        atualizado_em=datetime.now(),
    )
    db.session.add(row)
    db.session.commit()

    # Sincroniza a localizacao no ERP (tproduto) - nao falha a contagem se
    # o ERP estiver fora do ar, so avisa: o registro do inventario ja foi
    # salvo com sucesso independente disso.
    localizacao_erp = {"sincronizado": False}
    try:
        localizacao_erp["resposta"] = atualizar_localizacao_estoque(codigo_produto, local_codigo)
        localizacao_erp["sincronizado"] = True
    except LocalizacaoEstoqueNaoEncontrada as exc:
        localizacao_erp["erro"] = str(exc)
    except Exception as exc:  # noqa: BLE001
        current_app.logger.warning(
            "Falha ao sincronizar localização de estoque no ERP (código=%s, local=%s): %s",
            codigo_produto, local_codigo, exc,
        )
        localizacao_erp["erro"] = str(exc)

    # Consulta o saldo e o custo medio do GRV UMA VEZ, no exato momento da
    # contagem, e grava esse snapshot no proprio registro
    # (qtde_grv_no_momento/custo_medio_no_momento/grv_consultado_em) - a
    # tela de consulta (Inventario Realizado) usa esse valor gravado em vez
    # de reconsultar o GRV em tempo real depois, senao uma contagem que
    # batia com o GRV na hora em que foi feita passaria a aparecer como
    # divergente so porque o estoque girou normalmente depois. Tambem
    # detecta divergencia e, se houver, abre automaticamente um ajuste no
    # Modulo 02 (Validacao) pro gestor revisar, com o custo medio junto pra
    # calcular o impacto financeiro (R$) - nao falha a contagem se o GRV
    # estiver indisponivel, so fica sem snapshot e sem ajuste.
    ajuste_aberto = None
    try:
        estoque_grv = buscar_estoque_grv()
        qtde_grv = qtde_grv_para(row.codigo_produto, row.local_codigo, estoque_grv)
        custo_medio = custo_medio_para(row.codigo_produto, row.local_codigo, estoque_grv)
        row.qtde_grv_no_momento = qtde_grv
        row.custo_medio_no_momento = custo_medio
        row.grv_consultado_em = datetime.now()
        db.session.commit()
        ajuste = ajuste_svc.detectar_divergencia(row, qtde_grv, custo_medio)
        if ajuste:
            ajuste_aberto = {"id": ajuste.id, "diferenca": ajuste.diferenca}
    except Exception as exc:  # noqa: BLE001
        current_app.logger.warning(
            "Falha ao consultar o GRV pra gravar snapshot/detectar divergencia (código=%s, local=%s): %s",
            codigo_produto, local_codigo, exc,
        )

    return jsonify({
        "sucesso": True,
        "registro": _fmt_registro(row),
        "localizacao_erp": localizacao_erp,
        "ajuste_aberto": ajuste_aberto,
    }), 201


@logistica_inventario_bp.route("/api/integracao/inventario/material-local", methods=["GET"])
def inventario_material_local_integracao():
    if not _token_integracao_inventario_valido():
        return jsonify({"sucesso": False, "erro": "Token de integracao invalido."}), 401
    return jsonify(_montar_payload_material_local())


@logistica_inventario_bp.route("/api/logistica/inventario-inicial/sincronizar-grv", methods=["POST"])
@permission_required(PERMISSION)
def sincronizar_inventario_grv():
    payload = _montar_payload_material_local()
    _salvar_snapshot_inventario_json(payload)

    grv = {}
    try:
        grv = _enviar_para_grv(payload)
    except Exception as exc:
        grv = {"enviado": False, "erro": str(exc)}

    base_url = request.url_root.rstrip("/")
    return jsonify(
        {
            "sucesso": True,
            "total": payload.get("total", 0),
            "snapshot_url": f"{base_url}/static/{INVENTARIO_EXPORT_JSON_REL_PATH}",
            "integracao_url": f"{base_url}/api/integracao/inventario/material-local",
            "grv": grv,
        }
    )


# ── Fluxo de ajuste de estoque (Modulos 02-04, itens divergentes) ─────────
PERMISSION_VALIDACAO = "PAGE_LOGISTICA_INVENTARIO_VALIDACAO"
PERMISSION_FINANCE = "PAGE_LOGISTICA_INVENTARIO_FINANCE"
PERMISSION_FISCAL = "PAGE_LOGISTICA_INVENTARIO_FISCAL"
PERMISSION_PULAR_ETAPA = "PAGE_LOGISTICA_INVENTARIO_PULAR_ETAPA"

# Mesmos rotulos exibidos na tela (ver LABEL_STATUS em
# logistica_inventario_ajustes.html) - usado na exportacao Excel pra nao
# mostrar o nome interno do status_modulo pro usuario final.
LABEL_STATUS_AJUSTE = {
    "Validacao": "Aguardando gestor",
    "Relatorio": "Aguardando relatório",
    "Finance": "Aguardando Finance",
    "Fiscal": "Aguardando Fiscal",
    "Concluido": "Concluído",
    "Descartado": "Dif. Improcedente",
}


def _fmt_ajuste(a) -> dict:
    # Impacto financeiro (R$) da divergencia = diferenca (qtde) * custo
    # medio - calculado aqui na hora, nao armazenado (evita duplicar dado
    # derivado). Fica None se o GRV nao tinha custo pro codigo naquele
    # momento (custo_medio None) - a tela mostra "—" nesse caso, sem tentar
    # adivinhar um valor.
    diferenca_valor = (a.diferenca * a.custo_medio) if a.custo_medio is not None else None
    return {
        "id": a.id,
        "codigo_produto": a.codigo_produto,
        "local_codigo": a.local_codigo,
        "unidade_medida": a.unidade_medida,
        "qtde_contada": a.qtde_contada,
        "qtde_estoque_no_momento": a.qtde_estoque_no_momento,
        "diferenca": a.diferenca,
        "custo_medio": a.custo_medio,
        "diferenca_valor": diferenca_valor,
        "status_modulo": a.status_modulo,
        "status_slug": a.status_slug,
        "criado_em": a.criado_em.strftime("%d/%m/%Y %H:%M") if a.criado_em else None,
        "gestor_justificativa": a.gestor_justificativa,
        "gestor_confirmado_em": a.gestor_confirmado_em.strftime("%d/%m/%Y %H:%M") if a.gestor_confirmado_em else None,
        "gestor_confirmado_por": a.gestor_confirmado_por,
        "finance_observacao": a.finance_observacao,
        "finance_concluido_em": a.finance_concluido_em.strftime("%d/%m/%Y %H:%M") if a.finance_concluido_em else None,
        "finance_concluido_por": a.finance_concluido_por,
        "fiscal_nf_numero": a.fiscal_nf_numero,
        "fiscal_concluido_em": a.fiscal_concluido_em.strftime("%d/%m/%Y %H:%M") if a.fiscal_concluido_em else None,
        "fiscal_concluido_por": a.fiscal_concluido_por,
        # So indica se tem analise de causa raiz pendente/concluida - a
        # tela de Ajuste de Estoque mostra so um badge, os detalhes ficam
        # na tela dedicada de Analise de Causa Raiz.
        "analise_causa_status": a.analise_causa.status if a.analise_causa else None,
        "relatorio_id": a.relatorio_id,
        "relatorio_numero_documento": a.relatorio.numero_documento if a.relatorio else None,
    }


def _fmt_analise_causa(a: LogisticaInventarioAnaliseCausa) -> dict:
    ajuste = a.ajuste
    return {
        "id": a.id,
        "ajuste_id": a.ajuste_id,
        "status": a.status,
        "motivo_causa_raiz": a.motivo_causa_raiz,
        "solicitado_por": a.solicitado_por,
        "solicitado_em": a.solicitado_em.strftime("%d/%m/%Y %H:%M") if a.solicitado_em else None,
        "analisado_por": a.analisado_por,
        "analisado_em": a.analisado_em.strftime("%d/%m/%Y %H:%M") if a.analisado_em else None,
        "codigo_produto": ajuste.codigo_produto if ajuste else None,
        "local_codigo": ajuste.local_codigo if ajuste else None,
        "unidade_medida": ajuste.unidade_medida if ajuste else None,
        "qtde_contada": ajuste.qtde_contada if ajuste else None,
        "qtde_estoque_no_momento": ajuste.qtde_estoque_no_momento if ajuste else None,
        "diferenca": ajuste.diferenca if ajuste else None,
        "ajuste_status_modulo": ajuste.status_modulo if ajuste else None,
    }


@logistica_inventario_bp.route("/logistica/inventario/ajustes")
@permission_required_any(PERMISSION, PERMISSION_VALIDACAO, PERMISSION_FINANCE, PERMISSION_FISCAL)
def inventario_ajustes_page():
    return render_template(
        "logistica_inventario_ajustes.html",
        user=session["username"],
        user_role=session.get("role", ""),
        pode_validar=has_permission(PERMISSION_VALIDACAO),
        pode_finance=has_permission(PERMISSION_FINANCE),
        pode_fiscal=has_permission(PERMISSION_FISCAL),
        pode_pular_etapa=has_permission(PERMISSION_PULAR_ETAPA),
    )


@logistica_inventario_bp.route("/api/logistica/inventario-ajustes", methods=["GET"])
@permission_required_any(PERMISSION, PERMISSION_VALIDACAO, PERMISSION_FINANCE, PERMISSION_FISCAL)
def api_listar_ajustes():
    status_modulo = request.args.get("status") or None
    ajustes = ajuste_svc.listar_ajustes(status_modulo=status_modulo)
    return jsonify({
        "ajustes": [_fmt_ajuste(a) for a in ajustes],
        "pode_pular_etapa": has_permission(PERMISSION_PULAR_ETAPA),
        "pode_validar": has_permission(PERMISSION_VALIDACAO),
        "pode_finance": has_permission(PERMISSION_FINANCE),
        "pode_fiscal": has_permission(PERMISSION_FISCAL),
    })


@logistica_inventario_bp.route("/api/logistica/inventario-ajustes/exportar", methods=["GET"])
@permission_required_any(PERMISSION, PERMISSION_VALIDACAO, PERMISSION_FINANCE, PERMISSION_FISCAL)
def exportar_inventario_ajustes_excel():
    status_modulo = request.args.get("status") or None
    busca = request.args.get("busca") or ""
    ajustes = ajuste_svc.listar_ajustes(status_modulo=status_modulo, busca=busca)

    wb = Workbook()
    ws = wb.active
    ws.title = "Ajustes de Estoque"

    headers = [
        "Detectado em",
        "Local",
        "Cód. Produto",
        "Unidade",
        "Qtd Contada",
        "Qtd Estoque",
        "Diferença",
        "Custo Médio",
        "Diferença R$",
        "Status",
        "Gestor - Justificativa",
        "Gestor - Confirmado em",
        "Gestor - Confirmado por",
        "Finance - Observação",
        "Finance - Concluído em",
        "Finance - Concluído por",
        "Fiscal - NF",
        "Fiscal - Concluído em",
        "Fiscal - Concluído por",
    ]
    ws.append(headers)

    # Mesmo formato de moeda com sinal (+/-) que a tela mostra (ex.: "+R$
    # 69,96" / "-R$ 205,82") - sem isso o Excel exibe so o numero cru
    # (69.96), o que parecia "sem valor" comparado com a tela.
    FORMATO_MOEDA_SEM_SINAL = '"R$" #,##0.00'
    FORMATO_MOEDA_COM_SINAL = '+"R$" #,##0.00;-"R$" #,##0.00'
    COL_CUSTO_MEDIO = 8   # H
    COL_DIFERENCA_VALOR = 9  # I

    for a in ajustes:
        diferenca_valor = (a.diferenca * a.custo_medio) if a.custo_medio is not None else None
        linha = ws.max_row + 1
        ws.append([
            a.criado_em.strftime("%d/%m/%Y %H:%M:%S") if a.criado_em else "",
            a.local_codigo,
            a.codigo_produto,
            a.unidade_medida,
            float(a.qtde_contada or 0),
            float(a.qtde_estoque_no_momento or 0),
            float(a.diferenca or 0),
            a.custo_medio if a.custo_medio is not None else "—",
            diferenca_valor if diferenca_valor is not None else "—",
            LABEL_STATUS_AJUSTE.get(a.status_modulo, a.status_modulo),
            a.gestor_justificativa or "",
            a.gestor_confirmado_em.strftime("%d/%m/%Y %H:%M:%S") if a.gestor_confirmado_em else "",
            a.gestor_confirmado_por or "",
            a.finance_observacao or "",
            a.finance_concluido_em.strftime("%d/%m/%Y %H:%M:%S") if a.finance_concluido_em else "",
            a.finance_concluido_por or "",
            a.fiscal_nf_numero or "",
            a.fiscal_concluido_em.strftime("%d/%m/%Y %H:%M:%S") if a.fiscal_concluido_em else "",
            a.fiscal_concluido_por or "",
        ])
        if a.custo_medio is not None:
            ws.cell(row=linha, column=COL_CUSTO_MEDIO).number_format = FORMATO_MOEDA_SEM_SINAL
        if diferenca_valor is not None:
            ws.cell(row=linha, column=COL_DIFERENCA_VALOR).number_format = FORMATO_MOEDA_COM_SINAL

    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            val = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(val))
        ws.column_dimensions[col_letter].width = min(max(12, max_len + 2), 60)

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)

    nome_arquivo = f"ajustes_estoque_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(
        stream,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=nome_arquivo,
    )


@logistica_inventario_bp.route("/api/logistica/inventario-ajustes/<int:ajuste_id>/confirmar", methods=["POST"])
@permission_required_any(PERMISSION, PERMISSION_VALIDACAO, PERMISSION_FINANCE, PERMISSION_FISCAL)
def api_confirmar_ajuste(ajuste_id):
    if not has_permission(PERMISSION_VALIDACAO):
        return jsonify({"error": "Você não tem permissão pra validar divergências - fale com a gerência."}), 403
    ajuste = db.session.get(LogisticaInventarioAjuste, ajuste_id)
    if not ajuste:
        return jsonify({"error": "Ajuste não encontrado."}), 404
    payload = request.get_json(silent=True) or {}
    try:
        ajuste = ajuste_svc.confirmar_divergencia(
            ajuste,
            session.get("username", "desconhecido"),
            payload.get("justificativa"),
            solicitar_analise_causa=bool(payload.get("solicitar_analise_causa")),
        )
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400
    mensagem = "Divergência confirmada - aguardando entrar no relatório (FORM-08.52) pro Finance."
    if payload.get("solicitar_analise_causa"):
        mensagem += " Análise de causa raiz solicitada."
    return jsonify({"message": mensagem, "ajuste": _fmt_ajuste(ajuste)})


# ── Relatorio de Ajuste (FORM-08.52 "Ajuste para Faturamento") - gera o ───
# formulario formal em PDF pra um LOTE de divergencias de uma vez, e ja
# confirma todas pro Finance juntas (substitui o "Confirmar divergencia"
# item a item quando o gestor precisa mandar varias juntas com o mesmo
# documento).
@logistica_inventario_bp.route("/api/logistica/inventario-ajustes/relatorio/opcoes", methods=["GET"])
@permission_required_any(PERMISSION, PERMISSION_VALIDACAO, PERMISSION_FINANCE, PERMISSION_FISCAL)
def api_relatorio_ajuste_opcoes():
    return jsonify({
        "tipos_ajuste": RELATORIO_AJUSTE_TIPOS,
        "motivos_ajuste": RELATORIO_AJUSTE_MOTIVOS,
        "deposito_tipos": RELATORIO_AJUSTE_DEPOSITO_TIPOS,
        "max_itens": RELATORIO_AJUSTE_MAX_ITENS,
    })


@logistica_inventario_bp.route("/api/logistica/inventario-ajustes/relatorio", methods=["POST"])
@permission_required_any(PERMISSION, PERMISSION_VALIDACAO, PERMISSION_FINANCE, PERMISSION_FISCAL)
def api_gerar_relatorio_ajuste():
    if not has_permission(PERMISSION_VALIDACAO):
        return jsonify({"error": "Você não tem permissão pra validar divergências - fale com a gerência."}), 403
    payload = request.get_json(silent=True) or {}
    ajuste_ids = payload.get("ajuste_ids") or []
    try:
        ajuste_ids = [int(i) for i in ajuste_ids]
    except (TypeError, ValueError):
        return jsonify({"error": "Lista de itens inválida."}), 400

    try:
        relatorio = ajuste_svc.gerar_relatorio_ajuste(
            ajuste_ids,
            session.get("username", "desconhecido"),
            tipo_ajuste=payload.get("tipo_ajuste"),
            tipo_ajuste_detalhe=payload.get("tipo_ajuste_detalhe"),
            motivo_ajuste=payload.get("motivo_ajuste"),
            motivo_ajuste_detalhe=payload.get("motivo_ajuste_detalhe"),
            deposito_tipo=payload.get("deposito_tipo"),
            deposito_local=payload.get("deposito_local"),
            responsavel=payload.get("responsavel"),
            solicitante=payload.get("solicitante"),
            depto=payload.get("depto"),
            observacoes_ajuste=payload.get("observacoes_ajuste"),
            observacoes_itens=payload.get("observacoes_itens"),
            solicitar_analise_causa=bool(payload.get("solicitar_analise_causa")),
        )
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400

    return jsonify({
        "message": f"Relatório {relatorio.numero_documento} gerado - {len(ajuste_ids)} item(ns) enviado(s) pro Finance.",
        "relatorio_id": relatorio.id,
        "numero_documento": relatorio.numero_documento,
        "pdf_url": f"/api/logistica/inventario-ajustes/relatorio/{relatorio.id}.pdf",
    })


@logistica_inventario_bp.route("/api/logistica/inventario-ajustes/relatorio/<int:relatorio_id>.pdf", methods=["GET"])
@permission_required_any(PERMISSION, PERMISSION_VALIDACAO, PERMISSION_FINANCE, PERMISSION_FISCAL)
def api_relatorio_ajuste_pdf(relatorio_id):
    relatorio = ajuste_svc.buscar_relatorio_ajuste(relatorio_id)
    if not relatorio:
        return jsonify({"error": "Relatório não encontrado."}), 404
    try:
        pdf_bytes = gerar_relatorio_ajuste_pdf(relatorio, relatorio.ajustes)
    except Exception as exc:
        current_app.logger.exception("Falha ao gerar PDF do relatório de ajuste %s", relatorio_id)
        return jsonify({"error": f"Falha ao gerar PDF: {exc}"}), 502
    nome_arquivo = f"FORM-08.52_{relatorio.numero_documento.replace('/', '-')}.pdf"
    return send_file(
        BytesIO(pdf_bytes),
        mimetype="application/pdf",
        as_attachment=False,
        download_name=nome_arquivo,
    )


@logistica_inventario_bp.route("/api/logistica/inventario-ajustes/<int:ajuste_id>/descartar", methods=["POST"])
@permission_required_any(PERMISSION, PERMISSION_VALIDACAO, PERMISSION_FINANCE, PERMISSION_FISCAL)
def api_descartar_ajuste(ajuste_id):
    if not has_permission(PERMISSION_VALIDACAO):
        return jsonify({"error": "Você não tem permissão pra validar divergências - fale com a gerência."}), 403
    ajuste = db.session.get(LogisticaInventarioAjuste, ajuste_id)
    if not ajuste:
        return jsonify({"error": "Ajuste não encontrado."}), 404
    payload = request.get_json(silent=True) or {}
    try:
        ajuste = ajuste_svc.descartar_divergencia(ajuste, session.get("username", "desconhecido"), payload.get("motivo"))
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400
    return jsonify({"message": "Diferença marcada como improcedente.", "ajuste": _fmt_ajuste(ajuste)})

@logistica_inventario_bp.route("/api/logistica/inventario-ajustes/<int:ajuste_id>/finance-concluir", methods=["POST"])
def api_finance_concluir_ajuste(ajuste_id):
    if not has_permission(PERMISSION_FINANCE):
        return jsonify({"error": "Você não tem permissão de Finance nesse fluxo - fale com a gerência."}), 403
    ajuste = LogisticaInventarioAjuste.query.get(ajuste_id)
    if not ajuste:
        return jsonify({"error": "Ajuste não encontrado."}), 404
    payload = request.get_json(silent=True) or {}
    try:
        ajuste = ajuste_svc.concluir_finance(ajuste, session.get("username", "desconhecido"), payload.get("observacao"))
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400
    return jsonify({"message": "Ajuste de estoque confirmado. Liberado pro Fiscal.", "ajuste": _fmt_ajuste(ajuste)})


@logistica_inventario_bp.route("/api/logistica/inventario-ajustes/<int:ajuste_id>/fiscal-concluir", methods=["POST"])
@permission_required_any(PERMISSION, PERMISSION_VALIDACAO, PERMISSION_FINANCE, PERMISSION_FISCAL)
def api_fiscal_concluir_ajuste(ajuste_id):
    if not has_permission(PERMISSION_FISCAL):
        return jsonify({"error": "Você não tem permissão de Fiscal nesse fluxo - fale com a gerência."}), 403
    ajuste = LogisticaInventarioAjuste.query.get(ajuste_id)
    if not ajuste:
        return jsonify({"error": "Ajuste não encontrado."}), 404
    payload = request.get_json(silent=True) or {}
    try:
        ajuste = ajuste_svc.concluir_fiscal(ajuste, session.get("username", "desconhecido"), payload.get("nf_numero"))
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400
    return jsonify({"message": "NF de ajuste confirmada. Processo concluído.", "ajuste": _fmt_ajuste(ajuste)})


@logistica_inventario_bp.route("/api/logistica/inventario-ajustes/<int:ajuste_id>/atualizar-grv", methods=["POST"])
@roles_required("Admin")
def api_atualizar_grv_ajuste(ajuste_id):
    ajuste = db.session.get(LogisticaInventarioAjuste, ajuste_id)
    if not ajuste:
        return jsonify({"error": "Ajuste não encontrado."}), 404
    try:
        estoque_grv = buscar_estoque_grv(forcar_atualizacao=True)
        qtde_grv = qtde_grv_para(ajuste.codigo_produto, ajuste.local_codigo, estoque_grv)
        if qtde_grv is None:
            return jsonify({"error": "Produto não encontrado no estoque atual do GRV."}), 404
        custo_medio = custo_medio_para(ajuste.codigo_produto, ajuste.local_codigo, estoque_grv)
        ajuste.qtde_estoque_no_momento = qtde_grv
        ajuste.diferenca = float(ajuste.qtde_contada or 0) - qtde_grv
        ajuste.custo_medio = custo_medio
        if ajuste.contagem_id:
            contagem = db.session.get(LogisticaInventarioInicial, ajuste.contagem_id)
            if contagem:
                contagem.qtde_grv_no_momento = qtde_grv
                contagem.custo_medio_no_momento = custo_medio
                contagem.grv_consultado_em = datetime.now()
        db.session.commit()
    except Exception as exc:
        db.session.rollback()
        return jsonify({"error": f"Não foi possível atualizar o GRV: {exc}"}), 502
    return jsonify({"message": "Quantidade do GRV atualizada.", "ajuste": _fmt_ajuste(ajuste)})


@logistica_inventario_bp.route("/api/logistica/inventario-ajustes/<int:ajuste_id>/estornar", methods=["POST"])
@roles_required("Admin")
def api_estornar_ajuste(ajuste_id):
    ajuste = db.session.get(LogisticaInventarioAjuste, ajuste_id)
    if not ajuste:
        return jsonify({"error": "Ajuste não encontrado."}), 404
    ajuste = ajuste_svc.estornar_para_validacao(ajuste)
    return jsonify({"message": "Ajuste estornado para validação.", "ajuste": _fmt_ajuste(ajuste)})


@logistica_inventario_bp.route("/api/logistica/inventario-ajustes/<int:ajuste_id>/pular-etapa", methods=["POST"])
@permission_required_any(PERMISSION, PERMISSION_VALIDACAO, PERMISSION_FINANCE, PERMISSION_FISCAL, PERMISSION_PULAR_ETAPA)
def api_pular_etapa_ajuste(ajuste_id):
    """"Pular Etapa" - avanço SEM validação nenhuma pra próxima etapa do
    ajuste, reservado pra casos excepcionais. Exige permissão extra de
    gerência (PAGE_LOGISTICA_INVENTARIO_PULAR_ETAPA), além do acesso normal
    ao módulo."""
    if not has_permission(PERMISSION_PULAR_ETAPA):
        return jsonify({"error": "Você não tem permissão pra usar o Pular Etapa - fale com a gerência."}), 403
    ajuste = LogisticaInventarioAjuste.query.get(ajuste_id)
    if not ajuste:
        return jsonify({"error": "Ajuste não encontrado."}), 404
    usuario = session.get("username", "desconhecido")
    try:
        ajuste = ajuste_svc.pular_etapa(ajuste, usuario)
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400
    return jsonify({"message": f"Etapa avançada para {ajuste.status_modulo}.", "ajuste": _fmt_ajuste(ajuste)})


# ── Analise de Causa Raiz (fila separada, nao bloqueia Finance/Fiscal) ────
@logistica_inventario_bp.route("/logistica/inventario/analise-causa")
@permission_required_any(PERMISSION, PERMISSION_VALIDACAO, PERMISSION_FINANCE, PERMISSION_FISCAL)
def inventario_analise_causa_page():
    return render_template(
        "logistica_inventario_analise_causa.html",
        user=session["username"],
        user_role=session.get("role", ""),
    )


@logistica_inventario_bp.route("/api/logistica/inventario-analise-causa", methods=["GET"])
@permission_required_any(PERMISSION, PERMISSION_VALIDACAO, PERMISSION_FINANCE, PERMISSION_FISCAL)
def api_listar_analise_causa():
    status = request.args.get("status") or None
    analises = ajuste_svc.listar_analises_causa(status=status)
    return jsonify({"analises": [_fmt_analise_causa(a) for a in analises]})


@logistica_inventario_bp.route("/api/logistica/inventario-analise-causa/<int:analise_id>/preencher", methods=["POST"])
@permission_required_any(PERMISSION, PERMISSION_VALIDACAO, PERMISSION_FINANCE, PERMISSION_FISCAL)
def api_preencher_analise_causa(analise_id):
    analise = db.session.get(LogisticaInventarioAnaliseCausa, analise_id)
    if not analise:
        return jsonify({"error": "Análise não encontrada."}), 404
    payload = request.get_json(silent=True) or {}
    try:
        analise = ajuste_svc.preencher_analise_causa(analise, payload.get("motivo"), session.get("username", "desconhecido"))
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400
    return jsonify({"message": "Análise de causa raiz registrada.", "analise": _fmt_analise_causa(analise)})


# ===================================================================
# Controle de Chapas
# ===================================================================
_FATOR_PESO_KG = {
    "KG": 1.0, "KGM": 1.0,
    "T": 1000.0, "TO": 1000.0, "TON": 1000.0, "TONELADA": 1000.0, "TONELADAS": 1000.0,
}

# Densidades (g/cm3) - mesmas famílias da calculadora Vimetal.
CHAPA_DENSIDADES = {
    "aco_carbono": 7.85, "inox": 8.00, "aluminio": 2.71,
    "cobre": 8.96, "latao": 8.53, "bronze": 8.80,
}


def _chapa_kg_do_item(item) -> float:
    """Peso da NF em KG (a NF de chapa vem em KG ou T; T converte para KG)."""
    unidade = re.sub(r"[^A-Z]", "", str(item.unidade_comercial or "").upper())
    return float(item.qtd_real or 0) * _FATOR_PESO_KG.get(unidade, 1.0)


def _chapa_lotes_por_item(itens) -> dict[int, str]:
    """Best-effort: mapa item_id -> lote (do GRV). Fallback fica com a AR."""
    try:
        from ..services.erp_lancamento_service import buscar_entradas_chapa_lote
    except Exception:
        return {}
    try:
        entradas = buscar_entradas_chapa_lote(itens) or []
    except Exception:
        return {}
    # Indexa lote por (numero_nota, codigo/descricao normalizados).
    por_chave: dict[tuple[str, str], str] = {}
    for entrada in entradas:
        nota = str(entrada.get("numero_nota") or "").strip()
        for it in entrada.get("itens") or []:
            lote = str(it.get("lote") or "").strip()
            if not lote:
                continue
            for campo in (it.get("cod_interno"), it.get("descricao")):
                chave = _normalizar_busca(campo)
                if chave:
                    por_chave[(nota, chave)] = lote
    saida: dict[int, str] = {}
    for item in itens:
        nota = str(item.numero_nota or "").strip()
        for campo in (item.codigo_grv, item.codigo, item.descricao):
            lote = por_chave.get((nota, _normalizar_busca(campo)))
            if lote:
                saida[item.id] = lote
                break
    return saida


def _chapa_serializar_calculo(calc: ChapaCalculo | None) -> dict:
    if not calc:
        return {"material": "", "formato": "", "dimensoes": None, "peso_por_peca": None, "calculo_id": None,
                "atualizado_por": "", "atualizado_em": None}
    try:
        dimensoes = json.loads(calc.dimensoes) if calc.dimensoes else None
    except Exception:
        dimensoes = None
    return {
        "material": calc.material or "",
        "formato": calc.formato or "",
        "dimensoes": dimensoes,
        "peso_por_peca": calc.peso_por_peca,
        "calculo_id": calc.id,
        "atualizado_por": calc.atualizado_por or calc.criado_por or "",
        "atualizado_em": calc.atualizado_em.isoformat() if calc.atualizado_em else None,
    }


@logistica_inventario_bp.route("/logistica/estoque/chapas")
@permission_required(PERMISSION)
def estoque_chapas_page():
    return render_template(
        "logistica_estoque_chapas.html",
        user=session["username"],
        user_role=session.get("role", ""),
    )


@logistica_inventario_bp.route("/api/logistica/chapas", methods=["GET"])
@permission_required(PERMISSION)
def api_chapas_listar():
    termo = _normalizar_busca(request.args.get("q"))
    # Chapas = itens que passaram pela conferência cega marcados como chapa
    # (o conferente selecionou "é chapa" e informou a UND -> qtd_chapas_und) E
    # que estão LANÇADOS no Sync. Ao estornar o lançamento/conferência, o item
    # sai daqui automaticamente (deixa de estar "Lançado") e só volta ao ser
    # lançado de novo.
    itens = (
        ItemNota.query
        .filter(ItemNota.qtd_chapas_und.isnot(None))
        .filter(ItemNota.qtd_chapas_und > 0)
        .filter(ItemNota.status == "Lançado")
        .order_by(ItemNota.id.desc())
        .all()
    )
    calc_por_item = {c.item_nota_id: c for c in ChapaCalculo.query.all() if c.item_nota_id is not None}
    lote_por_item = _chapa_lotes_por_item(itens)

    # Saldo/saída/reservado por LOTE (best-effort do GRV; vazio se a bridge não responder).
    try:
        from ..services.erp_estoque_service import buscar_saldo_chapa_por_lote
        saldo_info = buscar_saldo_chapa_por_lote([(i.codigo_grv or i.codigo) for i in itens])
    except Exception:
        saldo_info = {"por_lote": [], "por_codigo": {}, "fontes": {}}

    def _codnorm(c):
        return re.sub(r"[^A-Z0-9]", "", str(c or "").upper())

    # Saída/reservado SÓ por lote exato. Nada de distribuir o total do código
    # (isso somava o histórico inteiro do GRV num lote novo e zerava ele).
    por_lote = {}
    for r in saldo_info.get("por_lote") or []:
        chave = (_codnorm(r.get("codigo")), str(r.get("lote") or "").strip())
        por_lote[chave] = {"kg_saida": float(r.get("kg_saida") or 0), "kg_reservado": float(r.get("kg_reservado") or 0)}

    # Saldo REAL por código (GRV, depósito principal). É a informação básica:
    # em estoque / reservado / disponível de verdade. A baixa já vem refletida
    # aqui (o GRV decrementa o saldo), então não recalculamos nada por fora.
    saldo_cod = {_codnorm(c): v for c, v in (saldo_info.get("saldo_codigo") or {}).items()}
    reservas_os_map = {_codnorm(c): (lst or []) for c, lst in (saldo_info.get("reservas_os") or {}).items()}

    linhas = []
    for i in itens:
        codigo = i.codigo_grv or i.codigo
        calc = calc_por_item.get(i.id)
        calc_dict = _chapa_serializar_calculo(calc)
        lote = lote_por_item.get(i.id) or i.numero_lancamento or ""
        info = por_lote.get((_codnorm(codigo), str(lote).strip()), {})
        linhas.append({
            "item_id": i.id, "numero_nota": i.numero_nota, "codigo": codigo,
            "codigo_norm": _codnorm(codigo), "descricao": i.descricao,
            "fornecedor": i.fornecedor or "", "ar": i.numero_lancamento or "",
            "lote": lote,
            "conferente": i.usuario_conferencia or "", "unidade_nf": i.unidade_comercial or "",
            "kg_nf": _chapa_kg_do_item(i), "und": float(i.qtd_chapas_und or 0),
            "peso": float(calc.peso_por_peca) if (calc and calc.peso_por_peca) else None,
            "kg_saida": info.get("kg_saida", 0.0), "kg_reservado": info.get("kg_reservado", 0.0),
            **calc_dict,
        })

    itens_out = []
    for l in linhas:
        kg_nf = l["kg_nf"]; und = l["und"]; peso = l["peso"]
        real_cod = saldo_cod.get(l["codigo_norm"])
        kg_saida = float(l.get("kg_saida") or 0)
        kg_res = float(l.get("kg_reservado") or 0)
        kg_saldo = max(kg_nf - kg_saida, 0.0)
        kg_disp = max(kg_saldo - kg_res, 0.0)
        peso_calc_total = (peso * und) if (peso and und > 0) else None
        pct_diff = ((peso_calc_total - kg_nf) / kg_nf * 100.0) if (peso_calc_total is not None and kg_nf > 0) else None
        # nº de chapas = UND informada menos as baixadas (kg que saiu ÷ peso da peça).
        chapas_baixadas = (kg_saida / peso) if peso else None
        chapas_estoque = max(und - chapas_baixadas, 0.0) if chapas_baixadas is not None else None
        chapas_reservadas = (kg_res / peso) if peso else None
        chapas_disp = max(chapas_estoque - chapas_reservadas, 0.0) if (chapas_estoque is not None and chapas_reservadas is not None) else None
        # Histórico = o código não tem mais saldo real no GRV (tudo consumido).
        # Sem dado do GRV (bridge fora), mantém em estoque (nunca falso-histórico).
        historico = bool(real_cod is not None and float(real_cod.get("qtde_total") or 0) <= 0.0001)
        out = {
            "item_id": l["item_id"], "numero_nota": l["numero_nota"], "codigo": l["codigo"],
            "descricao": l["descricao"], "fornecedor": l["fornecedor"], "ar": l["ar"],
            "lote": l["lote"], "conferente": l["conferente"], "unidade_nf": l["unidade_nf"],
            "kg_nf": kg_nf, "und": und,
            "kg_saida": round(kg_saida, 2), "kg_saldo": round(kg_saldo, 2),
            "kg_reservado": round(kg_res, 2), "kg_disponivel": round(kg_disp, 2),
            "chapas_em_estoque": chapas_estoque,
            "chapas_baixadas": chapas_baixadas,
            "chapas_reservadas": chapas_reservadas,
            "chapas_disponiveis": chapas_disp,
            "peso_calculado_total": peso_calc_total, "pct_diferenca": pct_diff,
            "historico": historico,
            "material": l.get("material", ""), "formato": l.get("formato", ""),
            "dimensoes": l.get("dimensoes"), "peso_por_peca": l.get("peso_por_peca"),
            "calculo_id": l.get("calculo_id"), "atualizado_por": l.get("atualizado_por", ""),
            "atualizado_em": l.get("atualizado_em"),
        }
        if termo:
            alvo = _normalizar_busca(" ".join(str(v) for v in [
                out["codigo"], out["descricao"], out["lote"],
                out["numero_nota"], out["ar"], out["fornecedor"], out["conferente"],
            ]))
            if termo not in alvo:
                continue
        itens_out.append(out)

    # Saldo real e reservas por OS agrupados por código (pro cabeçalho do item).
    saldo_codigo_out = {}
    reservas_os_out = {}
    for l in itens_out:
        cn = _codnorm(l["codigo"])
        r = saldo_cod.get(cn)
        if r is not None and l["codigo"] not in saldo_codigo_out:
            saldo_codigo_out[l["codigo"]] = {
                "em_estoque_kg": round(float(r.get("qtde_total") or 0), 2),
                "reservado_kg": round(float(r.get("qtde_reservada") or 0), 2),
                "disponivel_kg": round(float(r.get("qtde_disponivel") or 0), 2),
                "unidade": r.get("unidade") or "KG",
                "terceiro": bool(r.get("terceiro")),
                "cliente": r.get("cliente") or "",
            }
        ros = reservas_os_map.get(cn)
        if ros and l["codigo"] not in reservas_os_out:
            reservas_os_out[l["codigo"]] = ros

    ativos = [l for l in itens_out if not l["historico"]]
    codigos_ativos = {l["codigo"] for l in ativos}
    if saldo_codigo_out:
        total_estoque_kg = round(sum(saldo_codigo_out.get(c, {}).get("em_estoque_kg", 0.0) for c in codigos_ativos), 2)
        total_reservado_kg = round(sum(saldo_codigo_out.get(c, {}).get("reservado_kg", 0.0) for c in codigos_ativos), 2)
        total_disponivel_kg = round(sum(saldo_codigo_out.get(c, {}).get("disponivel_kg", 0.0) for c in codigos_ativos), 2)
    else:
        total_estoque_kg = round(sum(l["kg_saldo"] for l in ativos), 2)
        total_reservado_kg = round(sum(l["kg_reservado"] for l in ativos), 2)
        total_disponivel_kg = round(sum(l["kg_disponivel"] for l in ativos), 2)
    resumo = {
        "chapas": len(ativos),
        "codigos": len(codigos_ativos),
        "total_kg": total_estoque_kg,
        "reservado_kg": total_reservado_kg,
        "disponivel_kg": total_disponivel_kg,
        "com_divergencia": sum(1 for l in itens_out if l["pct_diferenca"] is not None and abs(l["pct_diferenca"]) > 2.0),
    }
    return jsonify({
        "resumo": resumo, "densidades": CHAPA_DENSIDADES,
        "itens": itens_out, "saldo_codigo": saldo_codigo_out,
        "reservas_os": reservas_os_out, "fontes": saldo_info.get("fontes") or {},
    })


@logistica_inventario_bp.route("/api/logistica/chapas/diagnostico", methods=["GET"])
@permission_required(PERMISSION)
def api_chapas_diagnostico():
    """Diagnóstico saída->lote / reservas-OS no GRV. Abrir logado, ex.:
    /api/logistica/chapas/diagnostico?saida=54087&codigo=19-01-00564"""
    saida = (request.args.get("saida") or "").strip()
    codigo = (request.args.get("codigo") or "").strip()
    from ..services.erp_estoque_service import diagnosticar_chapa_lote
    return jsonify(diagnosticar_chapa_lote(saida, codigo))


@logistica_inventario_bp.route("/api/logistica/chapas/calculo", methods=["POST"])
@permission_required(PERMISSION)
def api_chapas_salvar_calculo():
    data = request.get_json(silent=True) or {}
    try:
        item_id = int(data.get("item_id"))
    except (TypeError, ValueError):
        return jsonify({"error": "Item inválido."}), 400

    item = ItemNota.query.get(item_id)
    if not item:
        return jsonify({"error": "Item não encontrado."}), 404

    material = str(data.get("material") or "").strip()[:40]
    formato = str(data.get("formato") or "").strip()[:40]
    dimensoes = data.get("dimensoes") if isinstance(data.get("dimensoes"), dict) else {}
    try:
        peso_por_peca = float(data.get("peso_por_peca"))
    except (TypeError, ValueError):
        return jsonify({"error": "Peso por peça inválido."}), 400
    if peso_por_peca <= 0:
        return jsonify({"error": "Peso por peça deve ser maior que zero."}), 400

    usuario = session.get("username", "desconhecido")
    dimensoes_json = json.dumps(dimensoes, ensure_ascii=False)

    calc = ChapaCalculo.query.filter_by(numero_nota=item.numero_nota, item_nota_id=item.id).first()
    estado_anterior = _chapa_serializar_calculo(calc) if calc else None
    if not calc:
        calc = ChapaCalculo(
            numero_nota=item.numero_nota,
            item_nota_id=item.id,
            codigo=item.codigo_grv or item.codigo,
            criado_por=usuario,
        )
        db.session.add(calc)
    calc.material = material
    calc.formato = formato
    calc.dimensoes = dimensoes_json
    calc.peso_por_peca = peso_por_peca
    calc.atualizado_por = usuario
    calc.atualizado_em = datetime.now()
    db.session.flush()

    db.session.add(ChapaCalculoLog(
        chapa_calculo_id=calc.id,
        alterado_por=usuario,
        dados_anteriores=json.dumps(estado_anterior, ensure_ascii=False) if estado_anterior else "",
        dados_novos=json.dumps(_chapa_serializar_calculo(calc), ensure_ascii=False),
    ))
    db.session.commit()

    kg_nf = _chapa_kg_do_item(item)
    und = float(item.qtd_chapas_und or 0)
    peso_total = peso_por_peca * und if und > 0 else None
    pct = ((peso_total - kg_nf) / kg_nf * 100.0) if (peso_total is not None and kg_nf > 0) else None
    return jsonify({
        "message": "Cálculo salvo.",
        "calculo_id": calc.id,
        "peso_calculado_total": peso_total,
        "pct_diferenca": pct,
    })


@logistica_inventario_bp.route("/api/logistica/chapas/calculo/<int:calculo_id>/logs", methods=["GET"])
@permission_required(PERMISSION)
def api_chapas_calculo_logs(calculo_id):
    calc = ChapaCalculo.query.get(calculo_id)
    if not calc:
        return jsonify({"logs": []})
    logs = (
        ChapaCalculoLog.query
        .filter_by(chapa_calculo_id=calculo_id)
        .order_by(ChapaCalculoLog.alterado_em.desc(), ChapaCalculoLog.id.desc())
        .all()
    )
    def _parse(txt):
        try:
            return json.loads(txt) if txt else None
        except Exception:
            return None
    return jsonify({
        "logs": [
            {
                "alterado_por": lg.alterado_por or "",
                "alterado_em": lg.alterado_em.strftime("%d/%m/%Y %H:%M") if lg.alterado_em else "",
                "antes": _parse(lg.dados_anteriores),
                "depois": _parse(lg.dados_novos),
            }
            for lg in logs
        ]
    })
